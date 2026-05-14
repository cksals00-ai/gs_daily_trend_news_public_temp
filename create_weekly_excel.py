import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

wb = openpyxl.Workbook()

# Common styles
title_font = Font(name='Arial', size=14, bold=True)
subtitle_font = Font(name='Arial', size=10, italic=True, color='666666')
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
data_font = Font(name='Arial', size=11)
bold_font = Font(name='Arial', size=11, bold=True)
indent_font = Font(name='Arial', size=10)
sub_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
total_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
pct_fmt = '0.0%'
num_fmt = '#,##0'


def create_sheet(ws, title_text, subtitle_text, data_rows, sub_rows, start_data_row=4):
    # Title
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = title_text
    c.font = title_font
    c.alignment = Alignment(horizontal='center')

    # Subtitle
    ws.merge_cells('A2:F2')
    c = ws['A2']
    c.value = subtitle_text
    c.font = subtitle_font
    c.alignment = Alignment(horizontal='center')

    # Headers
    headers = ['세그먼트', '목표(Budget)', 'OTB', '달성률(%)', '전년동기(LY)', 'YoY(%)']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_data_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    row = start_data_row + 1
    main_rows_indices = []  # track main segment rows for summing

    for seg_name, budget, otb, ly in data_rows:
        # Column A: segment name
        cell_a = ws.cell(row=row, column=1, value=seg_name)

        if seg_name == '합계':
            # Use SUM formulas referencing the 3 main segment rows
            r1, r2, r3 = main_rows_indices
            ws.cell(row=row, column=2).value = f'=B{r1}+B{r2}+B{r3}'
            ws.cell(row=row, column=3).value = f'=C{r1}+C{r2}+C{r3}'
            ws.cell(row=row, column=5).value = f'=E{r1}+E{r2}+E{r3}'
        else:
            ws.cell(row=row, column=2, value=budget)
            ws.cell(row=row, column=3, value=otb)
            ws.cell(row=row, column=5, value=ly)
            main_rows_indices.append(row)

        ws.cell(row=row, column=2).number_format = num_fmt
        ws.cell(row=row, column=3).number_format = num_fmt
        ws.cell(row=row, column=5).number_format = num_fmt

        # Formula: 달성률 = OTB / Budget
        ws.cell(row=row, column=4).value = f'=IF(B{row}=0,"",C{row}/B{row})'
        ws.cell(row=row, column=4).number_format = pct_fmt

        # Formula: YoY = (OTB - LY) / LY
        ws.cell(row=row, column=6).value = f'=IF(E{row}=0,"",(C{row}-E{row})/E{row})'
        ws.cell(row=row, column=6).number_format = pct_fmt

        # Styling
        is_total = (seg_name == '합계')
        for c in range(1, 7):
            cell = ws.cell(row=row, column=c)
            cell.border = thin_border
            cell.font = bold_font if is_total else data_font
            if c >= 2:
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
            if is_total:
                cell.fill = total_fill

        row += 1

        # Insert sub-rows after Inbound
        if seg_name == 'Inbound':
            for country, c_budget in sub_rows:
                ws.cell(row=row, column=1, value=f'  {country}').font = indent_font
                ws.cell(row=row, column=2, value=c_budget).font = indent_font
                ws.cell(row=row, column=2).number_format = num_fmt
                for ci in [3, 4, 5, 6]:
                    ws.cell(row=row, column=ci, value='-').font = indent_font
                    ws.cell(row=row, column=ci).alignment = Alignment(horizontal='center')
                for c in range(1, 7):
                    ws.cell(row=row, column=c).fill = sub_fill
                    ws.cell(row=row, column=c).border = thin_border
                    ws.cell(row=row, column=c).alignment = ws.cell(row=row, column=c).alignment.copy(vertical='center')
                row += 1

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 12


# ===== Sheet 1: 당월(5월) =====
ws1 = wb.active
ws1.title = '당월(5월)실적'

data1 = [
    ('OTA', 33580, 31164, 21030),
    ('G-OTA', 23215, 20397, 14396),
    ('Inbound', 4410, 4473, 3487),
    ('합계', 61205, 56034, 38913),
]
sub1 = [
    ('홍콩', 723),
    ('대만', 2791),
    ('중국', 0),
    ('베트남', 118),
    ('기타', 778),
]
create_sheet(ws1,
    '주간회의 보고자료 - 당월(2026년 5월) 세그먼트별 실적 현황',
    '기준일: 2026-05-13 | 데이터: 온북(27+43) OTB',
    data1, sub1)

# ===== Sheet 2: 익월(6월) =====
ws2 = wb.create_sheet('익월(6월)OTB')

data2 = [
    ('OTA', 41050, 14151, 13889),
    ('G-OTA', 27375, 13090, 11306),
    ('Inbound', 4915, 1897, 2033),
    ('합계', 73340, 29138, 27228),
]
sub2 = [
    ('홍콩', 579),
    ('대만', 3636),
    ('중국', 0),
    ('베트남', 132),
    ('기타', 568),
]
create_sheet(ws2,
    '주간회의 보고자료 - 익월(2026년 6월) OTB 현황',
    '기준일: 2026-05-13 | 데이터: 온북(27+43) OTB',
    data2, sub2)

# Save
script_dir = os.path.dirname(os.path.abspath(__file__))
output_path = os.path.join(script_dir, '주간회의_당월익월_현황_20260514.xlsx')
wb.save(output_path)
print(f'Successfully created: {output_path}')
