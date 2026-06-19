#!/usr/bin/env python3
"""
긴급 추출: 2026년 1~5월 델피노 · OTA/G-OTA(27번 온북 예약자료) → xlsx
- 소스(투숙월 분리, 중복 없음):
    · 1~4월(202601-202604): 재전송(20260101-20260430) 파일
    · 5월(202605):          스냅샷 20260619 파일 (재전송은 5월 미커버, 스냅샷이 마감 후 보유)
  ※ 라이브 스냅샷은 마감월(1~4월)을 드롭하므로 재전송으로 백필 — month-boundary 메모 참조.
- 필터/매핑은 2025 추출과 동일(parse_raw_db 표준): 델피노, 세그먼트 53/72→OTA·A4/A5→G-OTA, 투숙(판매일자) 기준.
- 매출 = 합계금액(1박객실료×객실수, 행=1박) ÷ 1.1 = 공급가액(원), 백만 환산.
- DB 시트: 원본 컬럼 + [세그먼트, _RN, _매출공급가(원), _매출(백만)]  ← 세그먼트 행 단위 명시.
"""
import sys, glob
from pathlib import Path
from collections import defaultdict

PROJECT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT / "scripts"))
from parse_raw_db import classify_segment, extract_channel, normalize_property

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F_REISSUE = glob.glob(str(PROJECT / "data/raw_db/2026/27.*재전송(20260101-20260430)*"))[0]
F_SNAP = glob.glob(str(PROJECT / "data/raw_db/2026/27.*생성시간*"))[0]
# (소스파일, 허용 투숙월 set)
SOURCES = [
    (F_REISSUE, {"202601", "202602", "202603", "202604"}),
    (F_SNAP, {"202605"}),
]
OUT = Path.home() / "Desktop" / "델피노_OTA_GOTA_2026_1-5월_온북집계.xlsx"

SEASON_ORDER = [
    "비수기 주중", "비수기 금요일", "비수기 주말",
    "여름 성수기주중", "여름 성수기금요일", "여름 성수기토요일", "여름 최성수기",
    "겨울 성수기금요일", "겨울 성수기토요일", "겨울 최성수기", "연휴",
]
def season_rank(s):
    return SEASON_ORDER.index(s) if s in SEASON_ORDER else len(SEASON_ORDER)

agg = defaultdict(lambda: {"rn": 0, "rev": 0})   # (prop, month, season, seg, channel)
db_rows = []
tot_rn = tot_rev = 0
HDR = None

for src, months_ok in SOURCES:
    with open(src, encoding="cp949") as f:
        hdr = [h.strip() for h in f.readline().rstrip("\n\r").split(";")]
        if HDR is None:
            HDR = hdr
        ix = {h: i for i, h in enumerate(hdr)}
        i_cprop, i_oprop = ix["변경사업장명"], ix["영업장명"]
        i_sell, i_code, i_agent = ix["판매일자"], ix["변경예약집계코드"], ix["AGENT명"]
        i_rooms, i_rate, i_season = ix["객실수"], ix["1박객실료"], ix["시즌명"]
        i_mem, i_usr = ix["회원명"], ix["이용자명"]
        seen = set()
        for line in f:
            ls = line.rstrip("\n\r")
            h = hash(ls)
            if h in seen:
                continue
            seen.add(h)
            p = ls.split(";")
            if len(p) < len(hdr):
                continue
            prop = normalize_property(p[i_cprop].strip() or p[i_oprop].strip())
            if "델피노" not in prop:
                continue
            sell = p[i_sell].strip()
            if len(sell) < 6 or sell[:6] not in months_ok:
                continue
            month = sell[:6]
            code = p[i_code].strip()
            seg = classify_segment(code, "", "", "")
            if seg not in ("OTA", "G-OTA"):
                continue
            mem, usr = p[i_mem].strip(), p[i_usr].strip()
            if mem and usr and mem == usr and code != "58":
                continue
            if "매출조정" in mem or "매출조정" in usr:
                continue
            channel = extract_channel(p[i_agent].strip(), code)
            season = p[i_season].strip() or "(미지정)"
            rooms = int(p[i_rooms].strip() or 0)
            rate = int(p[i_rate].strip() or 0)
            rn = rooms if rooms > 0 else 1
            rev_won = int(rate * rn / 1.1)
            agg[(prop, month, season, seg, channel)]["rn"] += rn
            agg[(prop, month, season, seg, channel)]["rev"] += rev_won
            tot_rn += rn
            tot_rev += rev_won
            # DB행: 원본 + 세그먼트(행 단위 명시) + 계산열
            db_rows.append(p + [seg, rn, rev_won, round(rev_won / 1_000_000, 4)])

print(f"필터 통과 행수: {len(db_rows):,}")
print(f"합계 RN: {tot_rn:,}  매출(공급가 원): {tot_rev:,} = {tot_rev/1_000_000:,.1f} 백만")

# ─── 엑셀 ───
wb = openpyxl.Workbook()
HDR_FILL = PatternFill("solid", fgColor="1F4E78"); SUB_FILL = PatternFill("solid", fgColor="DDEBF7")
TOT_FILL = PatternFill("solid", fgColor="FCE4D6"); HDR_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True); THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CEN = Alignment(horizontal="center", vertical="center"); RIGHT = Alignment(horizontal="right")

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CEN; cell.border = BORDER

ws = wb.active; ws.title = "집계"
cols = ["영업장", "투숙월", "시즌", "세그먼트", "거래처", "RN(객실수)", "매출(백만,공급가)"]
ws.append(cols); style_header(ws, 1, len(cols))

def sort_key(item):
    (prop, month, season, seg, ch), v = item
    return (prop, month, season_rank(season), season, 0 if seg == "OTA" else 1, -v["rev"])

ordered = sorted(agg.items(), key=sort_key)
r = 2; sub_rn = sub_rev = month_rn = month_rev = 0

def write_subtotal(ws, r, label, rn, rev, fill):
    ws.cell(row=r, column=1, value=label).font = BOLD
    ws.cell(row=r, column=6, value=rn).font = BOLD
    ws.cell(row=r, column=7, value=round(rev / 1_000_000, 1)).font = BOLD
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = fill; ws.cell(row=r, column=c).border = BORDER
    ws.cell(row=r, column=6).alignment = RIGHT; ws.cell(row=r, column=7).alignment = RIGHT
    return r + 1

prev = None
for (prop, month, season, seg, ch), v in ordered:
    if prev and (prev[1] != month or prev[2] != season):
        r = write_subtotal(ws, r, f"  └ {prev[1]} {prev[2]} 소계", sub_rn, sub_rev, SUB_FILL); sub_rn = sub_rev = 0
    if prev and prev[1] != month:
        r = write_subtotal(ws, r, f"▶ {prev[1]} 월 합계", month_rn, month_rev, TOT_FILL); month_rn = month_rev = 0
    ws.append([prop, month, season, seg, ch, v["rn"], round(v["rev"] / 1_000_000, 2)])
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = BORDER
    ws.cell(row=r, column=6).alignment = RIGHT; ws.cell(row=r, column=7).alignment = RIGHT
    sub_rn += v["rn"]; sub_rev += v["rev"]; month_rn += v["rn"]; month_rev += v["rev"]
    prev = (prop, month, season, seg, ch); r += 1
if prev:
    r = write_subtotal(ws, r, f"  └ {prev[1]} {prev[2]} 소계", sub_rn, sub_rev, SUB_FILL)
    r = write_subtotal(ws, r, f"▶ {prev[1]} 월 합계", month_rn, month_rev, TOT_FILL)
ws.cell(row=r, column=1, value="■ 총 합계 (델피노 OTA+G-OTA 2026.1~5)").font = Font(bold=True, size=12)
ws.cell(row=r, column=6, value=tot_rn).font = Font(bold=True, size=12)
ws.cell(row=r, column=7, value=round(tot_rev / 1_000_000, 1)).font = Font(bold=True, size=12)
for c in range(1, 8):
    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFC000"); ws.cell(row=r, column=c).border = BORDER
ws.freeze_panes = "A2"
for i, w in enumerate([16, 10, 16, 9, 14, 12, 16], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# DB 시트
ws2 = wb.create_sheet("DB")
db_hdr = HDR + ["세그먼트(OTA/GOTA)", "_RN", "_매출공급가(원)", "_매출(백만)"]
ws2.append(db_hdr); style_header(ws2, 1, len(db_hdr))
for row in db_rows:
    ws2.append(row)
ws2.freeze_panes = "B2"; ws2.column_dimensions["A"].width = 16
seg_col = len(HDR) + 1
ws2.column_dimensions[get_column_letter(seg_col)].width = 16
ws2.auto_filter.ref = f"A1:{get_column_letter(len(db_hdr))}1"

# 기준 시트
ws3 = wb.create_sheet("기준")
notes = [
    ["항목", "내용"],
    ["소스", "27.온라인영업팀 예약자료 (FIT, 27번만). 28(취소)·43(IB) 미사용 → 취소 미차감 그로스 온북"],
    ["소스 파일 분리", "1~4월=재전송(20260101-20260430), 5월=스냅샷 20260619 (라이브가 마감월 드롭, 재전송이 백필)"],
    ["사업장", "델피노 (08. 델피노)"],
    ["세그먼트", "변경예약집계코드 기준: 53/72→OTA, A4/A5→G-OTA (그 외 제외). DB 행에 세그먼트 컬럼 명시"],
    ["거래처", "AGENT명 → 채널 매핑 (parse_raw_db OTA_CHANNEL_MAP)"],
    ["시즌", "원본 '시즌명' 컬럼"],
    ["연도/기간", "투숙(판매일자) 2026년 1~5월 (202601~202605). 각 행=1박(연박 일자 전개)"],
    ["RN", "객실수 합 (행=1박)"],
    ["매출", "합계금액(1박객실료×객실수) ÷ 1.1 = 공급가액(원), 백만 환산"],
    ["정제", "행 중복제거 + 자체예약(회원명=이용자명) 제거 + 매출조정 제거"],
    ["", ""],
    ["검증: 집계 RN 합", tot_rn],
    ["검증: 집계 매출(백만)", round(tot_rev / 1_000_000, 1)],
    ["검증: DB 행수", len(db_rows)],
]
for row in notes:
    ws3.append(row)
style_header(ws3, 1, 2); ws3.column_dimensions["A"].width = 20; ws3.column_dimensions["B"].width = 85
for r_ in range(2, len(notes) + 1):
    ws3.cell(row=r_, column=2).alignment = Alignment(wrap_text=True, vertical="top")

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"\n저장: {OUT}")

db_rn = sum(row[-3] for row in db_rows); db_rev = sum(row[-2] for row in db_rows)
print("\n=== 교차검증 (집계 vs DB) ===")
print(f"RN  집계={tot_rn:,}  DB={db_rn:,}  일치={tot_rn==db_rn}")
print(f"매출원 집계={tot_rev:,}  DB={db_rev:,}  일치={tot_rev==db_rev}")
