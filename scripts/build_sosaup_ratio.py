#!/usr/bin/env python3
"""
build_sosaup_ratio.py
=====================

소사업(객실 외) 매출 추정용 **전년 실당소사업 비율**을 소스 엑셀에서 1회 추출 →
``data/sosaup_ratio.json`` 으로 저장한다.

방법(가): 비율(천원/실)만 저장해 두고, 실제 소사업 매출은
``build_rm_fcst_excel.py`` 가 ``현재 RM FCST 객실수 × 비율`` 로 매주 재계산한다.

산식 / 매핑 / fallback (사용자 확정):
  소사업매출(VAT제외) = RM FCST 객실수 × (전년 동월 GS 소사업매출 ÷ 전년 동월 GS 객실수)
  - 소사업 = GS 영업조직, 객실 외 7유형(식음+부대+아쿠아+스포츠+골프+유통+기타)
  - 세그먼트: 인바운드=Inbound, 국내OTA=OTA, 해외OTA=G-OTA
  - 비발디: 소노캄 비발디파크 → 01·02·03,  소노펠리체 비발디파크 → 04·05,
            소노펠리체CC·스키&오션월드는 제외(rm_fcst에 객실 없음)
  - 고양 = 소노캄 고양 + 소노펠리체 컨벤션 (동일 입지, rm_fcst는 20.소노캄고양 하나)
  - fallback: 전년 동월 소사업매출=0 또는 객실수=0(작년 미운영/미입력)이면
              해당 사업장·세그먼트를 **2026-05(전체시트)** 비율로 대체

데이터 출처:
  - 전년 소사업매출: ``Ⅲ-3. 유형별매출(월별)`` (2025년, 이미 VAT제외=공급가액)
  - 전년 객실수:     ``Ⅲ-1. 판매객실수(월별)`` 섹션3(2025년 실적)
  - fallback:        ``전체`` 시트(2026-05, 계=VAT포함 → ÷1.1 로 VAT제외 환산)

사용:
  python scripts/build_sosaup_ratio.py [소스엑셀경로]
  (인자 생략 시 SRC_DEFAULT 사용. 새 소사업 엑셀이 오면 경로를 넘겨 재추출)

단위: 비율 = 천원/실 (VAT제외)
"""

import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "data" / "sosaup_ratio.json"
RM_FCST = REPO / "data" / "rm_fcst.json"

# 소스 엑셀 기본 경로 — 전년도(고정) 비율 원천. 레포에 영구 저장(data/sources/, gitignore).
# 전년도 기준이므로 재추출은 전년 데이터 정정 시에만 필요. 새 파일은 CLI 인자로 교체.
SRC_DEFAULT = str(REPO / "data" / "sources" / "소사업_전년비율_소스_2025.xlsx")

SEGS = ["OTA", "G-OTA", "Inbound"]
CHAN = {"인바운드": "Inbound", "국내OTA": "OTA", "해외OTA": "G-OTA"}
NONROOM = {"식음", "부대", "아쿠아", "스포츠", "골프", "유통", "기타"}
# 전체시트(2026-05) 행사구분 → 세그먼트
EVENT_SEG = {
    "외국인(동남아)": "Inbound", "외국인(일본)": "Inbound",
    "대매점": "OTA", "패키지(대매점)": "OTA",
    "해외OTA": "G-OTA", "패키지(해외OTA)": "G-OTA",
}

# rm_fcst 사업장 → (Ⅲ-3 유형별 블록, Ⅲ-1 객실수 블록, 전체시트 fallback 사업장명)
PROP_MAP = {
    "01.벨비발디":     (["소노캄 비발디파크"],   ["소노캄 비발디파크"],   ["비발디파크"]),
    "02.캄비발디":     (["소노캄 비발디파크"],   ["소노캄 비발디파크"],   ["비발디파크"]),
    "03.펫비발디":     (["소노캄 비발디파크"],   ["소노캄 비발디파크"],   ["비발디파크"]),
    "04.펠리체비발디":  (["소노펠리체 비발디파크"], ["소노펠리체 비발디파크"], ["비발디파크"]),
    "05.빌리지비발디":  (["소노펠리체 비발디파크"], ["소노펠리체 비발디파크"], ["비발디파크"]),
    "06.양평":        (["소노휴 양평"],        ["소노벨 양평"],        ["소노벨 양평"]),
    "07.델피노":       (["델피노"],            ["델피노"],            ["델피노"]),
    "08.쏠비치양양":    (["쏠비치 양양"],        ["쏠비치 양양"],        ["쏠비치 양양"]),
    "09.쏠비치삼척":    (["쏠비치 삼척"],        ["쏠비치 삼척"],        ["쏠비치 삼척"]),
    "10.소노벨단양":    (["소노벨 단양"],        ["소노벨 단양"],        ["소노벨 단양"]),
    "11.소노캄경주":    (["소노캄 경주"],        ["소노캄 경주"],        ["소노캄 경주"]),
    "12.소노벨청송":    (["소노벨 청송"],        ["소노벨 청송"],        ["소노벨 청송"]),
    "13.소노벨천안":    (["소노벨 천안"],        ["소노벨 천안"],        ["소노벨 천안"]),
    "14.소노벨변산":    (["소노벨 변산"],        ["소노벨 변산"],        ["소노벨 변산"]),
    "15.소노캄여수":    (["소노캄 여수"],        ["소노캄 여수"],        ["소노캄 여수"]),
    "16.소노캄거제":    (["소노캄 거제"],        ["소노캄 거제"],        ["소노캄 거제"]),
    "17.쏠비치진도":    (["쏠비치 진도"],        ["쏠비치 진도"],        ["쏠비치 진도"]),
    "18.소노벨제주":    (["소노벨 제주"],        ["소노벨 제주"],        ["소노벨 제주"]),
    "19.소노캄제주":    (["소노캄 제주"],        ["소노캄 제주"],        ["소노캄 제주"]),
    "20.소노캄고양":    (["소노캄 고양", "소노펠리체 컨벤션"],
                       ["소노캄 고양", "소노펠리체 컨벤션"], ["소노캄 고양"]),
    "21.소노문해운대":  (["소노문 해운대"],      ["소노문 해운대"],      ["소노문 해운대"]),
    "22.쏠비치남해":    (["쏠비치 남해"],        ["쏠비치 남해"],        ["쏠비치 남해"]),
    "23.르네블루":     ([],                  ["르네블루 바이 쏠비치"], ["르네블루"]),
}

# 비발디 클러스터(01·02·03 / 04·05)는 각 사업장이 같은 블록을 가리켜 동일 비율로 산출됨.

FALLBACK_MONTH = "2026-05"  # 전체시트 기준월


def _nz(v):
    return re.sub(r"\s", "", v) if isinstance(v, str) else ""


def _block_index(ws):
    """월별 시트의 사업장 블록: 정제된 사업장명 → (start_row, end_row)."""
    found = []
    r = 1
    while r <= ws.max_row:
        b = ws.cell(r, 2).value
        if isinstance(b, str) and b.startswith("Ⅲ"):
            nm = re.sub(r"^[0-9\-]+\)\s*", "", (ws.cell(r + 3, 2).value or "").replace("※", "")).strip()
            found.append((r, nm))
        r += 1
    idx = {}
    for i, (s, nm) in enumerate(found):
        end = found[i + 1][0] if i + 1 < len(found) else ws.max_row + 1
        idx[nm] = (s, end)
    return idx


def _gs_sosaup(ws, span, col):
    """Ⅲ-3 유형별매출: GS 조직, 세그먼트별 객실 외 7유형 합(이미 VAT제외)."""
    s, e = span
    res = {k: 0 for k in SEGS}
    in_gs = False
    cur = None
    for r in range(s, e):
        B, C, D = _nz(ws.cell(r, 2).value), _nz(ws.cell(r, 3).value), _nz(ws.cell(r, 4).value)
        if B == "GS":
            in_gs = True
            cur = None
            continue
        if in_gs and B in ("마케팅", "MICE", "레저마케팅"):
            break
        if in_gs and C in CHAN:
            cur = CHAN[C]
            continue
        if in_gs and C == "년간계약사":
            cur = None
            continue
        if in_gs and cur and D in NONROOM:
            v = ws.cell(r, col).value or 0
            if isinstance(v, (int, float)):
                res[cur] += v
    return res


def _gs_rooms(ws, span, col):
    """Ⅲ-1 판매객실수: GS 조직, 세그먼트(채널 헤더 행) 객실수 합."""
    s, e = span
    res = {k: 0 for k in SEGS}
    in_gs = False
    for r in range(s, e):
        B, C = _nz(ws.cell(r, 2).value), _nz(ws.cell(r, 3).value)
        if B == "GS":
            in_gs = True
            continue
        if in_gs and B in ("마케팅", "MICE", "레저마케팅"):
            break
        if in_gs and C in CHAN:
            v = ws.cell(r, col).value or 0
            if isinstance(v, (int, float)):
                res[CHAN[C]] += v
    return res


def _sum_blocks(fn, idx, names, col):
    res = {k: 0 for k in SEGS}
    for n in names:
        if n in idx:
            sub = fn(idx[n], col)  # idx[n] = (start_row, end_row)
            for k in SEGS:
                res[k] += sub[k]
    return res


def _fallback_total(ws_total, names):
    """전체시트(2026-05) VAT제외 소사업/객실수 = (계-객실료)/1.1, 객실수."""
    soup = {k: 0 for k in SEGS}
    room = {k: 0 for k in SEGS}
    for r in range(2, ws_total.max_row + 1):
        if (ws_total.cell(r, 2).value or "") not in names:
            continue
        h = ws_total.cell(r, 3).value
        if h in EVENT_SEG:
            seg = EVENT_SEG[h]
            gye = ws_total.cell(r, 14).value or 0       # 계 (VAT포함)
            rfee = ws_total.cell(r, 5).value or 0       # 객실료 (VAT포함)
            soup[seg] += (gye - rfee) / 1.1             # VAT제외 환산
            room[seg] += ws_total.cell(r, 4).value or 0
    return soup, room


def compute_month(rev_ws, rn_ws, tot_ws, rev_idx, rn_idx, ref_ym):
    """ref_ym(전년, 예: '2025-06') 기준 사업장×세그먼트 비율(천원/실) 산출."""
    ref_m = int(ref_ym.split("-")[1])
    rev_col = 6 + ref_m   # Ⅲ-3 2025 실적: m1=col7 ... m12=col18
    rn_col = 42 + ref_m   # Ⅲ-1 섹션3(2025): m1=col43 ... m12=col54

    ratios = {}
    fallback_used = {}
    computed = {}  # 대표키 비율 캐시(클러스터 공유용)

    for prop, (rev_names, rn_names, fb_names) in PROP_MAP.items():
        rr = _sum_blocks(_gs_sosaup_ws(rev_ws), rev_idx, rev_names, rev_col)
        nr = _sum_blocks(_gs_rooms_ws(rn_ws), rn_idx, rn_names, rn_col)
        fb_soup = fb_room = None
        pr = {}
        used = []
        for seg in SEGS:
            if rr[seg] == 0 or nr[seg] == 0:           # 작년 미운영/미입력 → fallback
                if fb_soup is None:
                    fb_soup, fb_room = _fallback_total(tot_ws, fb_names)
                ratio = (fb_soup[seg] / fb_room[seg]) if fb_room[seg] else 0
                if ratio > 0:
                    used.append(seg)
            else:
                ratio = rr[seg] / nr[seg]
            pr[seg] = round(ratio, 4)
        ratios[prop] = pr
        if used:
            fallback_used[prop] = used
    return ratios, fallback_used


# openpyxl span 헬퍼: _gs_sosaup/_gs_rooms 를 (ws 고정) 부분적용
def _gs_sosaup_ws(ws):
    return lambda span, col: _gs_sosaup(ws, span, col)


def _gs_rooms_ws(ws):
    return lambda span, col: _gs_rooms(ws, span, col)


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(SRC_DEFAULT)
    if not src.exists():
        sys.exit(f"✗ 소스 엑셀 없음: {src}\n  새 소사업 엑셀 경로를 인자로 넘기세요.")

    wb = openpyxl.load_workbook(src, data_only=True)
    rev_ws = wb["Ⅲ-3. 유형별매출(월별)"]
    rn_ws = wb["Ⅲ-1. 판매객실수(월별)"]
    tot_ws = wb["전체"]
    rev_idx = _block_index(rev_ws)
    rn_idx = _block_index(rn_ws)

    # 전년 비율은 전년도(고정) 기준이므로 **연중 12개월 전체**를 산출한다.
    # → 마감월(온북)·미마감월(FCST) 어느 달이든 소사업 누적 계산이 가능.
    rm = json.loads(RM_FCST.read_text(encoding="utf-8"))
    # 대상 연도 = rm_fcst 커버 월의 연도(예: 2026). 없으면 현재 연도.
    base_year = None
    for arr in rm.get("_months_covered", []):
        s = f"{int(arr[0])}-{int(arr[1]):02d}" if isinstance(arr, (list, tuple)) else str(arr)
        base_year = int(s.split("-")[0])
        break
    if base_year is None:
        base_year = datetime.now().year
    covered = [f"{base_year}-{m:02d}" for m in range(1, 13)]

    by_month = {}
    for fym in covered:
        y, m = fym.split("-")
        ref_ym = f"{int(y) - 1}-{m}"
        ratios, fb = compute_month(rev_ws, rn_ws, tot_ws, rev_idx, rn_idx, ref_ym)
        by_month[fym] = {
            "ref_year_month": ref_ym,
            "fallback_month": FALLBACK_MONTH,
            "ratios": ratios,
            "fallback_used": fb,
        }

    out = {
        "_generated_at": datetime.now().isoformat() + "Z",
        "_source_excel": src.name,
        "_unit": "천원/실 (VAT제외, 공급가액 기준)",
        "_definition": "GS 영업조직, 객실 외 7유형(식음+부대+아쿠아+스포츠+골프+유통+기타)",
        "_segment_map": {"인바운드": "Inbound", "국내OTA": "OTA", "해외OTA": "G-OTA"},
        "_method": (
            "소사업매출 = RM FCST 객실수 × 비율. 전년 동월 기준(Ⅲ-3 유형별매출/Ⅲ-1 판매객실수). "
            "비발디: 소노캄→01·02·03, 소노펠리체→04·05, CC·스키오션 제외. 고양=고양+컨벤션. "
            "fallback: 전년 동월 소사업=0 또는 객실수=0이면 2026-05(전체시트, VAT제외) 비율 사용."
        ),
        "by_month": by_month,
    }
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✓ sosaup_ratio.json 생성 → {OUT}")
    for fym, blk in by_month.items():
        nfb = len(blk["fallback_used"])
        print(f"  {fym} (전년 {blk['ref_year_month']}): 사업장 {len(blk['ratios'])}개, fallback {nfb}개")


if __name__ == "__main__":
    main()
