#!/usr/bin/env python3
"""
긴급 추출: 2025 델피노 · OTA/G-OTA(27번 온북 예약자료) → xlsx
- 소스: data/raw_db/2025/27.온라인영업팀... 예약자료 (FIT 예약, 27번만). 28(취소)·43(IB) 미사용.
- 필터: 사업장 '델피노', 세그먼트 OTA/G-OTA, 투숙(판매일자) 2025.
- 세그먼트/거래처/사업장/시즌 매핑은 parse_raw_db 표준 로직 그대로 재사용.
- 매출 = 합계금액(=1박객실료×객실수, 각 행=1박) ÷ 1.1 → 공급가액(원), 백만원 환산.
- 출력: 시트1 집계(영업장×월×시즌×세그먼트×거래처, RN·매출백만, 소계·합계),
        시트2 DB(필터 통과 원본 컬럼 + 계산 RN·매출), 시트3 기준.
"""
import sys, glob
from pathlib import Path
from collections import defaultdict, OrderedDict

PROJECT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT / "scripts"))
from parse_raw_db import classify_segment, extract_channel, normalize_property  # 표준 로직 재사용

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = glob.glob(str(PROJECT / "data/raw_db/2025/27.*"))[0]
OUT = Path.home() / "Desktop" / "델피노_OTA_GOTA_2025_온북집계.xlsx"

# ─── 시즌 정렬 순서 (시즌명 기준) ───
SEASON_ORDER = [
    "비수기 주중", "비수기 금요일", "비수기 주말",
    "여름 성수기주중", "여름 성수기금요일", "여름 성수기토요일", "여름 최성수기",
    "겨울 성수기금요일", "겨울 성수기토요일", "겨울 최성수기",
    "연휴",
]
def season_rank(s):
    return SEASON_ORDER.index(s) if s in SEASON_ORDER else len(SEASON_ORDER)

# ─── 파싱 ───
with open(SRC, encoding="cp949") as f:
    HDR = [h.strip() for h in f.readline().rstrip("\n\r").split(";")]
ix = {h: i for i, h in enumerate(HDR)}

i_cprop, i_oprop = ix["변경사업장명"], ix["영업장명"]
i_sell = ix["판매일자"]
i_code = ix["변경예약집계코드"]
i_agent = ix["AGENT명"]
i_rooms, i_rate = ix["객실수"], ix["1박객실료"]
i_season = ix["시즌명"]
i_mem, i_usr = ix["회원명"], ix["이용자명"]

agg = defaultdict(lambda: {"rn": 0, "rev": 0})   # key=(prop, month, season, seg, channel)
db_rows = []                                      # 필터 통과 원본행 + 계산열
tot_rn = 0
tot_rev = 0

with open(SRC, encoding="cp949") as f:
    f.readline()  # header
    seen = set()
    for line in f:
        ls = line.rstrip("\n\r")
        h = hash(ls)
        if h in seen:        # 행 단위 중복 제거 (parse_raw_db 동일)
            continue
        seen.add(h)
        p = ls.split(";")
        if len(p) < len(HDR):
            continue
        # 사업장
        prop = normalize_property(p[i_cprop].strip() or p[i_oprop].strip())
        if "델피노" not in prop:
            continue
        # 투숙(판매)일 2025
        sell = p[i_sell].strip()
        if not (len(sell) >= 6 and "202501" <= sell[:6] <= "202512"):
            continue
        month = sell[:6]
        # 세그먼트 = 코드 기반 (표준) → OTA/G-OTA만
        code = p[i_code].strip()
        seg = classify_segment(code, "", "", "")
        if seg not in ("OTA", "G-OTA"):
            continue
        # 거래처 자체예약(회원명=이용자명)·매출조정 제거 (parse_raw_db 동일)
        mem, usr = p[i_mem].strip(), p[i_usr].strip()
        if mem and usr and mem == usr and code != "58":
            continue
        if "매출조정" in mem or "매출조정" in usr:
            continue
        # 거래처 = AGENT명 → 채널 (표준)
        channel = extract_channel(p[i_agent].strip(), code)
        season = p[i_season].strip() or "(미지정)"
        # RN / 매출(공급가, 원) = 1박객실료×객실수 ÷ 1.1, 행별 int 절사 (BI 동일)
        rooms = int(p[i_rooms].strip() or 0)
        rate = int(p[i_rate].strip() or 0)
        rn = rooms if rooms > 0 else 1
        rev_won = int(rate * rn / 1.1)

        k = (prop, month, season, seg, channel)
        agg[k]["rn"] += rn
        agg[k]["rev"] += rev_won
        tot_rn += rn
        tot_rev += rev_won
        db_rows.append(p + [rn, rev_won, round(rev_won / 1_000_000, 4)])

print(f"필터 통과 행수: {len(db_rows):,}")
print(f"합계 RN: {tot_rn:,}  매출(공급가 원): {tot_rev:,}  = {tot_rev/1_000_000:,.1f} 백만")

# ─── 엑셀 작성 ───
wb = openpyxl.Workbook()

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="DDEBF7")
TOT_FILL = PatternFill("solid", fgColor="FCE4D6")
HDR_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CEN = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right")

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CEN; cell.border = BORDER

# ── 시트1: 집계 ──
ws = wb.active
ws.title = "집계"
cols = ["영업장", "투숙월", "시즌", "세그먼트", "거래처", "RN(객실수)", "매출(백만,공급가)"]
ws.append(cols)
style_header(ws, 1, len(cols))

# 정렬: 영업장→월→시즌→세그먼트→매출 desc
def sort_key(item):
    (prop, month, season, seg, ch), v = item
    return (prop, month, season_rank(season), season, 0 if seg == "OTA" else 1, -v["rev"])

ordered = sorted(agg.items(), key=sort_key)

r = 2
cur_month = cur_season = None
sub_rn = sub_rev = 0
month_rn = month_rev = 0

def write_subtotal(ws, r, label, rn, rev, fill):
    ws.cell(row=r, column=1, value=label).font = BOLD
    ws.cell(row=r, column=6, value=rn).font = BOLD
    ws.cell(row=r, column=7, value=round(rev / 1_000_000, 1)).font = BOLD
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=c).border = BORDER
    ws.cell(row=r, column=6).alignment = RIGHT
    ws.cell(row=r, column=7).alignment = RIGHT
    return r + 1

prev = None
for (prop, month, season, seg, ch), v in ordered:
    # 시즌 소계
    if prev and (prev[1] != month or prev[2] != season):
        r = write_subtotal(ws, r, f"  └ {prev[1]} {prev[2]} 소계", sub_rn, sub_rev, SUB_FILL)
        sub_rn = sub_rev = 0
    # 월 소계
    if prev and prev[1] != month:
        r = write_subtotal(ws, r, f"▶ {prev[1]} 월 합계", month_rn, month_rev, TOT_FILL)
        month_rn = month_rev = 0
    row = [prop, month, season, seg, ch, v["rn"], round(v["rev"] / 1_000_000, 2)]
    ws.append(row)
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = BORDER
    ws.cell(row=r, column=6).alignment = RIGHT
    ws.cell(row=r, column=7).alignment = RIGHT
    sub_rn += v["rn"]; sub_rev += v["rev"]
    month_rn += v["rn"]; month_rev += v["rev"]
    prev = (prop, month, season, seg, ch)
    r += 1
# 마지막 소계/월계
if prev:
    r = write_subtotal(ws, r, f"  └ {prev[1]} {prev[2]} 소계", sub_rn, sub_rev, SUB_FILL)
    r = write_subtotal(ws, r, f"▶ {prev[1]} 월 합계", month_rn, month_rev, TOT_FILL)
# 총합계
ws.cell(row=r, column=1, value="■ 총 합계 (델피노 OTA+G-OTA 2025)").font = Font(bold=True, size=12)
ws.cell(row=r, column=6, value=tot_rn).font = Font(bold=True, size=12)
ws.cell(row=r, column=7, value=round(tot_rev / 1_000_000, 1)).font = Font(bold=True, size=12)
for c in range(1, 8):
    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFC000")
    ws.cell(row=r, column=c).border = BORDER
ws.freeze_panes = "A2"
widths = [16, 10, 16, 9, 14, 12, 16]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── 시트2: DB (원본 컬럼 + 계산열) ──
ws2 = wb.create_sheet("DB")
db_hdr = HDR + ["_RN", "_매출공급가(원)", "_매출(백만)"]
ws2.append(db_hdr)
style_header(ws2, 1, len(db_hdr))
for row in db_rows:
    ws2.append(row)
ws2.freeze_panes = "B2"
ws2.column_dimensions["A"].width = 16
ws2.auto_filter.ref = f"A1:{get_column_letter(len(db_hdr))}1"

# ── 시트3: 기준/검증 ──
ws3 = wb.create_sheet("기준")
notes = [
    ["항목", "내용"],
    ["소스", "27.온라인영업팀(대매점,OTA,패키지) 예약자료 (FIT 예약) — 27번만 사용"],
    ["미사용", "28(취소)·43/44(Inbound) 미반영 → 취소 미차감(그로스 온북)"],
    ["사업장 필터", "변경사업장명/영업장명에 '델피노' 포함 (08. 델피노)"],
    ["세그먼트", "변경예약집계코드 기준 (표준): 53/72→OTA, A4/A5→G-OTA. 그 외 제외"],
    ["거래처", "AGENT명 → 채널 매핑 (parse_raw_db OTA_CHANNEL_MAP)"],
    ["시즌", "원본 '시즌명' 컬럼 그대로 사용"],
    ["연도 기준", "투숙(판매일자) 2025년 (202501~202512). 각 행=1박(연박은 일자별 전개)"],
    ["RN", "객실수 합 (각 행 1박 단위)"],
    ["매출", "매출 = 합계금액(1박객실료×객실수) ÷ 1.1 = 공급가액(원), 백만원 환산"],
    ["정제", "행 중복제거 + 자체예약(회원명=이용자명) 제거 + 매출조정 제거 (parse_raw_db 동일)"],
    ["", ""],
    ["검증: 집계 RN 합", tot_rn],
    ["검증: 집계 매출(백만)", round(tot_rev / 1_000_000, 1)],
    ["검증: DB 행수", len(db_rows)],
]
for row in notes:
    ws3.append(row)
style_header(ws3, 1, 2)
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 80
for r_ in range(2, len(notes) + 1):
    ws3.cell(row=r_, column=2).alignment = Alignment(wrap_text=True, vertical="top")

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"\n저장: {OUT}")

# ─── 교차 검증 ───
db_rn = sum(row[-3] for row in db_rows)
db_rev = sum(row[-2] for row in db_rows)
print("\n=== 교차검증 (집계 vs DB) ===")
print(f"RN  집계={tot_rn:,}  DB={db_rn:,}  일치={tot_rn==db_rn}")
print(f"매출원 집계={tot_rev:,}  DB={db_rev:,}  일치={tot_rev==db_rev}")
