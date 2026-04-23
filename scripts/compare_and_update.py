#!/usr/bin/env python3
"""
compare_and_update.py
===========================================================
작업 1: DB 파싱 결과 vs 사업계획 Budget 비교 (콘솔 출력)
작업 2: daily_notes.json property_performance를 DB + Budget 데이터로 교체
===========================================================
- DB 데이터: data/db_aggregated.json (parse_raw_db.py 실행 결과)
- Budget: data/raw_db/budget/★최종★... xlsx
- 출력: data/daily_notes.json (property_performance 키만 교체)
"""
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"
BUDGET_XLSX = DATA_DIR / "raw_db" / "budget" / "★최종★(검토완료)_2026년 객실 사업계획_총량 수립(2차+사업장변경건).xlsx"
DB_JSON = DATA_DIR / "db_aggregated.json"
NOTES_JSON = DATA_DIR / "daily_notes.json"

# ── 투숙월 (HTML에서 표시할 3개월) ──
STAY_MONTHS = ["202604", "202605", "202606"]  # db_aggregated.json key format
STAY_MONTHS_DISP = ["2026-04", "2026-05", "2026-06"]  # daily_notes.json key format

# ── Budget xlsx Grand Total 행 (1-indexed) ──
BUDGET_GRAND_TOTAL_ROWS = [26, 50, 74, 98, 122, 146, 170, 194, 218, 242, 266, 290]
BUDGET_MONTHS_LABEL = ["1월", "2월", "3월", "4월", "5월", "6월",
                        "7월", "8월", "9월", "10월", "11월", "12월"]
# Grand Total cols: RN=col6, ADR=col8, REV(천원)=col10
BUDGET_COL_RN = 6
BUDGET_COL_ADR = 8
BUDGET_COL_REV = 10  # 千원 단위

# ── Budget 개별 시트명 → (display_name, region, db_props) ──
#   db_props: DB의 property names (by_property 키) 중 합산 대상
PROPERTY_DEFS = [
    # (sheet_name, display_name, region, [db_property_names])
    ("소노벨 비발디파크(A,B,C,호텔)",       "01.벨비발디",      "vivaldi", ["소노벨 비발디파크", "소노문 비발디파크"]),
    ("소노캄 비발디파크 (D동)",              "02.캄비발디",      "vivaldi", ["소노캄 비발디파크"]),
    ("소노펫 (D동+E동)",                     "03.펫비발디",      "vivaldi", ["소노펫 비발디파크"]),
    ("소노펠리체 비발디파크",                "04.펠리체비발디",  "vivaldi", ["소노펠리체 비발디파크"]),
    ("소노펠리체 빌리지 비발디파크",         "05.빌리지비발디",  "vivaldi", ["소노펠리체 빌리지 비발디파크"]),
    ("소노벨양평",                           "06.양평",          "central", ["소노휴 양평", "소노벨 양평"]),
    ("델피노",                               "07.델피노",        "central", ["델피노"]),
    ("쏠비치양양",                           "08.쏠비치양양",    "central", ["쏠비치 양양"]),
    ("쏠비치삼척",                           "09.쏠비치삼척",    "central", ["쏠비치 삼척"]),
    ("소노벨단양",                           "10.소노벨단양",    "central", ["소노문 단양", "소노벨 단양"]),
    ("소노캄경주",                           "11.소노캄경주",    "south",   ["소노벨 경주", "소노캄 경주"]),
    ("소노벨청송",                           "12.소노벨청송",    "central", ["소노벨 청송"]),
    ("소노벨천안",                           "13.소노벨천안",    "central", ["소노벨 천안"]),
    ("소노벨변산",                           "14.소노벨변산",    "central", ["소노벨 변산"]),
    ("소노캄여수",                           "15.소노캄여수",    "south",   ["소노캄 여수"]),
    ("소노캄 거제",                          "16.소노캄거제",    "south",   ["소노캄 거제"]),
    ("쏠비치 진도",                          "17.쏠비치진도",    "south",   ["쏠비치 진도"]),
    ("소노벨 제주",                          "18.소노벨제주",    "apac",    ["소노벨 제주"]),
    ("소노캄 제주",                          "19.소노캄제주",    "apac",    ["소노캄 제주"]),
    ("소노캄 고양",                          "20.소노캄고양",    "apac",    ["소노캄 고양"]),
    ("소노문 해운대",                        "21.소노문해운대",  "south",   ["소노문 해운대"]),
    ("쏠비치 남해",                          "22.쏠비치남해",    "south",   ["쏠비치 남해"]),
    ("르네블루 바이 쏠비치",                 "23.르네블루",      "central", ["르네블루"]),
]


def load_budget_monthly(wb, sheet_name: str) -> dict:
    """Budget 개별 시트에서 월별 Grand Total (RN, ADR, REV_千원) 추출"""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    result = {}
    for i, (row_idx, month_label) in enumerate(zip(BUDGET_GRAND_TOTAL_ROWS, BUDGET_MONTHS_LABEL)):
        rn = ws.cell(row_idx, BUDGET_COL_RN).value or 0
        adr = ws.cell(row_idx, BUDGET_COL_ADR).value or 0
        rev_cheonwon = ws.cell(row_idx, BUDGET_COL_REV).value or 0
        result[month_label] = {
            "rn": int(round(float(rn))) if rn else 0,
            "adr": round(float(adr), 1) if adr else 0.0,
            "rev_m": round(float(rev_cheonwon) / 1000, 2) if rev_cheonwon else 0.0,  # 千원 → 백만원
        }
    return result


def sum_db_props(db: dict, prop_names: list, month: str) -> dict:
    """DB에서 여러 property를 합산하여 월별 net_rn, net_rev(百万), adr 반환"""
    total_rn = 0
    total_rev = 0.0
    for pname in prop_names:
        m_data = db.get("by_property", {}).get(pname, {}).get(month, {})
        total_rn += m_data.get("net_rn", 0)
        total_rev += m_data.get("net_rev", 0.0)
    adr = round((total_rev * 1000) / total_rn) if total_rn > 0 else 0
    return {"net_rn": total_rn, "net_rev": round(total_rev, 2), "adr": adr}


def sum_db_annual(db: dict, prop_names: list, year: str) -> dict:
    """DB에서 연간(특정 year) 합산"""
    all_months = [m for m in db.get("meta", {}).get("months", []) if m.startswith(year)]
    total_rn = 0
    total_rev = 0.0
    for pname in prop_names:
        prop_data = db.get("by_property", {}).get(pname, {})
        for m in all_months:
            m_data = prop_data.get(m, {})
            total_rn += m_data.get("net_rn", 0)
            total_rev += m_data.get("net_rev", 0.0)
    adr = round((total_rev * 1000) / total_rn) if total_rn > 0 else 0
    return {"net_rn": total_rn, "net_rev": round(total_rev, 2), "adr": adr}


# ═══════════════════════════════════════════════════════════
# 작업 1: Budget vs DB 비교 출력
# ═══════════════════════════════════════════════════════════
def task1_compare(db: dict, wb) -> None:
    print("\n" + "=" * 90)
    print("📊 작업 1: 2026 사업계획 Budget vs DB 파싱 결과 비교")
    print("  ※ DB 실적: 2026년 YTD (1월~4월22일 투숙 기준)")
    print("  ※ REV 단위: 백만원 | ADR 단위: 천원")
    print("=" * 90)

    # --- 연간 Budget from 갑지(목표수립) ---
    ws_gap = wb["갑지(목표수립)"]
    annual_budget = {}
    import re as _re
    for row in ws_gap.iter_rows(values_only=True):
        name = row[1]
        if name and _re.match(r'\d+\.', str(name)):
            rn = row[3] or 0
            adr = row[4] or 0
            rev_cheonwon = row[6] or 0
            rev_m = round(float(rev_cheonwon) / 1000, 0)  # 千원 → 百万원
            annual_budget[str(name)] = {"rn": int(rn), "adr": round(float(adr), 1), "rev_m": rev_m}

    # --- 비교 테이블 헤더 ---
    FMT = "{:<28} {:>8} {:>8} {:>8} | {:>8} {:>8} {:>8} | {:>6} {:>6} {:>6}"
    SEP = "-" * 90
    print(FMT.format("사업장", "Bud.RN", "Bud.ADR", "Bud.REV", "DB.RN", "DB.ADR", "DB.REV", "RN%", "ADR%", "REV%"))
    print(FMT.format("", "(연간)", "(천원)", "(백만)", "YTD", "(천원)", "(백만)", "±", "±", "±"))
    print(SEP)

    # 1월~3월 합산 Budget (3월까지 완전한 월)
    compare_months = ["1월", "2월", "3월"]  # 완전한 월만 비교
    compare_month_keys = ["202601", "202602", "202603"]

    total_bud_rn = total_bud_rev = 0
    total_db_rn = total_db_rev = 0

    for sheet_name, display_name, region, db_props in PROPERTY_DEFS:
        # Budget 1~3월 합산
        bud_data = load_budget_monthly(wb, sheet_name)
        bud_rn_q1 = sum(bud_data.get(m, {}).get("rn", 0) for m in compare_months)
        bud_rev_q1 = sum(bud_data.get(m, {}).get("rev_m", 0) for m in compare_months)
        bud_adr_q1 = round((bud_rev_q1 * 1000) / bud_rn_q1) if bud_rn_q1 > 0 else 0

        # DB 1~3월 합산
        db_rn = db_rev = 0
        for m in compare_month_keys:
            d = sum_db_props(db, db_props, m)
            db_rn += d["net_rn"]
            db_rev += d["net_rev"]
        db_adr = round((db_rev * 1000) / db_rn) if db_rn > 0 else 0

        if bud_rn_q1 == 0 and db_rn == 0:
            continue  # 데이터 없는 사업장 스킵

        # 오차율
        def pct(actual, budget):
            if budget == 0:
                return "—"
            return f"{(actual/budget - 1)*100:+.0f}%"

        print(FMT.format(
            display_name[:28],
            f"{bud_rn_q1:,}", f"{bud_adr_q1:,}", f"{bud_rev_q1:,.0f}",
            f"{db_rn:,}", f"{db_adr:,}", f"{db_rev:,.0f}",
            pct(db_rn, bud_rn_q1), pct(db_adr, bud_adr_q1), pct(db_rev, bud_rev_q1)
        ))

        total_bud_rn += bud_rn_q1; total_bud_rev += bud_rev_q1
        total_db_rn += db_rn; total_db_rev += db_rev

    total_bud_adr = round((total_bud_rev * 1000) / total_bud_rn) if total_bud_rn > 0 else 0
    total_db_adr = round((total_db_rev * 1000) / total_db_rn) if total_db_rn > 0 else 0

    print(SEP)
    print(FMT.format(
        "【 전체 합계 (1~3월) 】",
        f"{total_bud_rn:,}", f"{total_bud_adr:,}", f"{total_bud_rev:,.0f}",
        f"{total_db_rn:,}", f"{total_db_adr:,}", f"{total_db_rev:,.0f}",
        f"{(total_db_rn/total_bud_rn-1)*100:+.0f}%" if total_bud_rn else "—",
        f"{(total_db_adr/total_bud_adr-1)*100:+.0f}%" if total_bud_adr else "—",
        f"{(total_db_rev/total_bud_rev-1)*100:+.0f}%" if total_bud_rev else "—",
    ))

    # --- 연간 Budget 요약 ---
    print("\n" + "─" * 60)
    print("📋 2026 Budget 연간 목표 (전체)")
    total_ann_bud_rn = sum(v["rn"] for v in annual_budget.values())
    total_ann_bud_rev = sum(v["rev_m"] for v in annual_budget.values())
    total_ann_bud_adr = round((total_ann_bud_rev * 1000) / total_ann_bud_rn) if total_ann_bud_rn else 0
    print(f"  RN:  {total_ann_bud_rn:>12,}  (연간 목표)")
    print(f"  ADR: {total_ann_bud_adr:>12,}  천원")
    print(f"  REV: {total_ann_bud_rev:>12,.0f}  백만원")

    # DB 2026 전체 YTD
    db_2026 = db.get("monthly_total", {})
    db_ytd_rn = sum(v.get("net_rn", 0) for k, v in db_2026.items() if k.startswith("2026"))
    db_ytd_rev = sum(v.get("net_rev", 0) for k, v in db_2026.items() if k.startswith("2026"))
    db_ytd_adr = round((db_ytd_rev * 1000) / db_ytd_rn) if db_ytd_rn > 0 else 0
    print(f"\n  DB 2026 YTD (1~4월22일):")
    print(f"  RN:  {db_ytd_rn:>12,}")
    print(f"  ADR: {db_ytd_adr:>12,}  천원")
    print(f"  REV: {db_ytd_rev:>12,.0f}  백만원")
    print(f"  → 예산 대비 진행률: RN {db_ytd_rn/total_ann_bud_rn*100:.1f}% "
          f"(1~4월 이론치 ~{(4/12)*100:.0f}%)")
    print("=" * 90)


# ═══════════════════════════════════════════════════════════
# 작업 2: daily_notes.json property_performance 교체
# ═══════════════════════════════════════════════════════════
def task2_update(db: dict, wb) -> None:
    print("\n" + "=" * 60)
    print("🔄 작업 2: daily_notes.json property_performance 교체")
    print("=" * 60)

    # --- 기존 daily_notes.json 로드 ---
    notes = json.loads(NOTES_JSON.read_text(encoding="utf-8"))

    # --- 권역별 property 리스트 구성 ---
    region_props = defaultdict(list)

    for sheet_name, display_name, region, db_props in PROPERTY_DEFS:
        # Budget 월별 로드
        bud = load_budget_monthly(wb, sheet_name)

        prop_entry = {"name": display_name}

        for sm, sm_disp, month_label in zip(
            STAY_MONTHS, STAY_MONTHS_DISP,
            ["4월", "5월", "6월"]
        ):
            db_month = sum_db_props(db, db_props, sm)  # sm = "202604" etc.
            bud_month = bud.get(month_label, {})  # budget month

            actual_rn = db_month["net_rn"]
            actual_rev = db_month["net_rev"]
            actual_adr = db_month["adr"]
            target_rn = bud_month.get("rn", 0)
            target_rev = bud_month.get("rev_m", 0.0)

            achievement = round((actual_rn / target_rn * 100), 1) if target_rn > 0 else 0.0

            # YoY: DB 전년도 같은 월
            prev_year_month = sm.replace("2026", "2025")
            db_last = sum_db_props(db, db_props, prev_year_month)
            last_rn = db_last["net_rn"]
            yoy_pct = round((actual_rn / last_rn - 1) * 100, 1) if last_rn > 0 else 0.0

            prop_entry[sm_disp] = {
                "rns": actual_rn,
                "adr": actual_adr,
                "rev": round(actual_rev, 0),
                "target_rns": target_rn,
                "target_rev": round(target_rev, 0),
                "achievement": achievement,
                "yoy_pct": yoy_pct,
                "last_rns": last_rn,
            }

        region_props[region].append(prop_entry)
        print(f"  ✓ {display_name} ({region})")

    # --- property_performance 교체 ---
    pp = {
        "_description": "DB 파싱(온북 27/28/43/44번) + 사업계획 xlsx 기반 자동 산출",
        "_status": "auto_synced",
        "_last_sync": db.get("generated_at", ""),
        "vivaldi":  region_props["vivaldi"],
        "central":  region_props["central"],
        "south":    region_props["south"],
        "apac":     region_props["apac"],
    }
    notes["property_performance"] = pp

    NOTES_JSON.write_text(json.dumps(notes, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n  ✅ daily_notes.json 저장 완료")
    print(f"  vivaldi: {len(region_props['vivaldi'])}개 사업장")
    print(f"  central: {len(region_props['central'])}개 사업장")
    print(f"  south:   {len(region_props['south'])}개 사업장")
    print(f"  apac:    {len(region_props['apac'])}개 사업장")
    print("=" * 60)


def main():
    if not DB_JSON.exists():
        print(f"❌ DB JSON 없음: {DB_JSON}")
        print("   먼저 parse_raw_db.py를 실행하세요.")
        sys.exit(1)

    if not BUDGET_XLSX.exists():
        print(f"❌ Budget xlsx 없음: {BUDGET_XLSX}")
        sys.exit(1)

    print(f"📂 DB JSON 로드: {DB_JSON}")
    db = json.loads(DB_JSON.read_text(encoding="utf-8"))

    print(f"📂 Budget xlsx 로드: {BUDGET_XLSX.name}")
    wb = openpyxl.load_workbook(str(BUDGET_XLSX), data_only=True)

    # 작업 1
    task1_compare(db, wb)

    # 작업 2
    task2_update(db, wb)

    print("\n✅ 완료! build.py를 실행하여 index.html을 갱신하세요:")
    print("   python scripts/build.py")


if __name__ == "__main__":
    main()
