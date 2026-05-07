#\!/usr/bin/env python3
"""
해외사업장 엑셀 → JSON 변환 스크립트
소노호텔앤리조트 GS 해외사업장 사업계획 2026 ver.xlsx → overseas_data.json

© 2026 GS팀 · Haein Kim Manager
"""
import json
import logging
import os
import sys
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

ROOT = Path(__file__).parent.parent
DATA_DIR = ROOT / "data"
DOCS_DIR = ROOT / "docs"
EXCEL_FILE = DATA_DIR / "overseas" / "overseas_2026.xlsx"
WRH_FILE = DATA_DIR / "overseas" / "2026년 WRH 객실계획.xlsx"
WRH_CLOSING_DIR = DATA_DIR / "overseas"
WRH_CLOSING_GLOB = "*WRH Monthly Closing*.xlsx"


def sv(v):
    """Safe numeric value"""
    if v is None:
        return 0
    if isinstance(v, str):
        if '#' in v or v.strip() == '-':
            return 0
        try:
            return float(v.replace(',', ''))
        except ValueError:
            return 0
    return float(v) if isinstance(v, (int, float)) else 0


def pct(v):
    """Convert ratio (0.37) or percentage to display percentage (37.0)"""
    val = sv(v)
    if abs(val) <= 2:
        return round(val * 100, 1)
    return round(val, 1)


def parse_overseas_excel(filepath):
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl 미설치 — pip install openpyxl")
        return None

    wb = openpyxl.load_workbook(filepath, data_only=True)
    logger.info(f"엑셀 로드: {filepath.name} (시트: {wb.sheetnames})")

    # ── SUMMARY ──
    ws = wb['Ⅰ. SUMMARY']
    update_date = str(ws.cell(3, 17).value)[:10] if ws.cell(3, 17).value else ''
    exchange_rate = sv(ws.cell(4, 16).value)

    hp_summary = []
    for r, name in [(8, '객실'), (9, '골프'), (10, 'F&B'), (11, '합계')]:
        row = {
            'name': name,
            'year': {
                'target': round(sv(ws.cell(r, 5).value), 1),
                'actual': round(sv(ws.cell(r, 6).value), 1),
                'rate': pct(ws.cell(r, 7).value)
            },
            'prev': round(sv(ws.cell(r, 8).value), 1),
            'yoy': pct(ws.cell(r, 9).value),
            'months': []
        }
        month_cols = [(11, '1월'), (14, '2월'), (17, '3월'), (20, '4월')]
        for col_s, mn in month_cols:
            t = sv(ws.cell(r, col_s).value)
            a = sv(ws.cell(r, col_s + 1).value)
            if t > 0 or a > 0:
                row['months'].append({
                    'month': mn,
                    'target': round(t, 1),
                    'actual': round(a, 1),
                    'rate': pct(ws.cell(r, col_s + 2).value) if ws.cell(r, col_s + 2).value else 0
                })
        hp_summary.append(row)

    guam_list = []
    for r, loc, seg in [(12, '망길라오', '아웃바운드'), (13, '망길라오', '로컬'),
                         (14, '탈로포포', '아웃바운드'), (15, '탈로포포', '로컬'),
                         (16, '괌합계', '합계')]:
        row = {
            'location': loc, 'segment': seg,
            'year': {
                'target': sv(ws.cell(r, 5).value),
                'actual': sv(ws.cell(r, 6).value),
                'rate': pct(ws.cell(r, 7).value)
            },
            'prev': sv(ws.cell(r, 8).value),
            'yoy': pct(ws.cell(r, 9).value),
            'months': []
        }
        for col_s, mn in [(11, '1월'), (14, '2월'), (17, '3월'), (20, '4월')]:
            row['months'].append({
                'month': mn,
                'target': sv(ws.cell(r, col_s).value),
                'actual': sv(ws.cell(r, col_s + 1).value),
                'rate': pct(ws.cell(r, col_s + 2).value)
            })
        guam_list.append(row)

    # ── 하이퐁_DATA ──
    ws_hp = wb['하이퐁_DATA']
    hp_rates = {'target': sv(ws_hp.cell(1, 6).value), 'actual': sv(ws_hp.cell(2, 6).value)}
    hp_categories = []
    for r in range(7, 75):
        cat5 = ws_hp.cell(r, 5).value
        if cat5 is None:
            continue
        cat5 = str(cat5).strip()
        if not cat5:
            continue
        parent = str(ws_hp.cell(r, 3).value or '').strip()
        is_total = ('TOTAL' in cat5.upper() or '합' in cat5 or 'GRAND' in cat5.upper())
        entry = {
            'name': cat5, 'parent': parent, 'is_total': is_total,
            'total': {
                'target': round(sv(ws_hp.cell(r, 6).value), 2),
                'actual': round(sv(ws_hp.cell(r, 7).value), 2),
                'prev': round(sv(ws_hp.cell(r, 8).value), 2)
            },
            'months': []
        }
        for m in range(5):
            base = 9 + m * 3
            entry['months'].append({
                'target': round(sv(ws_hp.cell(r, base).value), 2),
                'actual': round(sv(ws_hp.cell(r, base + 1).value), 2),
                'prev': round(sv(ws_hp.cell(r, base + 2).value), 2)
            })
        hp_categories.append(entry)

    # ── 망길/탈로 DATA ──
    def extract_guam_sheet(sheet_name):
        ws_g = wb[sheet_name]
        cats = []
        for r in range(6, 58):
            cat2 = ws_g.cell(r, 2).value
            cat3 = ws_g.cell(r, 3).value
            if cat2 is None and cat3 is None:
                continue
            cat2 = str(cat2 or '').strip()
            cat3 = str(cat3 or '').strip()
            name = cat2 if cat2 else cat3
            if not name:
                continue
            is_sub = bool(cat3 and not cat2)
            is_total = 'TOTAL' in name.upper()
            entry = {
                'name': name, 'is_sub': is_sub, 'is_total': is_total,
                'total': {
                    'pax_budget': sv(ws_g.cell(r, 5).value),
                    'rev_budget': round(sv(ws_g.cell(r, 6).value), 1),
                    'pax_actual': sv(ws_g.cell(r, 7).value),
                    'rev_actual': round(sv(ws_g.cell(r, 8).value), 1),
                    'pax_rate': pct(ws_g.cell(r, 11).value),
                    'rev_rate': pct(ws_g.cell(r, 12).value),
                    'pax_prev': sv(ws_g.cell(r, 13).value),
                    'rev_prev': round(sv(ws_g.cell(r, 14).value), 1),
                },
                'months': []
            }
            for m in range(5):
                base = 21 + m * 16
                entry['months'].append({
                    'pax_budget': sv(ws_g.cell(r, base).value),
                    'rev_budget': round(sv(ws_g.cell(r, base + 1).value), 1),
                    'pax_actual': sv(ws_g.cell(r, base + 2).value),
                    'rev_actual': round(sv(ws_g.cell(r, base + 3).value), 1),
                    'pax_rate': pct(ws_g.cell(r, base + 6).value),
                    'rev_rate': pct(ws_g.cell(r, base + 7).value),
                })
            cats.append(entry)
        return cats

    mg_cats = extract_guam_sheet('망길_DATA')
    tl_cats = extract_guam_sheet('탈로_DATA')

    guam_exchange = 1480
    result = {
        'generated_at': update_date,
        'exchange_rate_vnd': exchange_rate,
        'exchange_rate_usd': guam_exchange,
        'haiphong': {
            'summary': hp_summary,
            'exchange_rates': hp_rates,
            'categories': hp_categories
        },
        'guam': {
            'summary': guam_list,
            'mangilao': mg_cats,
            'talopopo': tl_cats
        }
    }
    return result


def parse_wrh_excel(filepath):
    """WRH 객실계획 엑셀 → hawaii dict (프론트엔드 호환 구조)"""
    try:
        import openpyxl
    except ImportError:
        return None

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['⑴-4 WRHⅠ']
    logger.info(f"WRH 엑셀 로드: {filepath.name}")

    # Annual Total (col 6), row 5 = 23Y start
    ann = 5
    total = {
        'budget_rn': sv(ws.cell(ann + 9, 6).value),      # 26Y Budget RN
        'budget_rev': round(sv(ws.cell(ann + 10, 6).value)),  # 26Y Budget Rev
        'budget_adr': round(sv(ws.cell(ann + 11, 6).value), 2),
        'actual_rn': 0,   # 26Y Actual — WRH에는 미포함
        'actual_rev': 0,
        'actual_adr': 0,
        'ly_rn': sv(ws.cell(ann + 6, 6).value),           # 25Y RN
        'ly_rev': round(sv(ws.cell(ann + 7, 6).value)),
        'ly_adr': round(sv(ws.cell(ann + 8, 6).value), 2),
    }

    # Monthly data — 54 rows apart, starting at row 59
    month_starts = {m: 59 + (m - 1) * 54 for m in range(1, 13)}
    monthly = []
    for m in range(1, 13):
        r = month_starts[m]
        entry = {
            'month': m,
            'budget_rn': sv(ws.cell(r + 9, 6).value),
            'budget_rev': round(sv(ws.cell(r + 10, 6).value)),
            'budget_adr': round(sv(ws.cell(r + 11, 6).value), 2),
            'actual_rn': 0,
            'actual_rev': 0,
            'actual_adr': 0,
            'ly_rn': sv(ws.cell(r + 6, 6).value),
            'ly_rev': round(sv(ws.cell(r + 7, 6).value)),
        }
        monthly.append(entry)

    # Wholesale (col 14 = Wholesale Sub TL)
    # YTD = Jan ~ latest month with 25Y actual data
    import datetime
    current_month = min(datetime.date.today().month, 12)
    ytd_months = range(1, current_month + 1)
    ws25_ytd_rn = sum(sv(ws.cell(month_starts[m] + 6, 14).value) for m in ytd_months)
    ws25_ytd_rev = round(sum(sv(ws.cell(month_starts[m] + 7, 14).value) for m in ytd_months))
    ws26_ytd_rn = sum(sv(ws.cell(month_starts[m] + 9, 14).value) for m in ytd_months)
    ws26_ytd_rev = round(sum(sv(ws.cell(month_starts[m] + 10, 14).value) for m in ytd_months))

    wholesale = {
        '2025': {'ytd_rn': ws25_ytd_rn, 'ytd_rev': ws25_ytd_rev},
        '2026': {'ytd_rn': ws26_ytd_rn, 'ytd_rev': ws26_ytd_rev},
    }

    return {
        'total': total,
        'monthly': monthly,
        'wholesale': wholesale,
    }


def parse_wrh_monthly_closing(filepath):
    """WRH Monthly Closing 엑셀 → 월별 actual 객실/매출/IS 데이터.

    sheet '1) Rooms' 기반:
      Row 5 col5='Total' col6='Actual 2026'
      Row 6 col6=YTD 마감월, col7..18 = 1..12월
      Row 7 = OCC, R8 = Available Rooms, R9 = RN, R10 = ADR, R11 = RevPAR, R17 = Revenue
    sheet '2) IS(Summary)' 기반:
      Row 7 = Total Revenue, R21 = Departmental Profit, R33 = GOP,
      R39 = Net Operating Income, R41 = Income before tax
    """
    try:
        import openpyxl
    except ImportError:
        return None

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb['1) Rooms']
    logger.info(f"WRH Monthly Closing 로드: {filepath.name}")

    def cell(sheet, r, c):
        return sheet.cell(r, c).value

    ytd_through = int(sv(cell(ws, 6, 6)))
    rooms_total = {
        'ytd_through': ytd_through,
        'occ_pct': round(sv(cell(ws, 7, 6)) * 100, 2),
        'rn': int(sv(cell(ws, 9, 6))),
        'adr': round(sv(cell(ws, 10, 6)), 2),
        'revpar': round(sv(cell(ws, 11, 6)), 2),
        'rev': round(sv(cell(ws, 17, 6)), 2),
        'avail_rooms': int(sv(cell(ws, 8, 6))),
    }

    rooms_monthly = []
    for m in range(1, 13):
        col = 6 + m
        rn = int(sv(cell(ws, 9, col)))
        rooms_monthly.append({
            'month': m,
            'occ_pct': round(sv(cell(ws, 7, col)) * 100, 2),
            'rn': rn,
            'adr': round(sv(cell(ws, 10, col)), 2),
            'revpar': round(sv(cell(ws, 11, col)), 2),
            'rev': round(sv(cell(ws, 17, col)), 2),
            'has_actual': rn > 0,
        })

    income_statement = None
    if '2) IS(Summary)' in wb.sheetnames:
        ws2 = wb['2) IS(Summary)']
        is_rows = [
            (7, 'total_revenue'), (14, 'dept_expenses'), (21, 'dept_profit'),
            (28, 'undist_expenses'), (33, 'gop'), (34, 'fixed_charges'),
            (39, 'noi'), (40, 'd_and_a'), (41, 'income_before_tax'),
        ]
        income_statement = {'ytd': {}, 'monthly': {}}
        for r, key in is_rows:
            income_statement['ytd'][key] = round(sv(cell(ws2, r, 6)), 2)
            income_statement['monthly'][key] = [round(sv(cell(ws2, r, 6 + m)), 2) for m in range(1, 13)]

    return {
        'rooms_total': rooms_total,
        'rooms_monthly': rooms_monthly,
        'income_statement': income_statement,
        'source_file': filepath.name,
    }


def merge_wrh_actuals(hawaii, closing):
    """parse_wrh_excel 결과에 Monthly Closing actuals를 병합."""
    if not closing or not hawaii:
        return hawaii

    rt = closing['rooms_total']
    hawaii['total']['actual_rn'] = rt['rn']
    hawaii['total']['actual_rev'] = round(rt['rev'])
    hawaii['total']['actual_adr'] = rt['adr']
    hawaii['total']['actual_occ'] = rt['occ_pct']
    hawaii['total']['actual_revpar'] = rt['revpar']
    hawaii['total']['ytd_through'] = rt['ytd_through']

    by_month = {m['month']: m for m in closing['rooms_monthly']}
    for entry in hawaii['monthly']:
        m = entry['month']
        cm = by_month.get(m)
        if cm and cm['has_actual']:
            entry['actual_rn'] = cm['rn']
            entry['actual_rev'] = round(cm['rev'])
            entry['actual_adr'] = cm['adr']
            entry['actual_occ'] = cm['occ_pct']
            entry['actual_revpar'] = cm['revpar']
        else:
            entry['actual_occ'] = 0
            entry['actual_revpar'] = 0

    if closing.get('income_statement'):
        hawaii['income_statement'] = closing['income_statement']
    hawaii['actual_source'] = closing.get('source_file')
    return hawaii


def find_latest_closing_file():
    candidates = sorted(WRH_CLOSING_DIR.glob(WRH_CLOSING_GLOB))
    return candidates[-1] if candidates else None


def main():
    if not EXCEL_FILE.exists():
        logger.warning(f"엑셀 파일 없음: {EXCEL_FILE} — 기존 JSON 유지")
        return

    data = parse_overseas_excel(EXCEL_FILE)
    if not data:
        logger.error("파싱 실패")
        sys.exit(1)

    # ── 하와이 WRH Budget ──
    if WRH_FILE.exists():
        hawaii = parse_wrh_excel(WRH_FILE)
        if hawaii:
            data['hawaii'] = hawaii
            logger.info(f"✓ 하와이 WRH Budget 데이터 추가 (월 {len(hawaii['monthly'])}개)")
    else:
        logger.warning(f"WRH 파일 없음: {WRH_FILE} — 하와이 Budget 데이터 미포함")

    # ── 하와이 WRH Monthly Closing Actuals ──
    closing_file = find_latest_closing_file()
    if closing_file and 'hawaii' in data:
        try:
            closing = parse_wrh_monthly_closing(closing_file)
            if closing:
                merge_wrh_actuals(data['hawaii'], closing)
                logger.info(
                    f"✓ 하와이 WRH Actual 병합 ({closing_file.name}, "
                    f"YTD M{closing['rooms_total']['ytd_through']})"
                )
        except Exception as e:
            logger.warning(f"✗ WRH Monthly Closing 파싱 실패: {e}")
    elif not closing_file:
        logger.info("WRH Monthly Closing 파일 없음 — Actual 병합 건너뜀")

    # Save to data/
    out_data = DATA_DIR / "overseas_data.json"
    out_data.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    logger.info(f"✓ {out_data} 저장")

    # Save to docs/data/ (minified for frontend)
    out_docs = DOCS_DIR / "data" / "overseas_data.json"
    out_docs.parent.mkdir(parents=True, exist_ok=True)
    out_docs.write_text(json.dumps(data, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')
    logger.info(f"✓ {out_docs} 저장 ({out_docs.stat().st_size:,} bytes)")


if __name__ == '__main__':
    main()
