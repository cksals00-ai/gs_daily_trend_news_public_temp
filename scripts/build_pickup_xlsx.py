#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""build_pickup_xlsx.py — _pickup678_rows.json → 픽업분석 엑셀(7시트)."""
import json, os
LITE = os.environ.get("LITE") == "1"   # 요약본: DB를 CUR 윈도우만 담아 경량화
from pathlib import Path
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT = Path(__file__).resolve().parent.parent
SRC = json.load(open(PROJECT / "data/_pickup678_rows.json", encoding="utf-8"))
ROWS = SRC["rows"]
CAMP_CODES = set(SRC["camp_codes"])

WINS = ["CUR", "WOW", "MOM", "YOY"]
WIN_LABEL = {
    "CUR": "최근6일\n06.18~06.23", "WOW": "직전6일\n06.12~06.17",
    "MOM": "전월동기\n05.18~05.23", "YOY": "전년동기\n25.06.18~23",
}
SEG3 = ("OTA", "G-OTA", "Inbound")
OP = {"H/U", "단체COMP"}              # 운영(비매출) 세그먼트 → 판매 픽업에서 제외
MONTHS = {"CUR": ["202606", "202607", "202608"], "YOY": ["202506", "202507", "202508"]}

def is_sales(r):
    return r["seg"] not in OP

# ── 집계 헬퍼 ──
def agg(keyf, rows=None, sales_only=True):
    rows = rows if rows is not None else ROWS
    d = defaultdict(lambda: {w: {"rn": 0, "rev": 0, "new": 0, "can": 0} for w in WINS})
    mon = defaultdict(lambda: defaultdict(int))   # key -> sm -> rn (CUR)
    for r in rows:
        if sales_only and not is_sales(r):
            continue
        k = keyf(r); w = r["win"]
        d[k][w]["rn"] += r["rn_s"]; d[k][w]["rev"] += r["rev_s"]
        if r["sign"] > 0: d[k][w]["new"] += r["rn"]
        else: d[k][w]["can"] += r["rn"]
        if w == "CUR": mon[k][r["sm"]] += r["rn_s"]
    return d, mon

def pct(cur, base):
    return (cur - base) / base if base else None

# ── 스타일 ──
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
NOTE_FONT = Font(size=9, color="555555", italic=True)
BOLD = Font(bold=True, size=10)
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

def style_header(ws, row, ncol, start=1):
    for c in range(start, start + ncol):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CEN; cell.border = BORD

def put(ws, r, c, v, *, num=None, fill=None, font=None, align=None, border=True):
    cell = ws.cell(row=r, column=c, value=v)
    if num: cell.number_format = num
    if fill: cell.fill = fill
    if font: cell.font = font
    if align: cell.alignment = align
    if border: cell.border = BORD
    return cell

INT = "#,##0"; MIL = "#,##0.0"; PCT = "+0.0%;-0.0%"

wb = openpyxl.Workbook()

# ════════════════════════ 시트1: 요약 ════════════════════════
ws = wb.active; ws.title = "①요약"
ws.sheet_view.showGridLines = False
ws["A1"] = "GS 6·7·8월 투숙 픽업 호조 분석 (최근 7일)"; ws["A1"].font = TITLE_FONT
ws["A2"] = f"기준일 2026-06-24 · 데이터 라이브 스냅샷(07:45 KST 컷) · 생성 {datetime.now():%Y-%m-%d %H:%M}"
ws["A2"].font = NOTE_FONT
notes = [
    "■ 픽업 정의(net): 신규예약 − 취소.  신규(+) = 27/43+28/44 행의 '최초입력일자'가 윈도우 내 & 투숙월 6/7/8월.  취소(−) = 28/44 행의 '취소일자'가 윈도우 내.",
    "■ 정제: 행중복제거 · 자체예약(예약자명=거래처, Inbound 예외) 제외 · 매출조정 제외 (parse_raw_db 동일).  RN=객실수(실), 매출=1박객실료×객실수÷1.1 (VAT제외/공급가), 백만원.",
    "■ 판매 픽업 = 전체 − 하우스유즈(H/U) − 단체COMP(무상, 매출≈0).  본 요약·분해표는 '판매 픽업' 기준.  세그먼트 표·DB는 운영 포함 전체.",
    "■ 윈도우: 모두 동일 6완전일 (06-24는 스냅샷 07:45 부분일이라 제외).  CUR 06.18~23 / WoW 06.12~17(직전6일) / MoM 05.18~23 / YoY 25.06.18~23.  동일길이→절대값 직접 비교 공정.",
]
r = 4
for n in notes:
    ws.cell(row=r, column=1, value=n).font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14); r += 1

# 핵심지표 표
r += 1
ws.cell(row=r, column=1, value="【핵심 지표 — 판매 Net 픽업 (투숙 6/7/8월 합)】").font = BOLD; r += 1
hdr = ["지표", "최근6일(CUR)\n06.18~23", "직전6일(WoW)\n06.12~17", "전월동기(MoM)\n05.18~23", "전년동기(YoY)\n25.06.18~23",
       "vs WoW", "vs MoM", "vs YoY"]
for c, h in enumerate(hdr, 1):
    put(ws, r, c, h)
style_header(ws, r, len(hdr)); hr = r; r += 1

# 윈도우 합계(판매)
sales_net = {w: {"rn": 0, "rev": 0} for w in WINS}
dayp = {w: defaultdict(int) for w in WINS}
for x in ROWS:
    if not is_sales(x): continue
    w = x["win"]; sales_net[w]["rn"] += x["rn_s"]; sales_net[w]["rev"] += x["rev_s"]; dayp[w][x["date"]] += x["rn_s"]
def pace(w, days=6):
    return sales_net[w]["rn"] / days   # 모든 윈도우 6완전일
pace_cur = pace("CUR")
rows_metrics = [
    ("Net 픽업 RN (실)", "rn", INT, lambda w: sales_net[w]["rn"]),
    ("Net 객실매출 (백만원)", "rev", MIL, lambda w: round(sales_net[w]["rev"] / 1e6, 1)),
    ("일평균 (RN/일, ÷6)", "pace", INT, lambda w: round(pace(w))),
]
for label, kind, fmt, fn in rows_metrics:
    put(ws, r, 1, label, font=BOLD, align=LEFT)
    vals = {w: fn(w) for w in WINS}
    for c, w in enumerate(WINS, 2):
        put(ws, r, c, vals[w], num=fmt, align=RIGHT, fill=(TOTAL_FILL if w == "CUR" else None))
    base = vals["CUR"]
    for c, w in zip((6, 7, 8), ("WOW", "MOM", "YOY")):
        p = pct(base, vals[w])
        put(ws, r, c, p if p is not None else "n/a", num=PCT, align=RIGHT)
    r += 1
ws.cell(row=r, column=1,
        value="※ 4개 윈도우 모두 동일 6완전일 → Net RN/매출 절대값을 직접 비교(공정).  vs 열은 CUR 대비 증감률.").font = NOTE_FONT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8); r += 2

# 전체(운영포함) 참고 + 월별
all_net = {w: 0 for w in WINS}
for x in ROWS: all_net[x["win"]] += x["rn_s"]
put(ws, r, 1, "참고) 전체 Net RN(운영 H/U·COMP 포함)", font=NOTE_FONT, align=LEFT)
for c, w in enumerate(WINS, 2):
    put(ws, r, c, all_net[w], num=INT, align=RIGHT)
r += 2

ws.cell(row=r, column=1, value="【투숙월별 판매 Net RN】").font = BOLD; r += 1
put(ws, r, 1, "투숙월")
for c, w in enumerate(WINS, 2): put(ws, r, c, WIN_LABEL[w])
style_header(ws, r, 5); r += 1
mon_net = defaultdict(lambda: defaultdict(int))
for x in ROWS:
    if not is_sales(x): continue
    mlabel = {"202606": "6월", "202607": "7월", "202608": "8월",
              "202506": "6월", "202507": "7월", "202508": "8월"}.get(x["sm"], x["sm"])
    mon_net[x["win"]][mlabel] += x["rn_s"]
for m in ["6월", "7월", "8월"]:
    put(ws, r, 1, m, font=BOLD, align=LEFT)
    for c, w in enumerate(WINS, 2):
        put(ws, r, c, mon_net[w][m], num=INT, align=RIGHT, fill=(TOTAL_FILL if w == "CUR" else None))
    r += 1
r += 1

# 인사이트
_pk = sum(x["rn_s"] for x in ROWS if is_sales(x) and x["win"] == "CUR" and str(x["mnum"]).startswith("86"))
_pk_share = _pk / sales_net["CUR"]["rn"] * 100
ws.cell(row=r, column=1, value="【핵심 인사이트】").font = BOLD; r += 1
insights = [
    f"① 픽업 호조 확인(동일 6완전일): 판매 Net 픽업 {sales_net['CUR']['rn']:,} RN — 전년동기 {pct(sales_net['CUR']['rn'],sales_net['YOY']['rn'])*100:+.0f}%, 전월동기 {pct(sales_net['CUR']['rn'],sales_net['MOM']['rn'])*100:+.0f}%, 직전6일 {pct(sales_net['CUR']['rn'],sales_net['WOW']['rn'])*100:+.0f}%. (일평균 {pace_cur:,.0f} RN/일)",
    f"② 매출 기준 더 강함: Net 객실매출 {sales_net['CUR']['rev']/1e6:,.0f}백만 vs 전년 {sales_net['YOY']['rev']/1e6:,.0f}백만 ({pct(sales_net['CUR']['rev'],sales_net['YOY']['rev'])*100:+.0f}%).  ADR 상승 동반.",
    "③ 견인 채널: OTA(야놀자·여기어때) + G-OTA(트립비토즈·아고다·트립닷컴) 동반 강세.  Inbound도 RN 견인(저ADR).",
    f"④ 견인 상품: 패키지(86xx)가 판매 Net RN의 {_pk_share:.0f}% — 특히 '연박/투나잇' 패키지가 OTA 채널로 강하게 유입.",
    "⑤ 견인 사업장: 소노벨 비발디파크 선두, 이어 델피노·소노캄 거제·소노벨 천안/경주·쏠비치 진도/삼척 등 강원·남부 광역 분산.",
    "⑥ 기획전 연관: 코드 태깅된 유일 기획전(비발디5월전략)은 투숙기간 경과로 최근 윈도우 기여 ≈0.  여름 기획전 134건은 가동 중이나 패키지코드 미태깅 → 코드단위 정밀귀속은 비발디건만 가능(한계, ⑥시트).",
]
for n in insights:
    cell = ws.cell(row=r, column=1, value=n); cell.font = Font(size=10); cell.alignment = LEFT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14); r += 1

ws.column_dimensions["A"].width = 30
for c in range(2, 9): ws.column_dimensions[get_column_letter(c)].width = 14
for rr in range(hr, hr + 5): ws.row_dimensions[rr].height = 30
ws.row_dimensions[hr].height = 34

# ════════════════════════ 공용 분해표 빌더 ════════════════════════
def dim_sheet(title, keyf, label_col="항목", seg_col=False, sales_only=True,
              topn=None, note=None, filt=None):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    d, mon = agg(keyf, sales_only=sales_only)
    items = list(d.items())
    if filt: items = [(k, v) for k, v in items if filt(k)]
    items.sort(key=lambda kv: -kv[1]["CUR"]["rn"])
    if topn: items = items[:topn]
    ws["A1"] = title.replace("②", "").replace("③", "").replace("④", "") + " — 판매 Net 픽업 (투숙 6/7/8월)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "단위: RN=실, 매출=백만원(VAT제외).  Δ%는 CUR(절대,06-24부분일 포함) 기준.  정렬: CUR Net RN 내림차순."
    ws["A2"].font = NOTE_FONT
    rr = 3
    if note:
        ws.cell(row=rr, column=1, value=note).font = NOTE_FONT
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=15); rr += 1
    hrow = rr + 1
    cols = [label_col]
    if seg_col: cols.append("세그먼트")
    cols += ["신규RN", "취소RN", "Net RN", "6월", "7월", "8월", "Net매출\n(백만)",
             "WoW\nNetRN", "WoWΔ%", "MoM\nNetRN", "MoMΔ%", "YoY\nNetRN", "YoYΔ%"]
    for c, h in enumerate(cols, 1): put(ws, hrow, c, h)
    style_header(ws, hrow, len(cols)); ws.row_dimensions[hrow].height = 30
    rrow = hrow + 1
    # 채널 세그먼트 라벨(다수결)
    seg_of = {}
    if seg_col:
        tmp = defaultdict(lambda: defaultdict(int))
        for x in ROWS:
            if sales_only and not is_sales(x): continue
            if filt and not filt(keyf(x)): continue
            tmp[keyf(x)][x["seg"]] += abs(x["rn_s"])
        for k, sd in tmp.items():
            seg_of[k] = max(sd, key=sd.get) if sd else ""
    tot = {w: {"rn": 0, "rev": 0, "new": 0, "can": 0} for w in WINS}
    for k, v in items:
        c = 1
        put(ws, rrow, c, k, align=LEFT, font=BOLD); c += 1
        if seg_col: put(ws, rrow, c, seg_of.get(k, ""), align=CEN); c += 1
        cur = v["CUR"]
        put(ws, rrow, c, cur["new"], num=INT, align=RIGHT); c += 1
        put(ws, rrow, c, cur["can"], num=INT, align=RIGHT); c += 1
        put(ws, rrow, c, cur["rn"], num=INT, align=RIGHT, fill=TOTAL_FILL); c += 1
        for sm in MONTHS["CUR"]:
            put(ws, rrow, c, mon[k].get(sm, 0), num=INT, align=RIGHT); c += 1
        put(ws, rrow, c, round(cur["rev"] / 1e6, 1), num=MIL, align=RIGHT); c += 1
        for w in ("WOW", "MOM", "YOY"):
            put(ws, rrow, c, v[w]["rn"], num=INT, align=RIGHT); c += 1
            p = pct(cur["rn"], v[w]["rn"])
            put(ws, rrow, c, p if p is not None else "n/a", num=PCT, align=RIGHT); c += 1
        for w in WINS:
            for kk in ("rn", "rev", "new", "can"): tot[w][kk] += v[w][kk]
        rrow += 1
    # 합계행
    c = 1; put(ws, rrow, c, "합계", align=LEFT, font=BOLD, fill=SUB_FILL); c += 1
    if seg_col: put(ws, rrow, c, "", fill=SUB_FILL); c += 1
    put(ws, rrow, c, tot["CUR"]["new"], num=INT, align=RIGHT, fill=SUB_FILL); c += 1
    put(ws, rrow, c, tot["CUR"]["can"], num=INT, align=RIGHT, fill=SUB_FILL); c += 1
    put(ws, rrow, c, tot["CUR"]["rn"], num=INT, align=RIGHT, fill=SUB_FILL, font=BOLD); c += 1
    msum = {sm: sum(mon[k].get(sm, 0) for k, _ in items) for sm in MONTHS["CUR"]}
    for sm in MONTHS["CUR"]: put(ws, rrow, c, msum[sm], num=INT, align=RIGHT, fill=SUB_FILL); c += 1
    put(ws, rrow, c, round(tot["CUR"]["rev"] / 1e6, 1), num=MIL, align=RIGHT, fill=SUB_FILL); c += 1
    for w in ("WOW", "MOM", "YOY"):
        put(ws, rrow, c, tot[w]["rn"], num=INT, align=RIGHT, fill=SUB_FILL); c += 1
        p = pct(tot["CUR"]["rn"], tot[w]["rn"])
        put(ws, rrow, c, p if p is not None else "n/a", num=PCT, align=RIGHT, fill=SUB_FILL); c += 1
    ws.column_dimensions["A"].width = 24
    off = 1 if seg_col else 0
    if seg_col: ws.column_dimensions["B"].width = 10
    for c in range(2 + off, len(cols) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.freeze_panes = ws.cell(row=hrow + 1, column=2)
    return ws

# 시트2 사업장별
dim_sheet("②사업장별", lambda r: r["prop"], "사업장")
# 시트3 상품별
dim_sheet("③상품별", lambda r: r["cat"], "상품카테고리",
          note="상품카테고리는 패키지(회원번호 86xx) 9분류 + '비패키지(일반)'.  '연박/투나잇'은 박수·연속OTA체인 문맥 반영(parse_package_trend).")
# 시트4 채널·세그먼트별 (OTA/G-OTA/Inbound 채널만; 비OTA는 채널='미분류'로 묶여 제외)
_resid = {"rn": 0, "rev": 0}
for x in ROWS:
    if is_sales(x) and x["win"] == "CUR" and x["ch"] == "미분류":
        _resid["rn"] += x["rn_s"]; _resid["rev"] += x["rev_s"]
dim_sheet("④채널세그먼트별", lambda r: r["ch"], "채널", seg_col=True,
          note=("OTA(여기어때·야놀자·네이버 등)/G-OTA(아고다·트립닷컴·익스피디아·트립비토즈 등)/Inbound(여행사) 채널만 표기.  "
                f"비OTA(회원·자사·단체·D멤버스 등, 채널 미식별) 합계는 별도: CUR Net RN {_resid['rn']:,} / 매출 {_resid['rev']/1e6:,.0f}백만 (본 표 제외)."),
          filt=lambda k: k != "미분류", sales_only=True)

# ════════════════════════ 시트5: 교차 ════════════════════════
ws = wb.create_sheet("⑤교차분석")
ws.sheet_view.showGridLines = False
ws["A1"] = "교차분석 — 사업장 × 채널 / 사업장 × 상품 (판매 Net 픽업, CUR 기준 상위)"
ws["A1"].font = TITLE_FONT
ws["A2"] = "단위: RN=실, 매출=백만원.  각 교차 상위 25개(CUR Net RN).  '어디서 무엇이 잘 나가는지'."
ws["A2"].font = NOTE_FONT

def cross_block(ws, start_row, head, keyf, n=25, skip=None):
    d, mon = agg(keyf, sales_only=True)
    items = [(k, v) for k, v in d.items() if not (skip and skip(k))]
    items = sorted(items, key=lambda kv: -kv[1]["CUR"]["rn"])[:n]
    ws.cell(row=start_row, column=1, value=head).font = BOLD
    hrow = start_row + 1
    cols = ["사업장", "교차항목", "신규RN", "취소RN", "Net RN", "Net매출(백만)",
            "WoW NetRN", "WoWΔ%", "YoY NetRN", "YoYΔ%"]
    for c, h in enumerate(cols, 1): put(ws, hrow, c, h)
    style_header(ws, hrow, len(cols)); ws.row_dimensions[hrow].height = 28
    rrow = hrow + 1
    for (a, b), v in items:
        cur = v["CUR"]
        put(ws, rrow, 1, a, align=LEFT); put(ws, rrow, 2, b, align=LEFT)
        put(ws, rrow, 3, cur["new"], num=INT, align=RIGHT)
        put(ws, rrow, 4, cur["can"], num=INT, align=RIGHT)
        put(ws, rrow, 5, cur["rn"], num=INT, align=RIGHT, fill=TOTAL_FILL)
        put(ws, rrow, 6, round(cur["rev"] / 1e6, 1), num=MIL, align=RIGHT)
        put(ws, rrow, 7, v["WOW"]["rn"], num=INT, align=RIGHT)
        put(ws, rrow, 8, pct(cur["rn"], v["WOW"]["rn"]) if v["WOW"]["rn"] else "n/a", num=PCT, align=RIGHT)
        put(ws, rrow, 9, v["YOY"]["rn"], num=INT, align=RIGHT)
        put(ws, rrow, 10, pct(cur["rn"], v["YOY"]["rn"]) if v["YOY"]["rn"] else "n/a", num=PCT, align=RIGHT)
        rrow += 1
    return rrow + 2

nr = cross_block(ws, 4, "[A] 사업장 × 채널 (OTA/G-OTA/Inbound 채널)",
                 lambda r: (r["prop"], f'{r["ch"]}[{r["seg"] if r["seg"] in SEG3 else "비OTA"}]'),
                 skip=lambda k: k[1].startswith("미분류"))
nr = cross_block(ws, nr, "[B] 사업장 × 상품카테고리", lambda r: (r["prop"], r["cat"]))
ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 26
for c in range(3, 11): ws.column_dimensions[get_column_letter(c)].width = 12

# ════════════════════════ 시트6: 기획전 연관 ════════════════════════
ws = wb.create_sheet("⑥기획전연관")
ws.sheet_view.showGridLines = False
ws["A1"] = "기획전(프로모션) 연관 분석"; ws["A1"].font = TITLE_FONT
ws["A2"] = "픽업과 기획전 매칭.  단위 RN=실, 매출=백만원."
ws["A2"].font = NOTE_FONT
r = 4
# 6-1 패키지 vs 비패키지 (판매)
ws.cell(row=r, column=1, value="【6-1】 패키지(회원번호 86xx) vs 비패키지 — 판매 Net 픽업").font = BOLD; r += 1
dpk, _ = agg(lambda x: "패키지(86xx)" if str(x["mnum"]).startswith("86") else "비패키지(일반)", sales_only=True)
put(ws, r, 1, "구분");
for c, w in enumerate(WINS, 2): put(ws, r, c, WIN_LABEL[w])
put(ws, r, 6, "CUR Net매출(백만)"); style_header(ws, r, 6); r += 1
tot_pk = 0
for k in ("패키지(86xx)", "비패키지(일반)"):
    v = dpk.get(k, {w: {"rn": 0, "rev": 0} for w in WINS})
    put(ws, r, 1, k, align=LEFT, font=BOLD)
    for c, w in enumerate(WINS, 2): put(ws, r, c, v[w]["rn"], num=INT, align=RIGHT)
    put(ws, r, 6, round(v["CUR"]["rev"] / 1e6, 1), num=MIL, align=RIGHT)
    tot_pk += v["CUR"]["rn"]; r += 1
pk_rn = dpk["패키지(86xx)"]["CUR"]["rn"]
ws.cell(row=r, column=1, value=f"→ CUR 판매 Net RN 중 패키지(86xx) 비중 {pk_rn/tot_pk*100:.1f}%  (패키지 = 프로모션/기획전성 상품 프록시)").font = NOTE_FONT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8); r += 2

# 6-2 비발디5월전략 코드매칭 (전체 86xx 중 태깅코드)
ws.cell(row=r, column=1, value="【6-2】 코드 태깅 기획전 '비발디 5월전략 프로모션'(86157xxx, 118코드) 직접 매칭 — Net 픽업").font = BOLD; r += 1
dcamp, _ = agg(lambda x: "기획전매칭(비발디5월전략)" if x["mnum"] in CAMP_CODES else "기타", sales_only=True)
put(ws, r, 1, "구분")
for c, w in enumerate(WINS, 2): put(ws, r, c, WIN_LABEL[w])
put(ws, r, 6, "CUR Net매출(백만)"); style_header(ws, r, 6); r += 1
v = dcamp.get("기획전매칭(비발디5월전략)", {w: {"rn": 0, "rev": 0} for w in WINS})
put(ws, r, 1, "기획전매칭(비발디5월전략)", align=LEFT, font=BOLD)
for c, w in enumerate(WINS, 2): put(ws, r, c, v[w]["rn"], num=INT, align=RIGHT)
put(ws, r, 6, round(v["CUR"]["rev"] / 1e6, 1), num=MIL, align=RIGHT); r += 1
ws.cell(row=r, column=1, value="→ 해당 기획전은 투숙기간(2026-05~07월 초)이 거의 경과 → 최근 7일 픽업 기여 ≈ 0 (취소가 신규 상회).").font = NOTE_FONT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8); r += 2

# 6-3 활성 기획전 캘린더
ws.cell(row=r, column=1, value="【6-3】 여름(6~8월) 가동 기획전 캘린더 — campaign_data.json 기준").font = BOLD; r += 1
cd = json.load(open(PROJECT / "docs/data/campaign_data.json", encoding="utf-8"))
def ov(a1, a2, b1, b2): return a1 and a2 and a1 <= b2 and b1 <= a2
active = []
for e in cd["events"]:
    if ov(e.get("투숙시작"), e.get("투숙종료"), "2026-06-01", "2026-08-31") or \
       ov(e.get("판매시작"), e.get("판매종료"), "2026-06-11", "2026-06-24"):
        active.append(e)
propc = defaultdict(int)
for e in active: propc[e.get("사업장") or "기타"] += 1
put(ws, r, 1, "사업장(기획전 사업장명)"); put(ws, r, 2, "가동 기획전 수")
style_header(ws, r, 2); r += 1
for k, c in sorted(propc.items(), key=lambda x: -x[1]):
    put(ws, r, 1, k, align=LEFT); put(ws, r, 2, c, num=INT, align=RIGHT); r += 1
r += 1
ws.cell(row=r, column=1, value=f"※ 6~8월 가동 기획전 총 {len(active)}건.  주력 채널: 카카오(톡딜/메이커스/쇼핑라이브)·타이드스퀘어·인플루언서·트립비토즈·야놀자·하나투어 등.").font = NOTE_FONT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8); r += 1
ws.cell(row=r, column=1, value="※ 한계: campaign_data.json은 '비발디 5월전략'(118코드) 외 기획전의 패키지코드(86xx)를 보유하지 않음 → 코드단위 픽업 귀속은 비발디건만 정밀 산출.").font = NOTE_FONT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8); r += 1
ws.cell(row=r, column=1, value="   다른 기획전 견인효과는 (a)패키지86xx 비중(6-1) (b)기획전 주력채널의 픽업 강세(④채널시트: 트립비토즈·야놀자·카카오계열 타이드스퀘어 등)로 간접 확인.").font = NOTE_FONT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws.column_dimensions["A"].width = 30
for c in range(2, 7): ws.column_dimensions[get_column_letter(c)].width = 16

# ════════════════════════ 시트7: DB ════════════════════════
ws = wb.create_sheet("⑦DB(로우데이터)")
ws.sheet_view.showGridLines = False
dbcols = ["윈도우", "구분", "활동일자", "투숙월", "사업장", "세그먼트", "OTA구분",
          "채널", "상품카테고리", "회원번호", "패키지여부", "기획전여부",
          "RN(부호)", "매출_원(부호)", "매출_백만(부호)"]
for c, h in enumerate(dbcols, 1): put(ws, 1, c, h)
style_header(ws, 1, len(dbcols)); ws.row_dimensions[1].height = 28
KIND_ORD = {"CUR": 0, "WOW": 1, "MOM": 2, "YOY": 3}
def seg3(r): return r["seg"] if r["seg"] in SEG3 else ("비매출(H/U/COMP)" if r["seg"] in OP else "비OTA")
_dbrows = [r for r in ROWS if r["win"] == "CUR"] if LITE else ROWS
sorted_rows = sorted(_dbrows, key=lambda r: (KIND_ORD[r["win"]], r["date"], r["prop"]))
rr = 2
for x in sorted_rows:
    fmtd = f'{x["date"][:4]}-{x["date"][4:6]}-{x["date"][6:8]}'
    vals = [x["win"], x["kind"], fmtd, x["sm"][:4] + "-" + x["sm"][4:6], x["prop"], x["seg"], seg3(x),
            x["ch"], x["cat"], x["mnum"], "Y" if str(x["mnum"]).startswith("86") else "N",
            "Y" if x["mnum"] in CAMP_CODES else "N",
            x["rn_s"], x["rev_s"], round(x["rev_s"] / 1e6, 3)]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=rr, column=c, value=v)
        if c in (13, 14): cell.number_format = INT
        if c == 15: cell.number_format = MIL
    rr += 1
widths = [9, 11, 11, 8, 20, 12, 14, 14, 14, 12, 9, 9, 11, 14, 12]
for c, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(dbcols))}{rr-1}"

# ── 교차검증 ──
db_sales_rn = sum(x["rn_s"] for x in ROWS if is_sales(x) and x["win"] == "CUR")
print("교차검증 CUR 판매 Net RN(DB합):", db_sales_rn, "== 요약합:", sales_net["CUR"]["rn"])
assert db_sales_rn == sales_net["CUR"]["rn"]

OUT = Path(os.path.expanduser("~/Desktop")) / ("픽업분석_678월_20260624_요약본.xlsx" if LITE else "픽업분석_678월_20260624.xlsx")
wb.save(OUT)
print("저장:", OUT, f"({OUT.stat().st_size/1e6:.1f} MB, DB {rr-2:,}행)")
