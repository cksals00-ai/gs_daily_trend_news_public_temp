#!/usr/bin/env python3
"""
generate_otb_data.py
BI 26OTB 시트 기준으로 docs/data/otb_data.json 생성
- 사업장 순서: PROPERTY_DEFS (BI 시트 순서)
- 실적: db_aggregated.json (booking_rn, booking_rev, adr)
- 목표: budget XLSX Grand Total
- 전년 실적: db_aggregated.json 2025년 동월
- 월별 분리 데이터 포함 (월 필터 작동용)
"""
import calendar
import json
import sys
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"
DOCS_DATA_DIR = PROJECT_DIR / "docs" / "data"
BUDGET_XLSX = DATA_DIR / "raw_db" / "budget" / "★최종★(검토완료)_2026년 객실 사업계획_총량 수립(2차+사업장변경건).xlsx"
DB_JSON = DATA_DIR / "db_aggregated.json"
RM_FCST_JSON = DATA_DIR / "rm_fcst.json"
FCST_SEG_TREND_JSON = DATA_DIR / "fcst_segment_trend.json"  # 사업장×월×세그 RM FCST 분배본
AI_FCST_JSON = DOCS_DATA_DIR / "ai_fcst.json"               # manager_keyin_segments 포함
HOLIDAYS_KR_JSON = DATA_DIR / "holidays_kr.json"
DAILY_BOOKING_JSON = DATA_DIR / "daily_booking.json"
OUTPUT_JSON = DOCS_DATA_DIR / "otb_data.json"

# Budget/RM FCST 비교 base 세그먼트 (rns_budget이 이 3개 세그 합으로 정의됨)
BUDGET_SEGS = ('OTA', 'G-OTA', 'Inbound')

KST = timezone(timedelta(hours=9))

def _calc_target_months_tuple():
    """매월 2일부터 다음 3개월로 롤링. 1일은 전월 마감 실적 확인용."""
    now = datetime.now(KST)
    base_month = now.month if now.day >= 2 else (now.month - 1 if now.month > 1 else 12)
    months = []
    for i in range(3):
        m = base_month + i
        if m > 12: m -= 12
        months.append(m)
    return tuple(months)

TARGET_MONTHS = _calc_target_months_tuple()

# BI 26OTB 시트 사업장 순서 (compare_and_update.py PROPERTY_DEFS 동일)
PROPERTY_DEFS = [
    ("소노벨 비발디파크(A,B,C,호텔)",    "01.벨비발디",      "vivaldi", ["소노벨 비발디파크", "소노문 비발디파크"]),
    ("소노캄 비발디파크 (D동)",           "02.캄비발디",      "vivaldi", ["소노캄 비발디파크"]),
    ("소노펫 (D동+E동)",                  "03.펫비발디",      "vivaldi", ["소노펫 비발디파크"]),
    ("소노펠리체 비발디파크",             "04.펠리체비발디",  "vivaldi", ["소노펠리체 비발디파크"]),
    ("소노펠리체 빌리지 비발디파크",      "05.빌리지비발디",  "vivaldi", ["소노펠리체 빌리지 비발디파크"]),
    ("소노벨양평",                        "06.양평",          "central", ["소노휴 양평", "소노벨 양평"]),
    ("델피노",                            "07.델피노",        "central", ["델피노"]),
    ("쏠비치양양",                        "08.쏠비치양양",    "central", ["쏠비치 양양"]),
    ("쏠비치삼척",                        "09.쏠비치삼척",    "central", ["쏠비치 삼척"]),
    ("소노벨단양",                        "10.소노벨단양",    "central", ["소노문 단양", "소노벨 단양"]),
    ("소노캄경주",                        "11.소노캄경주",    "south",   ["소노벨 경주", "소노캄 경주"]),
    ("소노벨청송",                        "12.소노벨청송",    "central", ["소노벨 청송"]),
    ("소노벨천안",                        "13.소노벨천안",    "central", ["소노벨 천안"]),
    ("소노벨변산",                        "14.소노벨변산",    "central", ["소노벨 변산"]),
    ("소노캄여수",                        "15.소노캄여수",    "south",   ["소노캄 여수"]),
    ("소노캄 거제",                       "16.소노캄거제",    "south",   ["소노캄 거제"]),
    ("쏠비치 진도",                       "17.쏠비치진도",    "south",   ["쏠비치 진도"]),
    ("소노벨 제주",                       "18.소노벨제주",    "apac",    ["소노벨 제주"]),
    ("소노캄 제주",                       "19.소노캄제주",    "apac",    ["소노캄 제주"]),
    ("소노캄 고양",                       "20.소노캄고양",    "apac",    ["소노캄 고양"]),
    ("소노문 해운대",                     "21.소노문해운대",  "south",   ["소노문 해운대"]),
    ("쏠비치 남해",                       "22.쏠비치남해",    "south",   ["쏠비치 남해"]),
    ("르네블루 바이 쏠비치",              "23.르네블루",      "central", ["르네블루"]),
    ("팔라티움 해운대",                   "25.팔라티움",      "south",   []),  # 온북 DB 미포함 — daily_booking.json에서 보정
]

BUDGET_GRAND_TOTAL_ROWS = [26, 50, 74, 98, 122, 146, 170, 194, 218, 242, 266, 290]
BUDGET_COL_RN  = 6
BUDGET_COL_ADR = 8
BUDGET_COL_REV = 10  # 千원 단위

# 세그먼트별 행 오프셋 (Grand Total 행 기준 상대적 위치) — 예산이 있는 세그먼트
BUDGET_SEGMENT_KEYS = ["OTA", "G-OTA", "Inbound"]
SEGMENT_KEYS = ["OTA", "G-OTA", "Inbound"]  # 초기값, main()에서 DB 기반 동적 확장
SEGMENT_ROW_OFFSETS = {"OTA": -6, "G-OTA": -5, "Inbound": -9}

MONTHS_26 = ["202601","202602","202603","202604","202605","202606",
             "202607","202608","202609","202610","202611","202612"]
MONTHS_25 = ["202501","202502","202503","202504","202505","202506",
             "202507","202508","202509","202510","202511","202512"]
MONTH_LABEL = {
    "202601":"1월","202602":"2월","202603":"3월","202604":"4월",
    "202605":"5월","202606":"6월","202607":"7월","202608":"8월",
    "202609":"9월","202610":"10월","202611":"11월","202612":"12월",
}
BUDGET_MONTH_LABEL = ["1월","2월","3월","4월","5월","6월","7월","8월","9월","10월","11월","12월"]

# 리드타임 버킷 중간값 (가중평균 계산용)
LEAD_TIME_MIDPOINTS = {
    'same_day': 0, '1_3d': 2, '4_7d': 5.5, '1_2w': 10.5,
    '2_4w': 21, '1_2m': 45, '2m_plus': 75,
}


def load_budget(wb):
    """사업장별 월별 budget 및 세그먼트별 budget 반환
    Returns:
        budgets:     {display_name: {"1월": {rn, adr, rev_m}, ...}}
        seg_budgets: {display_name: {seg: {"1월": {rn, adr, rev_m}, ...}, ...}}
    """
    budgets = {}
    seg_budgets = {}
    for sheet_name, display_name, region, db_props in PROPERTY_DEFS:
        if sheet_name not in wb.sheetnames:
            budgets[display_name] = {}
            seg_budgets[display_name] = {s: {} for s in SEGMENT_KEYS}
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        monthly = {}
        monthly_seg = {s: {} for s in SEGMENT_KEYS}
        for i, (row_idx, m_label) in enumerate(zip(BUDGET_GRAND_TOTAL_ROWS, BUDGET_MONTH_LABEL)):
            r = rows[row_idx - 1]
            rn  = float(r[BUDGET_COL_RN  - 1] or 0)
            adr = float(r[BUDGET_COL_ADR - 1] or 0)
            rev_cheonwon = float(r[BUDGET_COL_REV - 1] or 0)
            monthly[m_label] = {
                "rn":  int(round(rn)),
                "adr": round(adr),
                "rev_m": round(rev_cheonwon / 1000, 2),  # 千원 → 百万원
            }
            # 세그먼트별 행 읽기
            for seg, offset in SEGMENT_ROW_OFFSETS.items():
                seg_r = rows[(row_idx + offset) - 1]
                seg_rn  = float(seg_r[BUDGET_COL_RN  - 1] or 0)
                seg_adr = float(seg_r[BUDGET_COL_ADR - 1] or 0)
                seg_rev = float(seg_r[BUDGET_COL_REV - 1] or 0)
                monthly_seg[seg][m_label] = {
                    "rn":    int(round(seg_rn)),
                    "adr":   round(seg_adr),
                    "rev_m": round(seg_rev / 1000, 2),
                }
        budgets[display_name] = monthly
        seg_budgets[display_name] = monthly_seg
    return budgets, seg_budgets


def sum_db(db_bp, prop_names, month_key):
    total_rn  = 0
    total_rev = 0.0
    for pname in prop_names:
        m = db_bp.get(pname, {}).get(month_key, {})
        total_rn  += m.get("booking_rn",  0)
        total_rev += m.get("booking_rev", 0.0)
    adr = round((total_rev * 1_000_000) / total_rn) if total_rn > 0 else 0
    return {"rn": total_rn, "rev_m": round(total_rev, 2), "adr": adr}


def sum_db_segments(db_bps, prop_names, month_key):
    """by_property_segment에서 OTA+G-OTA+Inbound만 합산 (예산 세그먼트 기준)"""
    total_rn  = 0
    total_rev = 0.0
    for pname in prop_names:
        prop_segs = db_bps.get(pname, {})
        for seg in BUDGET_SEGMENT_KEYS:
            m = prop_segs.get(seg, {}).get(month_key, {})
            total_rn  += m.get("booking_rn",  0)
            total_rev += m.get("booking_rev", 0.0)
    adr = round((total_rev * 1_000_000) / total_rn) if total_rn > 0 else 0
    return {"rn": total_rn, "rev_m": round(total_rev, 2), "adr": adr}


def sum_seg_budget(seg_budgets, display_name, bud_labels):
    """seg_budgets에서 OTA+G-OTA+Inbound 합산 → budget rn/rev/adr"""
    rn  = 0
    rev = 0.0
    for seg in SEGMENT_KEYS:
        prop_seg = seg_budgets.get(display_name, {}).get(seg, {})
        rn  += sum(prop_seg.get(l, {}).get("rn",    0)    for l in bud_labels)
        rev += sum(prop_seg.get(l, {}).get("rev_m", 0.0)  for l in bud_labels)
    adr = round(rev * 1_000_000 / rn) if rn > 0 else 0
    return {"rn": rn, "rev_m": rev, "adr": adr}


def _load_local_holidays():
    """holidays_kr.json에서 공휴일 데이터 로드. {date_str: {name, type}} 형식."""
    if HOLIDAYS_KR_JSON.exists():
        try:
            data = json.loads(HOLIDAYS_KR_JSON.read_text(encoding="utf-8"))
            return data.get("holidays", {})
        except Exception:
            pass
    return {}


def _count_weekday_holidays_local(holidays_dict, year, month):
    """holidays_kr.json 기반 평일 공휴일 수 계산."""
    from datetime import date as _date
    count = 0
    prefix = f"{year}{month:02d}"
    for date_str, info in holidays_dict.items():
        if not date_str.startswith(prefix):
            continue
        try:
            d = _date(int(date_str[:4]), int(date_str[4:6]), int(date_str[6:8]))
            if d.weekday() < 5:  # 월~금
                count += 1
        except Exception:
            pass
    return count


def _count_weekends_in_month(year, month):
    """해당 월의 토요일·일요일 수 반환 (토, 일 각각)."""
    from datetime import date as _date
    days = calendar.monthrange(year, month)[1]
    sat = sun = 0
    for d in range(1, days + 1):
        wd = _date(year, month, d).weekday()
        if wd == 5:
            sat += 1
        elif wd == 6:
            sun += 1
    return sat, sun


def _count_consecutive_holidays(holidays_dict, year, month):
    """해당 월에서 연휴 블록(공휴일+주말 연속)의 최대 길이와 총 연휴일수 반환.
    Returns: (max_consecutive, total_holiday_days, holiday_blocks)
      holiday_blocks: [(start_day, end_day, length), ...]
    """
    from datetime import date as _date, timedelta as _td
    days_in_month = calendar.monthrange(year, month)[1]
    # 해당 월의 "쉬는 날" 세트 (주말 + 공휴일)
    off_days = set()
    for d in range(1, days_in_month + 1):
        dt = _date(year, month, d)
        ds = f"{year}{month:02d}{d:02d}"
        if dt.weekday() >= 5 or ds in holidays_dict:
            off_days.add(d)
    # 연속 블록 추출
    blocks = []
    if off_days:
        sorted_days = sorted(off_days)
        start = sorted_days[0]
        prev = start
        for d in sorted_days[1:]:
            if d == prev + 1:
                prev = d
            else:
                blocks.append((start, prev, prev - start + 1))
                start = d
                prev = d
        blocks.append((start, prev, prev - start + 1))
    max_consec = max((b[2] for b in blocks), default=0)
    return max_consec, len(off_days), blocks


def build_holiday_factors(target_months=None, cur_year=2026, base_year=2025):
    """올해 vs 작년 평일 공휴일 차이 + 주말 분포 + 연휴 길이 → FCST 보정 계수.

    보정 요소:
    1) 평일 공휴일 수 차이: 1일당 +3%
    2) 토요일 수 차이: 1일당 +2.5% (리조트 특성상 토요일 영향 큼)
    3) 연휴 블록 길이 보너스: 3일 연휴 → +1%, 5일+ → +3%, 9일+ → +5%
    4) 대체공휴일 가산: 대체공휴일 1건당 +1.5% 추가
    ±15% 캡 적용.
    """
    if target_months is None:
        target_months = tuple(range(1, 13))
    holidays_dict = _load_local_holidays()
    factors = {}
    for m in target_months:
        # (1) 평일 공휴일 차이
        hol_delta = (_count_weekday_holidays_local(holidays_dict, cur_year, m)
                   - _count_weekday_holidays_local(holidays_dict, base_year, m))
        adj = hol_delta * 0.03

        # (2) 토요일 수 차이
        sat_cur, _ = _count_weekends_in_month(cur_year, m)
        sat_base, _ = _count_weekends_in_month(base_year, m)
        adj += (sat_cur - sat_base) * 0.025

        # (3) 연휴 블록 길이 보너스 (올해 vs 작년 비교)
        max_c_cur, _, _ = _count_consecutive_holidays(holidays_dict, cur_year, m)
        max_c_base, _, _ = _count_consecutive_holidays(holidays_dict, base_year, m)
        def _block_bonus(max_c):
            if max_c >= 9: return 0.05
            if max_c >= 5: return 0.03
            if max_c >= 3: return 0.01
            return 0.0
        adj += _block_bonus(max_c_cur) - _block_bonus(max_c_base)

        # (4) 대체공휴일 가산
        prefix_cur = f"{cur_year}{m:02d}"
        prefix_base = f"{base_year}{m:02d}"
        sub_cur = sum(1 for ds, info in holidays_dict.items()
                      if ds.startswith(prefix_cur) and "대체" in info.get("name", ""))
        sub_base = sum(1 for ds, info in holidays_dict.items()
                       if ds.startswith(prefix_base) and "대체" in info.get("name", ""))
        adj += (sub_cur - sub_base) * 0.015

        factors[m] = round(max(0.85, min(1.15, 1.0 + adj)), 4)
    return factors


def get_adj_rn(adj_by_prop, prop_names, month_key):
    total = 0
    for pname in prop_names:
        total += adj_by_prop.get(pname, {}).get(month_key, {}).get("adjustment_rn", 0)
    return total


def _calc_fcst(act_rn, act_rev, month_num, now_kst, bud_rn, bud_rev=0,
               ly_pickup_ratio=None):
    """월말 FCST 계산 (OTB pickup 모델 기반).

    - 과거 월: 실적=FCST
    - 미래 월: None 반환 (호출자가 fallback 처리)
    - 진행 중인 월: act_rn × ly_pickup_ratio
        ly_pickup_ratio = LY 풀년 실적 / LY 동기간(base_date) OTB
        OTB는 stay-date 기준이라 elapsed_ratio 외삽이 불가 (pickup으로 채워짐).
        ly_pickup_ratio 부재 시 보수적으로 act_rn 그대로 사용.
    """
    cur_year  = now_kst.year
    cur_month = now_kst.month

    if cur_year == 2026 and month_num == cur_month:
        days_in_month = calendar.monthrange(cur_year, month_num)[1]
        elapsed_ratio = now_kst.day / days_in_month
    elif month_num < cur_month or cur_year < 2026:
        elapsed_ratio = 1.0
    else:
        elapsed_ratio = 0.0

    if elapsed_ratio == 0.0:
        return None, None, None, None, None

    if elapsed_ratio >= 1.0:
        rns_fcst = act_rn
        rev_fcst = act_rev
    elif ly_pickup_ratio and ly_pickup_ratio > 0:
        rns_fcst = round(act_rn * ly_pickup_ratio)
        rev_fcst = act_rev * ly_pickup_ratio
    else:
        # LY pickup 데이터 부재 — act_rn을 보수적으로 fcst로 사용
        rns_fcst = act_rn
        rev_fcst = act_rev

    adr_fcst     = round(rev_fcst * 1_000_000 / rns_fcst) if rns_fcst > 0 else 0
    fcst_ach     = round(rns_fcst / bud_rn  * 100, 1) if bud_rn  > 0 else 0.0
    rev_fcst_ach = round(rev_fcst / bud_rev * 100, 1) if bud_rev > 0 else 0.0

    return rns_fcst, rev_fcst, adr_fcst, fcst_ach, rev_fcst_ach


def _compute_ly_pickup_ratio(adj_by_prop, db_bps, db_bp, db_props, last_keys):
    """LY pickup ratio = LY 풀년 실적 / LY 동기간(base_date) OTB.

    OTA+G-OTA+Inbound 세그먼트 기준 (parse_yoy_bookings 필터와 일치).
    pname별 adj_by_prop[pname][mk]['booking_rn'] = LY base_date 시점 OTB.
    db_bps에서 동일 사업장+세그 합산 = LY 풀년 최종 실적.

    Returns: float ratio (>1.0 일반적: 추가 pickup 발생) 또는 None.
    """
    if not adj_by_prop:
        return None
    ly_full = 0
    ly_otb_at_base = 0
    for mk in last_keys:
        # LY 풀년 실적 (3 budget 세그)
        if db_bps is not None:
            ly_full += sum_db_segments(db_bps, db_props, mk)["rn"]
        else:
            ly_full += sum_db(db_bp, db_props, mk)["rn"]
        # LY base_date OTB (3 budget 세그 — parse_yoy_bookings에서 필터됨)
        for pname in db_props:
            ly_otb_at_base += adj_by_prop.get(pname, {}).get(mk, {}).get("booking_rn", 0)
    if ly_otb_at_base <= 0 or ly_full <= 0:
        return None
    return ly_full / ly_otb_at_base


def _detect_outlier_years(hist, threshold=0.4):
    """과거 실적에서 이상치 연도 감지 (코로나 등).
    중앙값 대비 ±threshold 이상 벗어나면 이상치로 판정.
    Returns: {year: weight} — 이상치는 낮은 가중치, 정상은 1.0.
    """
    if len(hist) < 3:
        return {y: 1.0 for y in hist}
    vals = sorted(hist.values())
    median = vals[len(vals) // 2]
    weights = {}
    for y, v in hist.items():
        if median > 0:
            deviation = abs(v - median) / median
            if deviation > threshold:
                weights[y] = 0.2  # 이상치 → 가중치 대폭 축소
            else:
                weights[y] = 1.0
        else:
            weights[y] = 1.0
    return weights


def _time_weighted_avg(hist, outlier_weights=None):
    """시간 가중 이동평균. 최근 연도에 높은 가중치 부여.
    2022→1, 2023→2, 2024→3, 2025→4 (선형 가중).
    outlier_weights로 이상치 연도 가중치 추가 축소.
    Returns: weighted average RN.
    """
    if not hist:
        return 0
    years = sorted(hist.keys())
    min_yr = min(years)
    w_total = 0.0
    v_total = 0.0
    for y in years:
        time_w = (y - min_yr + 1)  # 2022→1, 2023→2, ...
        ow = (outlier_weights or {}).get(y, 1.0)
        w = time_w * ow
        w_total += w
        v_total += hist[y] * w
    return v_total / w_total if w_total > 0 else 0


def _calc_seasonality_share(db_bp, db_bps, db_props, year):
    """특정 연도의 월별 비중 패턴 계산.
    Returns: {month_num: share} (12개월 합=1.0)
    """
    monthly = {}
    total = 0
    for m in range(1, 13):
        mk = f"{year}{m:02d}"
        if db_bps is not None:
            rn = sum_db_segments(db_bps, db_props, mk)["rn"]
        else:
            rn = sum_db(db_bp, db_props, mk)["rn"]
        monthly[m] = rn
        total += rn
    if total > 0:
        return {m: monthly[m] / total for m in range(1, 13)}
    return {m: 1.0 / 12 for m in range(1, 13)}


def _calc_fcst_enhanced(act_rn, month_num, now_kst, bud_rn, db_bp, db_props,
                         holiday_factors=None, db_bps=None,
                         rm_trend_snapshots=None, ly_pickup_ratio=None):
    """고도화된 AI FCST — 8가지 보정 로직 적용.

    핵심 개선:
    1) 요일 분포 보정: build_holiday_factors에서 토요일 수 차이 반영 (외부)
    2) 세그먼트별 트렌드 분화: OTA/G-OTA/Inbound 각각 성장률 계산 후 합산
    3) 계절성 패턴: 4개년 월별 비중 패턴 → 연간 추정치 × 비중
    4) 시간 가중 이동평균: 2025 > 2024 > 2023 > 2022 순 가중치
    5) 이상치 필터링: 코로나(2022) 등 비정상 연도 자동 감지 + 가중치 축소
    6) Confidence Interval: 과거 4년 분산 기반 상한/하한 (fcst_lo, fcst_hi)
    7) 리드타임 반영: RM 스냅샷에서 pickup 패턴 활용 (미래월)
    8) 연휴 길이 세분화: build_holiday_factors에서 연휴 블록 길이 반영 (외부)

    Returns: (fcst_rn, fcst_ach, fcst_lo, fcst_hi)
    """
    import math

    cur_year = now_kst.year
    cur_month = now_kst.month
    day = now_kst.day

    # ── elapsed 판별 ──
    if cur_year == 2026 and month_num == cur_month:
        days_in_month = calendar.monthrange(cur_year, month_num)[1]
        elapsed_ratio = day / days_in_month
    elif month_num < cur_month or cur_year < 2026:
        elapsed_ratio = 1.0
    else:
        elapsed_ratio = 0.0

    # 완료 월 → 실적 그대로
    if elapsed_ratio == 1.0:
        fcst_ach = round(act_rn / bud_rn * 100, 1) if bud_rn > 0 else 0.0
        return act_rn, fcst_ach, act_rn, act_rn

    # ── helper: 3세그먼트 합산 실적 ──
    def _seg_rn(props, mk):
        if db_bps is not None:
            return sum_db_segments(db_bps, props, mk)["rn"]
        return sum_db(db_bp, props, mk)["rn"]

    def _seg_rn_by_segment(props, mk, seg):
        """단일 세그먼트 실적 조회."""
        if db_bps is not None:
            total = 0
            for pname in props:
                total += db_bps.get(pname, {}).get(seg, {}).get(mk, {}).get("booking_rn", 0)
            return total
        return 0

    # ── [5] 이상치 필터링 + [4] 시간 가중 과거 데이터 수집 ──
    hist = {}  # {year: rn}
    for yr in [2022, 2023, 2024, 2025]:
        mk = f"{yr}{month_num:02d}"
        rn = _seg_rn(db_props, mk)
        if rn > 0:
            hist[yr] = rn

    outlier_w = _detect_outlier_years(hist)
    tw_avg = _time_weighted_avg(hist, outlier_w)

    # ── 방법 1: 단순 OTB pickup ratio 적용 ──
    # OTB는 stay-date 기준 → 일별 외삽 불가 (act_rn / elapsed_ratio는 폭주).
    # ly_pickup_ratio = LY 풀년 / LY 동기간 OTB. 부재 시 act_rn 보수적 사용.
    if ly_pickup_ratio and ly_pickup_ratio > 0 and elapsed_ratio < 1.0:
        fcst_simple = round(act_rn * ly_pickup_ratio)
    else:
        fcst_simple = act_rn

    # ── [2] 세그먼트별 트렌드 분화 ──
    segments = ["OTA", "G-OTA", "Inbound"]
    seg_fcst_total = 0
    seg_has_data = False

    if db_bps is not None and elapsed_ratio == 0:
        # 미래월: 세그먼트별로 개별 YoY 계산
        for seg in segments:
            seg_hist = {}
            for yr in [2022, 2023, 2024, 2025]:
                mk = f"{yr}{month_num:02d}"
                srn = _seg_rn_by_segment(db_props, mk, seg)
                if srn > 0:
                    seg_hist[yr] = srn

            if not seg_hist:
                continue

            seg_outlier_w = _detect_outlier_years(seg_hist)
            seg_last_yr = seg_hist.get(2025, 0)

            # 세그먼트별 최근 완료 2개월 YoY
            seg_yoy_ratios = []
            for lookback in range(1, 4):
                rm = cur_month - lookback
                if rm < 1:
                    rm += 12
                r26 = _seg_rn_by_segment(db_props, f"2026{rm:02d}", seg)
                r25 = _seg_rn_by_segment(db_props, f"2025{rm:02d}", seg)
                if r25 > 0 and r26 > 0:
                    seg_yoy_ratios.append(r26 / r25)
                if len(seg_yoy_ratios) >= 2:
                    break

            if seg_last_yr > 0 and seg_yoy_ratios:
                avg_yoy = sum(seg_yoy_ratios) / len(seg_yoy_ratios)
                seg_fcst_total += round(seg_last_yr * avg_yoy)
                seg_has_data = True
            elif seg_last_yr > 0:
                seg_tw = _time_weighted_avg(seg_hist, seg_outlier_w)
                if seg_tw > 0 and seg_hist.get(2025, 0) > 0:
                    trend = seg_hist[2025] / seg_tw
                    seg_fcst_total += round(seg_last_yr * max(0.8, min(1.3, trend)))
                else:
                    seg_fcst_total += seg_last_yr
                seg_has_data = True

    # ── 방법 2: 과거 패턴 기반 (통합) ──
    mk_25 = f"2025{month_num:02d}"
    last_yr_rn = hist.get(2025, 0)
    if last_yr_rn == 0:
        last_yr_rn = _seg_rn(db_props, mk_25)

    if elapsed_ratio > 0 and hist:
        # 진행 중인 월: OTB pickup ratio 기반 (elapsed_ratio 외삽 금지 — OTB 폭주 방지)
        if ly_pickup_ratio and ly_pickup_ratio > 0 and act_rn > 0:
            fcst_pattern = round(act_rn * ly_pickup_ratio)
        elif last_yr_rn > 0 and act_rn > 0:
            # pickup 비율 부재 fallback — LY 풀년 × YoY (일별 외삽 X)
            # YoY 추정 불가 시 LY 풀년 그대로 사용 (보수적)
            fcst_pattern = last_yr_rn
        else:
            fcst_pattern = fcst_simple
    elif elapsed_ratio == 0 and hist:
        # 미래월
        yoy_ratios = []
        for lookback in range(1, 4):
            rm = cur_month - lookback
            if rm < 1:
                rm += 12
            r26 = _seg_rn(db_props, f"2026{rm:02d}")
            r25 = _seg_rn(db_props, f"2025{rm:02d}")
            if r25 > 0 and r26 > 0:
                yoy_ratios.append(r26 / r25)
            if len(yoy_ratios) >= 2:
                break

        if last_yr_rn > 0 and yoy_ratios:
            avg_yoy = sum(yoy_ratios) / len(yoy_ratios)
            fcst_pattern = round(last_yr_rn * avg_yoy)
        elif last_yr_rn > 0:
            # [4] 시간 가중 이동평균 기반 트렌드 적용
            if tw_avg > 0:
                trend = last_yr_rn / tw_avg
                fcst_pattern = round(last_yr_rn * max(0.8, min(1.3, trend)))
            else:
                fcst_pattern = last_yr_rn
        else:
            vals = [hist[y] for y in sorted(hist)]
            if len(vals) >= 2:
                trend = vals[-1] / vals[-2]
                fcst_pattern = round(vals[-1] * trend)
            else:
                fcst_pattern = vals[-1] if vals else 0
    else:
        fcst_pattern = fcst_simple

    # ── [3] 계절성 패턴 보정 (미래월에만 적용) ──
    if elapsed_ratio == 0 and len(hist) >= 2:
        # 4개년 월별 비중 평균 계산
        shares = []
        for yr in sorted(hist.keys()):
            if outlier_w.get(yr, 1.0) >= 0.5:  # 이상치 연도 제외
                s = _calc_seasonality_share(db_bp, db_bps, db_props, yr)
                if s.get(month_num, 0) > 0:
                    shares.append(s[month_num])
        if shares:
            avg_share = sum(shares) / len(shares)
            # 올해 1~(cur_month-1)월 합산으로 연간 추정
            ytd_rn = 0
            ytd_share = 0.0
            for m_past in range(1, cur_month):
                ytd_rn += _seg_rn(db_props, f"2026{m_past:02d}")
                # 평균 비중 합산
                past_shares = []
                for yr in sorted(hist.keys()):
                    if outlier_w.get(yr, 1.0) >= 0.5:
                        s = _calc_seasonality_share(db_bp, db_bps, db_props, yr)
                        past_shares.append(s.get(m_past, 0))
                if past_shares:
                    ytd_share += sum(past_shares) / len(past_shares)

            if ytd_rn > 0 and ytd_share > 0.05:
                annual_est = ytd_rn / ytd_share
                fcst_seasonal = round(annual_est * avg_share)
                # 계절성 FCST를 패턴 FCST와 블렌딩 (30:70)
                fcst_pattern = round(fcst_pattern * 0.7 + fcst_seasonal * 0.3)

    # ── [2] 세그먼트별 분화 결과 블렌딩 (미래월) ──
    if elapsed_ratio == 0 and seg_has_data and seg_fcst_total > 0:
        # 세그먼트별 결과와 통합 결과를 50:50 블렌딩
        fcst_pattern = round(fcst_pattern * 0.5 + seg_fcst_total * 0.5)

    # ── [7] 리드타임 반영 (미래월, RM 스냅샷 기반) ──
    if elapsed_ratio == 0 and rm_trend_snapshots and last_yr_rn > 0:
        # 과거 동월의 RM 스냅샷에서 pickup 패턴 분석
        # D-14 시점 FCST 대비 최종 실적의 평균 uplift 비율 계산
        pickup_ratios = []
        for snap in rm_trend_snapshots:
            snap_date_str = snap.get("_snapshot_date", "")
            snap_year = snap.get("_year", 0)
            if snap_year not in [2025, 2026]:
                continue
            # 스냅샷 월 추출
            try:
                snap_month = int(snap_date_str.split("-")[1])
            except (IndexError, ValueError):
                continue
            # 동월 스냅샷의 FCST vs 실적 비교
            if snap_month == month_num and snap_year == 2025:
                snap_props = snap.get("properties", {})
                for _, display_name, _, dp in PROPERTY_DEFS:
                    if dp != db_props:
                        continue
                    rm_key = f"{snap_year}-{month_num:02d}"
                    rm_entry = snap_props.get(display_name, {}).get(rm_key, {})
                    rm_fcst = rm_entry.get("rm_fcst_rn", 0)
                    if rm_fcst > 0 and last_yr_rn > 0:
                        pickup_ratios.append(last_yr_rn / rm_fcst)
                    break

        if pickup_ratios:
            avg_pickup = sum(pickup_ratios) / len(pickup_ratios)
            # pickup 비율이 1보다 크면 추가 유입 예상
            if 0.8 < avg_pickup < 1.5:
                # 현재 시점 RM FCST가 있으면 그것에 pickup 비율 적용
                fcst_pattern = round(fcst_pattern * min(1.15, max(0.9, avg_pickup)))

    # ── 공휴일/연휴 보정 (요일 분포 + 연휴 길이 포함) ──
    hf = holiday_factors.get(month_num, 1.0) if holiday_factors else 1.0
    fcst_pattern = round(fcst_pattern * hf)

    # ── 가중평균: elapsed 비율이 높을수록 단순외삽 신뢰도 ↑ ──
    if elapsed_ratio > 0.5:
        w_simple = min(elapsed_ratio * 1.2, 0.85)
        fcst_final = round(fcst_simple * w_simple + fcst_pattern * (1.0 - w_simple))
    elif elapsed_ratio > 0.1:
        w_simple = elapsed_ratio
        fcst_final = round(fcst_simple * w_simple + fcst_pattern * (1.0 - w_simple))
    elif elapsed_ratio > 0:
        fcst_final = fcst_pattern
    else:
        fcst_final = fcst_pattern

    # ── [6] Confidence Interval (과거 4년 분산 기반) ──
    if len(hist) >= 2:
        # 시간 가중 분산 계산
        years_sorted = sorted(hist.keys())
        min_yr = min(years_sorted)
        weighted_vals = []
        for y in years_sorted:
            tw = (y - min_yr + 1) * outlier_w.get(y, 1.0)
            weighted_vals.append((hist[y], tw))
        w_sum = sum(w for _, w in weighted_vals)
        w_mean = sum(v * w for v, w in weighted_vals) / w_sum if w_sum > 0 else 0
        w_var = sum(w * (v - w_mean) ** 2 for v, w in weighted_vals) / w_sum if w_sum > 0 else 0
        std_dev = math.sqrt(w_var)

        # 성장 트렌드 적용된 std_dev → 현재 FCST 대비 비율로 변환
        if w_mean > 0:
            cv = std_dev / w_mean  # 변동계수
            fcst_lo = round(fcst_final * (1.0 - cv))
            fcst_hi = round(fcst_final * (1.0 + cv))
        else:
            fcst_lo = fcst_final
            fcst_hi = fcst_final

        # 하한은 0 이하가 되지 않도록
        fcst_lo = max(0, fcst_lo)
    else:
        # 데이터 부족 시 ±10% 기본 구간
        fcst_lo = round(fcst_final * 0.9)
        fcst_hi = round(fcst_final * 1.1)

    fcst_ach = round(fcst_final / bud_rn * 100, 1) if bud_rn > 0 else 0.0
    return fcst_final, fcst_ach, fcst_lo, fcst_hi


def load_rm_fcst():
    """data/rm_fcst.json 로드 → {display_name: {"2026-MM": {rm_fcst_rn, rm_budget_rn}}}"""
    if not RM_FCST_JSON.exists():
        return {}
    try:
        data = json.loads(RM_FCST_JSON.read_text(encoding="utf-8"))
        return data.get("properties", {})
    except Exception:
        return {}


def load_segment_fcst():
    """세그먼트별 FCST 원본 데이터 로드 (분배 X, 원본 그대로).

    우선순위 (각 사업장×월×세그 단위):
      1) manager_keyin_segments[seg]: 매니저가 직접 키인한 값 (ai_fcst.json) — 최우선
      2) fcst_segment_trend snapshot rm_fcst_rn: RM 회의 결과를 4년 가중 historical mix로 분배한 공식 값
      3) ratios[prop][월][seg]: historical mix 점유율 (1~12월 모두 존재) — 미커버 월 fallback용

    Returns: {
        "manager_keyin": {prop: {ym: {seg: rn}}},
        "rm_seg_fcst":   {prop: {ym: {seg: {"rm_fcst_rn", "rm_budget_rn", "ratio"}}}},
        "ratios":        {prop: {month_int_str: {seg: ratio_pct}}},
    }
    """
    out = {"manager_keyin": {}, "manager_keyin_meta": {}, "rm_seg_fcst": {}, "ratios": {}}

    # 1) ai_fcst.json: manager_keyin_segments (테스트 입력은 제외)
    if AI_FCST_JSON.exists():
        try:
            ai = json.loads(AI_FCST_JSON.read_text(encoding="utf-8"))
            for prop, months in ai.get("properties", {}).items():
                for ym, v in months.items():
                    ks = v.get("manager_keyin_segments")
                    if not (ks and isinstance(ks, dict)):
                        continue
                    by = (v.get("manager_keyin_by") or "").strip().lower()
                    if by in ("test", "tester", "demo", ""):
                        # 테스트/미식별 키인은 무시 (안전장치)
                        print(f"  [WARN] manager_keyin 무시 (by='{by}'): {prop} {ym}")
                        continue
                    out["manager_keyin"].setdefault(prop, {})[ym] = ks
                    out["manager_keyin_meta"].setdefault(prop, {})[ym] = {
                        "by": v.get("manager_keyin_by"),
                        "at": v.get("manager_keyin_at"),
                    }
        except Exception:
            pass

    # 2) fcst_segment_trend.json: snapshots[-1] 최신 + ratios
    # New schema: properties[prop][ym] = {rm_fcst_rn, rm_budget_rn, segments: {OTA: {...}, ...}}
    # Old schema: properties[prop][ym] = {OTA: {...}, G-OTA: {...}, ...}
    if FCST_SEG_TREND_JSON.exists():
        try:
            fst = json.loads(FCST_SEG_TREND_JSON.read_text(encoding="utf-8"))
            snaps = fst.get("snapshots", [])
            if snaps:
                latest = snaps[-1]
                for prop, months in latest.get("properties", {}).items():
                    for ym, ym_data in months.items():
                        if isinstance(ym_data, dict) and isinstance(ym_data.get("segments"), dict):
                            segs = ym_data["segments"]
                        else:
                            segs = ym_data
                        out["rm_seg_fcst"].setdefault(prop, {})[ym] = segs
            out["ratios"] = fst.get("ratios", {})
        except Exception:
            pass

    return out


def sum_rm_seg_fcst(rm_fcst_props, display_name, ym):
    """rm_fcst.json segments(OTA/G-OTA/Inbound) PDF 원본 셀 합산.
    Returns: (rm_fcst_rn_sum, rm_budget_rn_sum, rm_fcst_rev_sum). 분배 X.
    rev_sum 단위: 백만원 (PDF 원본 그대로).
    """
    segs = rm_fcst_props.get(display_name, {}).get(ym, {}).get("segments", {})
    if not segs:
        return None, None, None
    fcst_sum = 0
    bud_sum = 0
    rev_sum = 0
    has_fcst = False
    has_bud = False
    has_rev = False
    for seg in ("OTA", "G-OTA", "Inbound"):
        s = segs.get(seg)
        if not isinstance(s, dict):
            continue
        v_f = s.get("rm_fcst_rn")
        v_b = s.get("rm_budget_rn")
        v_r = s.get("rm_fcst_rev_mil")
        if v_f is not None:
            fcst_sum += int(v_f)
            has_fcst = True
        if v_b is not None:
            bud_sum += int(v_b)
            has_bud = True
        if v_r is not None:
            rev_sum += int(v_r)
            has_rev = True
    return (fcst_sum if has_fcst else None, bud_sum if has_bud else None, rev_sum if has_rev else None)


def get_seg_fcst(seg_fcst_data, display_name, ym, seg, p_total_fcst=None, month_int=None,
                 s_bud_rn=0, s_lst_rn=0):
    """단일 (사업장, 월, 세그)에 대한 FCST RN 결정. 우선순위:
      1) manager_keyin_segments[seg]  (단, sanity check 통과 시)
      2) fcst_segment_trend snapshot rm_fcst_rn
      3) ratios[prop][month_int][seg] × p_total_fcst (RM 미커버 월 fallback)
      4) None (데이터 없음)

    Sanity check (manager_keyin):
      값이 max(seg_budget, LY 실적)의 3배를 초과하면 키인 오류로 간주, 무시.
      (테스트/오타 입력 방지)

    Args:
        seg_fcst_data: load_segment_fcst() 결과
        display_name: 사업장 표시명 (예: "01.벨비발디")
        ym: "YYYY-MM" 형식 (예: "2026-07")
        seg: "OTA" / "G-OTA" / "Inbound"
        p_total_fcst: 사업장 총 FCST RN (3순위 fallback에 필요)
        month_int: 월 정수 (1~12). 없으면 ym에서 추출.
        s_bud_rn: 해당 세그 예산 (sanity check용)
        s_lst_rn: 해당 세그 LY 풀년 실적 (sanity check용)
    """
    # 1) manager keyin (sanity check 적용)
    keyin = seg_fcst_data.get("manager_keyin", {}).get(display_name, {}).get(ym, {})
    if seg in keyin and keyin[seg] is not None:
        v = int(keyin[seg])
        sane_cap = max(s_bud_rn, s_lst_rn) * 3
        if sane_cap > 0 and v > sane_cap:
            print(f"  [WARN] manager_keyin 무시 (값 폭주: {v} > {sane_cap}): {display_name} {ym} {seg}")
        else:
            return v, "manager_keyin"

    # 2) RM segment FCST snapshot
    rm_block = seg_fcst_data.get("rm_seg_fcst", {}).get(display_name, {}).get(ym, {})
    if seg in rm_block:
        v = rm_block[seg].get("rm_fcst_rn")
        if v is not None:
            return int(v), "rm_seg_fcst"

    # 3) ratio × property total fcst (RM 미커버 월 fallback)
    if p_total_fcst is not None and p_total_fcst > 0:
        if month_int is None and ym and len(ym) >= 7:
            try:
                month_int = int(ym[5:7])
            except Exception:
                month_int = None
        if month_int is not None:
            ratios = seg_fcst_data.get("ratios", {}).get(display_name, {}).get(str(month_int), {})
            if ratios and seg in ratios:
                pct = ratios.get(seg)
                if pct is not None:
                    return int(round(p_total_fcst * pct / 100)), "ratio_fallback"

    return None, None


def build_ly_same_period_adjusted(db_bp, db_bps, db_seg, now_kst):
    """동기간 기준으로 LY 데이터를 필터링하여 동적 보정 데이터 생성.
    today가 2026-04-29라면, 2025년 데이터는 <= 2025-04-29 기준일 시점의 데이터만 포함.
    {
        "2025": {
            "by_property": {prop_name: {month_key: {booking_rn, booking_rev, adjustment_rn, adjustment_rev}, ...}, ...},
            "by_segment": {seg_name: {month_key: {...}, ...}, ...}
        }
    }
    """
    # 2025년의 동기간 기준일 (today와 같은 월일, 작년)
    ly_cutoff_date = f"{now_kst.year - 1}{now_kst.month:02d}{now_kst.day:02d}"

    # 아직 구현하지 않음 - 실제 booking_date 필터링이 필요하므로 skip
    # 현재 가용 데이터에서는 월 단위 aggregate만 있으므로, 정교한 동기간 필터링 불가
    return {}


def apply_ly_same_period_adjustment(db_bp, db_seg, db_bps, month_idx, now_kst):
    """특정 월에 대해 동기간 LY 보정을 적용 (future month인 경우만).
    Returns: (adj_ly_rn, adj_ly_rev) tuple for the month

    현재 month_idx에 대해:
    - 현재월이거나 미래월이면: 2025년 같은 월의 데이터를 사용하되,
      현재 진행 상황을 반영하기 위해 진행률로 스케일링
    - 과거월이면: 2025년 전체 월 데이터 사용
    """
    if month_idx <= 0 or month_idx > 12:
        return None, None

    cur_month = now_kst.month
    mk_26 = f"2026{month_idx:02d}"
    mk_25 = f"2025{month_idx:02d}"

    # 과거월: 전체 2025 데이터 사용
    if month_idx < cur_month:
        ly_rn = sum_db(db_bp, [p for _, _, _, props in PROPERTY_DEFS for p in props], mk_25)["rn"]
        ly_rev = sum_db(db_bp, [p for _, _, _, props in PROPERTY_DEFS for p in props], mk_25)["rev_m"]
        return ly_rn, ly_rev

    # 현재월 또는 미래월: 진행률 기반 스케일링
    if month_idx == cur_month:
        # 현재월: 현재까지 진행률만큼만 포함
        days_in_month = calendar.monthrange(2026, month_idx)[1]
        elapsed_ratio = now_kst.day / days_in_month

        # 2025년 같은 월의 동기간 데이터 (현재 진행률까지만)
        # 주의: 원본 데이터가 booking_date별 분해 정보가 없으므로 근사치 사용
        # 실제로는 booking_date를 필터링해야 하는데, 월 aggregate 데이터만 있음
        ly_all_rn = sum_db(db_bp, [p for _, _, _, props in PROPERTY_DEFS for p in props], mk_25)["rn"]
        ly_all_rev = sum_db(db_bp, [p for _, _, _, props in PROPERTY_DEFS for p in props], mk_25)["rev_m"]

        # 간단한 근사: 진행률로 2025년 데이터를 스케일
        # (이상적으로는 2025-01-01 ~ 2025-04-29의 booking_date 필터링)
        # 현재 구조에서는 그대로 반환 (동일월의 전체 데이터)
        return ly_all_rn, ly_all_rev

    # 미래월: 전체 2025 데이터 사용 (예측이므로)
    ly_rn = sum_db(db_bp, [p for _, _, _, props in PROPERTY_DEFS for p in props], mk_25)["rn"]
    ly_rev = sum_db(db_bp, [p for _, _, _, props in PROPERTY_DEFS for p in props], mk_25)["rev_m"]
    return ly_rn, ly_rev


def build_yoy_table(db_bp, budgets, seg_budgets, db_bps, adj_by_prop, holiday_factors,
                    months=(4, 5, 6), now_kst=None, rm_fcst_props=None, daily_bk=None,
                    rm_trend_snapshots=None):
    """사업장별 4·5·6월 YoY 추이 테이블 데이터 생성.
    rm_fcst_props: RM FCST 데이터 (load_rm_fcst() 결과)
    daily_bk: 온북 DB 미포함 사업장 보정 데이터
    rm_trend_snapshots: RM FCST 트렌드 스냅샷 (리드타임 반영용)
    """
    if now_kst is None:
        now_kst = datetime.now(KST)
    if rm_fcst_props is None:
        rm_fcst_props = {}
    rows = []
    for sheet_name, display_name, region, db_props in PROPERTY_DEFS:
        month_data = {}
        for m in months:
            mk_26 = f"2026{m:02d}"
            mk_25 = f"2025{m:02d}"
            bud_label = BUDGET_MONTH_LABEL[m - 1]

            act_rn = sum_db(db_bp, db_props, mk_26)["rn"]

            # 전년 보정값 포함
            base_rn = 0
            adj_rn  = 0
            for pname in db_props:
                adj_m = adj_by_prop.get(pname, {}).get(mk_25, {}) if adj_by_prop else {}
                base_rn += adj_m.get("booking_rn", 0)
                adj_rn  += adj_m.get("adjustment_rn", 0)
            if base_rn == 0:
                base_rn = sum_db(db_bp, db_props, mk_25)["rn"]

            if seg_budgets:
                bud_rn = sum_seg_budget(seg_budgets, display_name, [bud_label])["rn"]
            else:
                bud_rn = budgets.get(display_name, {}).get(bud_label, {}).get("rn", 0)

            # OTB pickup ratio (해당 월 기준)
            m_pickup_ratio = _compute_ly_pickup_ratio(
                adj_by_prop, db_bps, db_bp, db_props, [mk_25]
            )

            rns_fcst, _, _, fcst_ach, _ = _calc_fcst(
                act_rn, 0, m, now_kst, bud_rn, ly_pickup_ratio=m_pickup_ratio
            )

            # AI FCST (고도화: 세그먼트 분화 + 계절성 + 이상치 필터 + CI)
            rns_fcst_ai, fcst_ach_ai, fcst_lo, fcst_hi = _calc_fcst_enhanced(
                act_rn, m, now_kst, bud_rn, db_bp, db_props, holiday_factors,
                db_bps=db_bps, rm_trend_snapshots=rm_trend_snapshots,
                ly_pickup_ratio=m_pickup_ratio)

            # RM FCST (OTA+G-OTA+Inbound 세그합 — 대시보드 base와 일치)
            rm_key = f"2026-{m:02d}"
            rm_rn, rm_budget, rm_rev = sum_rm_seg_fcst(rm_fcst_props, display_name, rm_key)
            rm_ach = round(rm_rn / rm_budget * 100, 1) if (rm_rn and rm_budget and rm_budget > 0) else None
            rm_adr = round(rm_rev * 1_000_000 / rm_rn) if (rm_rev and rm_rn and rm_rn > 0) else None

            yoy = round((act_rn / base_rn - 1) * 100, 1) if base_rn > 0 else None
            month_data[m] = {
                "act_rn":       act_rn,
                "last_rn":      base_rn,
                "yoy":          yoy,
                "bud_rn":       bud_rn,
                "rns_fcst":     rns_fcst if rns_fcst else (rns_fcst_ai or act_rn),
                "fcst_ach":     fcst_ach if fcst_ach else (fcst_ach_ai or 0),
                "rns_fcst_ai":  rns_fcst_ai,
                "fcst_ach_ai":  fcst_ach_ai,
                "fcst_lo":      fcst_lo,
                "fcst_hi":      fcst_hi,
                "rm_fcst_rn":   rm_rn,
                "rm_fcst_rev":  rm_rev,
                "rm_fcst_adr":  rm_adr,
                "rm_budget_rn": rm_budget,
                "rm_fcst_ach":  rm_ach,
            }
        # 온북 DB 미포함 사업장: daily_booking 보정
        if daily_bk and display_name in daily_bk and not db_props:
            for m in months:
                bk = daily_bk[display_name].get(m, {})
                if bk:
                    act_rn = bk.get("actual_rns", 0)
                    bud_rn = bk.get("budget_rns", 0)
                    lst_rn = bk.get("ly_actual", 0)
                    yoy_v = round((act_rn / lst_rn - 1) * 100, 1) if lst_rn > 0 else None
                    month_data[m] = {
                        "act_rn":       act_rn,
                        "last_rn":      lst_rn,
                        "yoy":          yoy_v,
                        "bud_rn":       bud_rn,
                        "rns_fcst":     act_rn,
                        "fcst_ach":     round(act_rn / bud_rn * 100, 1) if bud_rn > 0 else 0,
                        "rns_fcst_ai":  act_rn,
                        "fcst_ach_ai":  round(act_rn / bud_rn * 100, 1) if bud_rn > 0 else 0.0,
                        "rm_fcst_rn":   None,
                        "rm_fcst_rev":  None,
                        "rm_fcst_adr":  None,
                        "rm_budget_rn": None,
                        "rm_fcst_ach":  None,
                    }

        rows.append({"name": display_name, "region": region, "months": month_data})

        # 세그먼트별 sub-row 추가
        if db_bps is not None and seg_budgets is not None:
            for seg in BUDGET_SEGMENT_KEYS:
                seg_month_data = {}
                for m in months:
                    mk_26 = f"2026{m:02d}"
                    mk_25 = f"2025{m:02d}"
                    bud_label = BUDGET_MONTH_LABEL[m - 1]
                    s_act_rn = 0
                    s_ly_rn = 0
                    for pname in db_props:
                        s_act_rn += db_bps.get(pname, {}).get(seg, {}).get(mk_26, {}).get("booking_rn", 0)
                        s_ly_rn += db_bps.get(pname, {}).get(seg, {}).get(mk_25, {}).get("booking_rn", 0)
                    sb = seg_budgets.get(display_name, {}).get(seg, {})
                    s_bud_rn = sb.get(bud_label, {}).get("rn", 0)
                    s_yoy = round((s_act_rn / s_ly_rn - 1) * 100, 1) if s_ly_rn > 0 else None
                    seg_month_data[m] = {
                        "act_rn":   s_act_rn,
                        "last_rn":  s_ly_rn,
                        "yoy":      s_yoy,
                        "bud_rn":   s_bud_rn,
                    }
                # 실적이 0인 세그먼트는 생략
                has_data = any(seg_month_data[m]["act_rn"] > 0 or seg_month_data[m]["last_rn"] > 0 for m in months)
                if has_data:
                    rows.append({"name": seg, "region": region, "months": seg_month_data, "is_segment": True, "parent": display_name})

    return rows


def build_segment_snapshot(db_seg, seg_budgets, month_idx, adj_by_segment=None, now_kst=None):
    """전체 세그먼트별 budget vs actual 요약 (예산 없는 세그먼트는 budget=0)
    adj_by_segment: 세그먼트별 동기간 보정 데이터 (2025년)
    now_kst: 현재 KST 시간 (미래월 판별용)
    """
    if now_kst is None:
        now_kst = datetime.now(KST)

    if month_idx == 0:
        target_keys = MONTHS_26
        ly_keys     = MONTHS_25
        bud_labels  = BUDGET_MONTH_LABEL
    else:
        target_keys = [MONTHS_26[month_idx - 1]]
        ly_keys     = [MONTHS_25[month_idx - 1]]
        bud_labels  = [BUDGET_MONTH_LABEL[month_idx - 1]]

    result = {}
    for seg in SEGMENT_KEYS:
        # Budget: 전체 사업장 합산
        bud_rn = 0
        bud_rev = 0.0
        for _, display_name, _, _ in PROPERTY_DEFS:
            prop_seg = seg_budgets.get(display_name, {}).get(seg, {})
            bud_rn  += sum(prop_seg.get(l, {}).get("rn",    0) for l in bud_labels)
            bud_rev += sum(prop_seg.get(l, {}).get("rev_m", 0) for l in bud_labels)
        bud_adr = round(bud_rev * 1_000_000 / bud_rn) if bud_rn > 0 else 0

        # Actual: db_seg 집계
        seg_db  = db_seg.get(seg, {})
        act_rn  = sum(seg_db.get(mk, {}).get("booking_rn",  0)   for mk in target_keys)
        act_rev = sum(seg_db.get(mk, {}).get("booking_rev", 0.0) for mk in target_keys)
        act_adr = round(act_rev * 1_000_000 / act_rn) if act_rn > 0 else 0

        # LY: 동기간 보정 적용 (미래월 판별: month_idx > 0이고 month_idx > now_kst.month이면 미래월)
        is_future_for_ly = (month_idx > 0 and month_idx > now_kst.month)

        if adj_by_segment and seg in adj_by_segment:
            seg_adj = adj_by_segment[seg]
            ly_rn  = sum(seg_adj.get(mk, {}).get("booking_rn", 0)  for mk in ly_keys)
            ly_rev_m = sum(seg_adj.get(mk, {}).get("booking_rev_m", 0.0) for mk in ly_keys)
            ly_rev = ly_rev_m  # 이미 백만원 단위
            ly_adr = round(ly_rev * 1_000_000 / ly_rn) if ly_rn > 0 else 0
        else:
            ly_rn   = sum(seg_db.get(mk, {}).get("booking_rn",  0)   for mk in ly_keys)
            ly_rev  = sum(seg_db.get(mk, {}).get("booking_rev", 0.0) for mk in ly_keys)
            ly_adr  = round(ly_rev * 1_000_000 / ly_rn) if ly_rn > 0 else 0

        rns_ach = round(act_rn  / bud_rn  * 100, 1) if bud_rn  > 0 else 0.0
        rev_ach = round(act_rev / bud_rev * 100, 1) if bud_rev > 0 else 0.0
        rns_yoy = round((act_rn / ly_rn - 1) * 100, 1) if ly_rn > 0 else 0.0

        # LY rev: adj_by_segment 있으면 백만원 단위, 없으면 원 단위
        ly_rev_won = round(ly_rev * 1_000_000) if adj_by_segment and seg in adj_by_segment else round(ly_rev * 1_000_000)

        result[seg] = {
            "rns_budget":      bud_rn,
            "rns_actual":      act_rn,
            "rns_achievement": rns_ach,
            "rns_last":        ly_rn,
            "rns_yoy":         rns_yoy,
            "today_booking":   0,
            "today_cancel":    0,
            "today_net":       0,
            "rev_budget":      round(bud_rev * 1_000_000),
            "rev_actual":      round(act_rev * 1_000_000),
            "rev_achievement": rev_ach,
            "rev_last":        ly_rev_won,
            "adr_budget":      bud_adr,
            "adr_actual":      act_adr,
            "adr_last":        ly_adr,
            "adr_vs_budget":   round((act_adr / bud_adr - 1) * 100, 1) if bud_adr > 0 else 0.0,
        }
    return result


def build_month_snapshot(db_bp, budgets, month_idx, db_seg=None, seg_budgets=None, db_bps=None,
                         adj_by_prop=None, adj_by_segment=None, holiday_factors=None, lead_time_by_prop=None, now_kst=None,
                         rm_fcst_props=None, rm_trend_snapshots=None, seg_fcst_data=None):
    """특정 월(0=전체, 1~12=해당월)에 대한 byProperty + summary + segmentData 반환
    실적/목표 모두 OTA+G-OTA+Inbound 세그먼트만 합산.

    seg_fcst_data: load_segment_fcst() 결과. 세그별 FCST는 분배 X, 원본 데이터 직접 사용.
    """
    if now_kst is None:
        now_kst = datetime.now(KST)
    if rm_fcst_props is None:
        rm_fcst_props = {}
    if seg_fcst_data is None:
        seg_fcst_data = {"manager_keyin": {}, "rm_seg_fcst": {}, "ratios": {}}

    if month_idx == 0:
        target_keys = MONTHS_26
        last_keys   = MONTHS_25
        bud_labels  = BUDGET_MONTH_LABEL
    else:
        target_keys = [MONTHS_26[month_idx - 1]]
        last_keys   = [MONTHS_25[month_idx - 1]]
        bud_labels  = [BUDGET_MONTH_LABEL[month_idx - 1]]

    props = []
    tot_bud_rn = tot_act_rn = tot_lst_rn = tot_rns_fcst = 0
    tot_bud_rev = tot_act_rev = tot_lst_rev = tot_rev_fcst = 0.0
    tot_ai_fcst_rn = 0
    tot_ai_fcst_lo = 0
    tot_ai_fcst_hi = 0
    tot_rm_fcst_rn = 0
    tot_rm_budget_rn = 0
    tot_rm_fcst_rev = 0

    for sheet_name, display_name, region, db_props in PROPERTY_DEFS:
        # Budget: OTA+G-OTA+Inbound 세그먼트만 합산
        if seg_budgets:
            b = sum_seg_budget(seg_budgets, display_name, bud_labels)
            bud_rn  = b["rn"]
            bud_adr = b["adr"]
            bud_rev = b["rev_m"]
        else:
            bud_monthly = budgets.get(display_name, {})
            bud_rn  = sum(bud_monthly.get(l, {}).get("rn",    0) for l in bud_labels)
            bud_adr = (sum(bud_monthly.get(l, {}).get("adr",   0) * bud_monthly.get(l, {}).get("rn", 0)
                           for l in bud_labels) / bud_rn) if bud_rn > 0 else 0
            bud_rev = sum(bud_monthly.get(l, {}).get("rev_m", 0) for l in bud_labels)

        # 실적: by_property_segment → OTA+G-OTA+Inbound만
        act_rn = act_rev = 0
        if db_bps is not None:
            for mk in target_keys:
                d = sum_db_segments(db_bps, db_props, mk)
                act_rn  += d["rn"]
                act_rev += d["rev_m"]
        else:
            for mk in target_keys:
                d = sum_db(db_bp, db_props, mk)
                act_rn  += d["rn"]
                act_rev += d["rev_m"]
        act_adr = round(act_rev * 1_000_000 / act_rn) if act_rn > 0 else 0

        # 전년 합산 (OTA+G-OTA+Inbound 3개 세그먼트만)
        lst_rn = 0
        lst_rev = 0.0
        if db_bps is not None:
            for mk in last_keys:
                d = sum_db_segments(db_bps, db_props, mk)
                lst_rn  += d["rn"]
                lst_rev += d["rev_m"]
        else:
            for mk in last_keys:
                d = sum_db(db_bp, db_props, mk)
                lst_rn  += d["rn"]
                lst_rev += d["rev_m"]

        # 미래월/현재월 LY 동기간 보정
        # month_idx > 0이고 month_idx >= now_kst.month인 경우: 같은 날짜까지의 2025 데이터만 포함
        is_future_or_current = (month_idx > 0 and month_idx >= now_kst.month)

        if is_future_or_current and adj_by_prop and lst_rn > 0:
            # adj_by_prop에 보정된 동기간 데이터가 있으면 사용
            adj_lst_rn = 0
            for pname in db_props:
                for mk in last_keys:
                    pm = adj_by_prop.get(pname, {}).get(mk, {})
                    # booking_rn = orig + adj (기준일 시점 동기간 RN). 없으면 final 사용.
                    if pm.get("booking_rn") is not None:
                        adj_lst_rn += pm.get("booking_rn", 0)
                    else:
                        adj_lst_rn += 0
            if adj_lst_rn > 0:
                # 매출은 동기간 RN 비율로 환산 (ADR 동일 가정)
                lst_rev = lst_rev * (adj_lst_rn / lst_rn)
                lst_rn = adj_lst_rn

        lst_adr = round(lst_rev * 1_000_000 / lst_rn) if lst_rn > 0 else 0

        rns_ach = round((act_rn / bud_rn * 100), 1) if bud_rn > 0 else 0.0
        rev_ach = round((act_rev / bud_rev * 100), 1) if bud_rev > 0 else 0.0
        rns_yoy = round((act_rn / lst_rn - 1) * 100, 1) if lst_rn > 0 else 0.0

        # 리드타임 분포 (사업장별, 대상 월 합산)
        lead_time = {}
        avg_lead_time = 0.0
        if lead_time_by_prop:
            lt_sum = defaultdict(int)
            for pname in db_props:
                prop_lt = lead_time_by_prop.get(pname, {})
                for mk in target_keys:
                    for bucket, cnt in prop_lt.get(mk, {}).items():
                        lt_sum[bucket] += cnt
            if lt_sum:
                lead_time = dict(lt_sum)
                total_rn = sum(lt_sum.values())
                if total_rn > 0:
                    avg_lead_time = round(
                        sum(cnt * LEAD_TIME_MIDPOINTS.get(b, 0) for b, cnt in lt_sum.items()) / total_rn, 1
                    )

        # FCST: elapsed_ratio 기반
        # AI FCST: 항상 독립 계산 (비교 표시용)
        ai_fcst_rn, ai_fcst_ach = 0, 0.0
        ai_fcst_lo, ai_fcst_hi = 0, 0  # Confidence Interval
        # RM FCST: rm_fcst.json에서 로드
        rm_rn_prop = None
        rm_budget_prop = None
        rm_ach_prop = None

        if month_idx > 0:
            # OTB pickup ratio (LY 풀년 / LY 동기간 OTB) — 진행 중인 월 FCST 산출용
            pickup_ratio = _compute_ly_pickup_ratio(
                adj_by_prop, db_bps, db_bp, db_props, last_keys
            )

            rns_fcst, rev_fcst, adr_fcst, fcst_ach, rev_fcst_ach = _calc_fcst(
                act_rn, act_rev, month_idx, now_kst, bud_rn, bud_rev,
                ly_pickup_ratio=pickup_ratio,
            )

            # AI FCST 독립 계산 (고도화)
            ai_fcst_rn, ai_fcst_ach, ai_fcst_lo, ai_fcst_hi = _calc_fcst_enhanced(
                act_rn, month_idx, now_kst, bud_rn, db_bp, db_props,
                holiday_factors, db_bps=db_bps, rm_trend_snapshots=rm_trend_snapshots,
                ly_pickup_ratio=pickup_ratio,
            )

            # RM FCST: rm_fcst.json segments(OTA+G-OTA+Inbound) PDF 원본 셀 합산. 분배 X.
            rm_key = f"2026-{month_idx:02d}"
            rm_rn_prop, rm_budget_prop, rm_rev_prop = sum_rm_seg_fcst(rm_fcst_props, display_name, rm_key)
            rm_ach_prop = round(rm_rn_prop / rm_budget_prop * 100, 1) if (rm_rn_prop and rm_budget_prop and rm_budget_prop > 0) else None
            rm_adr_prop = round(rm_rev_prop * 1_000_000 / rm_rn_prop) if (rm_rev_prop and rm_rn_prop and rm_rn_prop > 0) else None

            # 미래월 fallback: _calc_fcst가 None이면 AI FCST 사용
            if rns_fcst is None:
                if ai_fcst_rn is not None and ai_fcst_rn > 0:
                    rns_fcst = ai_fcst_rn
                    fcst_ach = ai_fcst_ach
                    # 매출 FCST = AI FCST RN × LY 풀년 ADR (OTA+G-OTA+Inbound 기준).
                    # ※ lst_rn/lst_rev는 위에서 동기간 보정(adj_by_prop)이 적용돼 양쪽 모두 축소된 값이라
                    #   ratio 계산 시 LY 풀년 ADR로 정규화하지 않으면 rev_fcst 폭주 (예: 92조).
                    mk_25 = f"2025{month_idx:02d}"
                    if db_bps is not None:
                        ly_full = sum_db_segments(db_bps, db_props, mk_25)
                    else:
                        ly_full = sum_db(db_bp, db_props, mk_25)
                    ly_full_rn = ly_full["rn"]
                    ly_full_rev = ly_full["rev_m"]
                    if ly_full_rn > 0:
                        ly_adr_m = ly_full_rev / ly_full_rn  # 백만원/RN
                        rev_fcst = rns_fcst * ly_adr_m
                    else:
                        rev_fcst = 0.0
                    adr_fcst = round(rev_fcst * 1_000_000 / rns_fcst) if rns_fcst > 0 else 0
                    rev_fcst_ach = round(rev_fcst / bud_rev * 100, 1) if bud_rev > 0 else 0.0
        else:
            # 전체 월: 개별 월(1~12월)별 FCST를 합산
            rns_fcst = 0
            rev_fcst = 0.0
            ai_fcst_rn = 0
            rm_rn_prop_sum = 0
            rm_budget_prop_sum = 0
            rm_rev_prop_sum = 0
            rm_rn_prop_has = False
            rm_rev_prop_has = False
            for mi in range(1, 13):
                mk_mi = f"2026{mi:02d}"
                if db_bps is not None:
                    mi_d = sum_db_segments(db_bps, db_props, mk_mi)
                else:
                    mi_d = sum_db(db_bp, db_props, mk_mi)
                mi_rn = mi_d["rn"]
                mi_rev = mi_d["rev_m"]
                if seg_budgets:
                    mi_bud = sum_seg_budget(seg_budgets, display_name, [BUDGET_MONTH_LABEL[mi - 1]])
                    mi_bud_rn = mi_bud["rn"]
                    mi_bud_rev = mi_bud["rev_m"]
                else:
                    mi_bud_rn = budgets.get(display_name, {}).get(BUDGET_MONTH_LABEL[mi - 1], {}).get("rn", 0)
                    mi_bud_rev = budgets.get(display_name, {}).get(BUDGET_MONTH_LABEL[mi - 1], {}).get("rev_m", 0)

                # OTB pickup ratio (월별: 이 월의 LY 동기간 기준)
                mi_pickup_ratio = _compute_ly_pickup_ratio(
                    adj_by_prop, db_bps, db_bp, db_props, [f"2025{mi:02d}"]
                )

                # AI FCST (월별 합산, 고도화)
                mi_ai, _, mi_lo, mi_hi = _calc_fcst_enhanced(
                    mi_rn, mi, now_kst, mi_bud_rn, db_bp, db_props,
                    holiday_factors, db_bps=db_bps, rm_trend_snapshots=rm_trend_snapshots,
                    ly_pickup_ratio=mi_pickup_ratio,
                )
                if mi_ai is not None and mi_ai > 0:
                    ai_fcst_rn += mi_ai
                    ai_fcst_lo += mi_lo if mi_lo is not None else 0
                    ai_fcst_hi += mi_hi if mi_hi is not None else 0

                # RM FCST (월별 합산): rm_fcst.json segments PDF 원본 셀 합산
                mi_rm_key = f"2026-{mi:02d}"
                mi_rm_rn, mi_rm_bud, mi_rm_rev = sum_rm_seg_fcst(rm_fcst_props, display_name, mi_rm_key)
                if mi_rm_rn is not None:
                    rm_rn_prop_sum += mi_rm_rn
                    rm_rn_prop_has = True
                if mi_rm_bud is not None:
                    rm_budget_prop_sum += mi_rm_bud
                if mi_rm_rev is not None:
                    rm_rev_prop_sum += mi_rm_rev
                    rm_rev_prop_has = True

                mi_rns_f, mi_rev_f, _, _, _ = _calc_fcst(
                    mi_rn, mi_rev, mi, now_kst, mi_bud_rn, mi_bud_rev,
                    ly_pickup_ratio=mi_pickup_ratio,
                )
                if mi_rns_f is not None:
                    rns_fcst += mi_rns_f
                    rev_fcst += mi_rev_f
                else:
                    # 미래월 fallback: AI FCST × LY 풀년 ADR (OTA+G-OTA+Inbound)
                    if mi_ai is not None and mi_ai > 0:
                        rns_fcst += mi_ai
                        mk_25_mi = f"2025{mi:02d}"
                        if db_bps is not None:
                            mi_lst = sum_db_segments(db_bps, db_props, mk_25_mi)
                        else:
                            mi_lst = sum_db(db_bp, db_props, mk_25_mi)
                        if mi_lst["rn"] > 0:
                            ly_adr_m = mi_lst["rev_m"] / mi_lst["rn"]  # 백만원/RN
                            rev_fcst += mi_ai * ly_adr_m
                        else:
                            rev_fcst += mi_bud_rev
            adr_fcst = round(rev_fcst * 1_000_000 / rns_fcst) if rns_fcst > 0 else 0
            fcst_ach = round(rns_fcst / bud_rn * 100, 1) if bud_rn > 0 else 0.0
            rev_fcst_ach = round(rev_fcst / bud_rev * 100, 1) if bud_rev > 0 else 0.0
            ai_fcst_ach = round(ai_fcst_rn / bud_rn * 100, 1) if bud_rn > 0 else 0.0
            rm_rn_prop = rm_rn_prop_sum if rm_rn_prop_has else None
            rm_budget_prop = rm_budget_prop_sum if rm_rn_prop_has else None
            rm_rev_prop = rm_rev_prop_sum if rm_rev_prop_has else None
            rm_ach_prop = round(rm_rn_prop / rm_budget_prop * 100, 1) if (rm_rn_prop and rm_budget_prop and rm_budget_prop > 0) else None
            rm_adr_prop = round(rm_rev_prop * 1_000_000 / rm_rn_prop) if (rm_rev_prop and rm_rn_prop and rm_rn_prop > 0) else None

        props.append({
            "name":            display_name,
            "region":          region,
            "rns_budget":      bud_rn,
            "rns_actual":      act_rn,
            "rns_achievement": rns_ach,
            "rns_last":        lst_rn,
            "rns_yoy":         rns_yoy,
            "rns_fcst":        rns_fcst,
            "fcst_achievement": fcst_ach,
            "ai_fcst_rn":      ai_fcst_rn,
            "ai_fcst_ach":     ai_fcst_ach,
            "ai_fcst_lo":      ai_fcst_lo,
            "ai_fcst_hi":      ai_fcst_hi,
            "rm_fcst_rn":      rm_rn_prop,
            "rm_fcst_rev":     rm_rev_prop,
            "rm_fcst_adr":     rm_adr_prop,
            "rm_budget_rn":    rm_budget_prop,
            "rm_fcst_ach":     rm_ach_prop,
            "adr_budget":      round(bud_adr),
            "adr_actual":      act_adr,
            "adr_last":        lst_adr,
            "rev_budget":      round(bud_rev * 1_000_000),
            "rev_actual":      round(act_rev * 1_000_000),
            "rev_last":        round(lst_rev * 1_000_000),
            "rev_achievement": rev_ach,
            "adr_yoy":         round((act_adr / lst_adr - 1) * 100, 1) if lst_adr > 0 else 0.0,
            "rev_yoy":         round((act_rev / lst_rev - 1) * 100, 1) if lst_rev > 0 else 0.0,
            "adr_fcst":        adr_fcst,
            "rev_fcst":        round(rev_fcst * 1_000_000) if rev_fcst is not None else None,
            "adr_fcst_achievement": round(adr_fcst / bud_adr * 100, 1) if (bud_adr > 0 and adr_fcst is not None) else None,
            "rev_fcst_achievement": rev_fcst_ach,
            "today_booking":   0,
            "today_cancel":    0,
            "today_net":       0,
            "today_booking_rev": 0,
            "today_cancel_rev":  0,
            "today_net_rev":     0,
            "lead_time":       lead_time,
            "avg_lead_time":   avg_lead_time,
        })
        tot_bud_rn   += bud_rn
        tot_act_rn   += act_rn
        tot_lst_rn   += lst_rn
        tot_rns_fcst += rns_fcst if rns_fcst is not None else 0
        tot_bud_rev  += bud_rev
        tot_act_rev  += act_rev
        tot_lst_rev  += lst_rev
        tot_rev_fcst += rev_fcst if rev_fcst is not None else 0.0
        tot_ai_fcst_rn += ai_fcst_rn if ai_fcst_rn is not None else 0
        tot_ai_fcst_lo += ai_fcst_lo if ai_fcst_lo is not None else 0
        tot_ai_fcst_hi += ai_fcst_hi if ai_fcst_hi is not None else 0
        if rm_rn_prop is not None:
            tot_rm_fcst_rn += rm_rn_prop
        if rm_budget_prop is not None:
            tot_rm_budget_rn += rm_budget_prop
        if rm_rev_prop is not None:
            tot_rm_fcst_rev += rm_rev_prop

    # 미래월 판별: month_idx가 단일 월이고 현재월보다 크면 FCST=None
    is_future_month = (month_idx > 0 and month_idx > now_kst.month)

    tot_rns_ach  = round(tot_act_rn  / tot_bud_rn  * 100, 1) if tot_bud_rn  > 0 else 0.0
    tot_rev_ach  = round(tot_act_rev / tot_bud_rev  * 100, 1) if tot_bud_rev  > 0 else 0.0
    tot_adr_act  = round(tot_act_rev * 1_000_000 / tot_act_rn)  if tot_act_rn  > 0 else 0
    tot_adr_bud  = round(tot_bud_rev * 1_000_000 / tot_bud_rn)  if tot_bud_rn  > 0 else 0
    tot_adr_lst  = round(tot_lst_rev * 1_000_000 / tot_lst_rn)  if tot_lst_rn  > 0 else 0
    if is_future_month:
        # 미래월이라도 enhanced FCST 합산값이 있으면 사용
        if tot_rns_fcst and tot_rns_fcst > 0:
            tot_adr_fcst = round(tot_rev_fcst * 1_000_000 / tot_rns_fcst) if tot_rns_fcst > 0 else 0
            tot_fcst_ach     = round(tot_rns_fcst / tot_bud_rn  * 100, 1) if tot_bud_rn  > 0 else 0.0
            tot_rev_fcst_ach = round(tot_rev_fcst / tot_bud_rev * 100, 1) if tot_bud_rev > 0 else 0.0
        else:
            tot_rns_fcst = None
            tot_rev_fcst = None
            tot_adr_fcst = None
            tot_fcst_ach = None
            tot_rev_fcst_ach = None
    else:
        tot_adr_fcst = round(tot_rev_fcst * 1_000_000 / tot_rns_fcst) if tot_rns_fcst > 0 else 0
        tot_fcst_ach     = round(tot_rns_fcst / tot_bud_rn  * 100, 1) if tot_bud_rn  > 0 else 0.0
        tot_rev_fcst_ach = round(tot_rev_fcst / tot_bud_rev * 100, 1) if tot_bud_rev > 0 else 0.0
    tot_yoy      = round((tot_act_rn  / tot_lst_rn  - 1) * 100, 1) if tot_lst_rn  > 0 else 0.0

    tot_ai_fcst_ach = round(tot_ai_fcst_rn / tot_bud_rn * 100, 1) if tot_bud_rn > 0 else 0.0
    tot_rm_fcst_ach = round(tot_rm_fcst_rn / tot_rm_budget_rn * 100, 1) if (tot_rm_budget_rn > 0 and tot_rm_fcst_rn > 0) else None

    summary = {
        "rns_budget":      tot_bud_rn,
        "rns_actual":      tot_act_rn,
        "rns_achievement": tot_rns_ach,
        "rns_last":        tot_lst_rn,
        "rns_yoy":         tot_yoy,
        "rns_fcst":        tot_rns_fcst,
        "fcst_achievement": tot_fcst_ach,
        "ai_fcst_rn":      tot_ai_fcst_rn,
        "ai_fcst_ach":     tot_ai_fcst_ach,
        "ai_fcst_lo":      tot_ai_fcst_lo,
        "ai_fcst_hi":      tot_ai_fcst_hi,
        "rm_fcst_rn":      tot_rm_fcst_rn if tot_rm_fcst_rn > 0 else None,
        "rm_fcst_rev":     tot_rm_fcst_rev if tot_rm_fcst_rev > 0 else None,
        "rm_fcst_adr":     round(tot_rm_fcst_rev * 1_000_000 / tot_rm_fcst_rn) if (tot_rm_fcst_rev > 0 and tot_rm_fcst_rn > 0) else None,
        "rm_budget_rn":    tot_rm_budget_rn if tot_rm_budget_rn > 0 else None,
        "rm_fcst_ach":     tot_rm_fcst_ach,
        "today_booking":   0,
        "today_cancel":    0,
        "today_net":       0,
        "today_booking_rev": 0,
        "today_cancel_rev":  0,
        "today_net_rev":     0,
        "rev_budget":      round(tot_bud_rev * 1_000_000),
        "rev_actual":      round(tot_act_rev * 1_000_000),
        "rev_last":        round(tot_lst_rev * 1_000_000),
        "rev_achievement": tot_rev_ach,
        "rev_yoy":         round((tot_act_rev / tot_lst_rev - 1) * 100, 1) if tot_lst_rev > 0 else 0.0,
        "rev_fcst":        round(tot_rev_fcst * 1_000_000) if tot_rev_fcst is not None else None,
        "rev_fcst_achievement": tot_rev_fcst_ach,
        "adr_budget":      tot_adr_bud,
        "adr_actual":      tot_adr_act,
        "adr_last":        tot_adr_lst,
        "adr_yoy":         round((tot_adr_act / tot_adr_lst - 1) * 100, 1) if tot_adr_lst > 0 else 0.0,
        "adr_fcst":        tot_adr_fcst,
        "adr_fcst_achievement": round(tot_adr_fcst / tot_adr_bud * 100, 1) if (tot_adr_bud > 0 and tot_adr_fcst is not None) else None,
        "adr_vs_budget":   round((tot_adr_act / tot_adr_bud - 1) * 100, 1) if tot_adr_bud > 0 else 0.0,
    }
    seg_data = {}
    if db_seg is not None and seg_budgets is not None:
        seg_data = build_segment_snapshot(db_seg, seg_budgets, month_idx, adj_by_segment=adj_by_segment, now_kst=now_kst)

    # ── 세그먼트별 byProperty (각 세그먼트 단독 기준) ──
    # FCST는 분배/추정하지 않고 ORIGINAL 데이터를 직접 사용:
    #   1) ai_fcst.json manager_keyin_segments (매니저가 키인한 세그별 FCST)
    #   2) fcst_segment_trend.json snapshots latest rm_fcst_rn (RM 회의 결과 4년 가중 mix 분배)
    #   3) historical mix ratio (RM 미커버 월: 7월 이후) — RM이 4-6월에 사용한 동일 mix 로직
    #   4) 과거월: actual = fcst
    prop_total_by_name = {p["name"]: p for p in props}
    by_prop_seg = {}
    if db_bps is not None and seg_budgets is not None:
        # 사업장별 세그먼트 actual/LY/budget 합 사전 계산
        prop_seg_data = {}
        for _, display_name, region, db_props in PROPERTY_DEFS:
            seg_aggs = {}
            for seg in SEGMENT_KEYS:
                sb = seg_budgets.get(display_name, {}).get(seg, {})
                s_bud_rn  = sum(sb.get(l, {}).get("rn",    0) for l in bud_labels)
                s_bud_rev = sum(sb.get(l, {}).get("rev_m", 0) for l in bud_labels)
                s_bud_adr = (sum(sb.get(l, {}).get("adr", 0) * sb.get(l, {}).get("rn", 0)
                                 for l in bud_labels) / s_bud_rn) if s_bud_rn > 0 else 0
                s_act_rn = s_act_rev = 0
                s_lst_rn = s_lst_rev = 0
                for mk in target_keys:
                    for pname in db_props:
                        m = db_bps.get(pname, {}).get(seg, {}).get(mk, {})
                        s_act_rn  += m.get("booking_rn",  0)
                        s_act_rev += m.get("booking_rev", 0.0)
                for mk in last_keys:
                    for pname in db_props:
                        m = db_bps.get(pname, {}).get(seg, {}).get(mk, {})
                        s_lst_rn  += m.get("booking_rn",  0)
                        s_lst_rev += m.get("booking_rev", 0.0)
                seg_aggs[seg] = {
                    "bud_rn": s_bud_rn, "bud_rev": s_bud_rev, "bud_adr": s_bud_adr,
                    "act_rn": s_act_rn, "act_rev": s_act_rev,
                    "lst_rn": s_lst_rn, "lst_rev": s_lst_rev,
                }
            prop_seg_data[display_name] = seg_aggs

        # RM FCST 세그별 PDF 원본 셀 (OTA/G-OTA/Inbound만 — rm_fcst_props.segments[seg]).
        # 단일월: 해당 월의 segments[seg]; 전체월: 12개월 합산.
        def _rm_seg(display_name: str, seg: str):
            """Return (rm_fcst_rn, rm_budget_rn) for this segment×month_idx. None if missing."""
            if seg not in ("OTA", "G-OTA", "Inbound"):
                return None, None
            prop_data = rm_fcst_props.get(display_name, {})
            if not prop_data:
                return None, None
            if month_idx > 0:
                ym = f"2026-{month_idx:02d}"
                seg_data = prop_data.get(ym, {}).get("segments", {}).get(seg, {})
                f = seg_data.get("rm_fcst_rn")
                b = seg_data.get("rm_budget_rn")
                return (f if f is not None else None,
                        b if b is not None else None)
            # 전체(연간): 12개월 segment 합
            f_sum = 0; b_sum = 0; has_f = False; has_b = False
            for mi in range(1, 13):
                ym_mi = f"2026-{mi:02d}"
                seg_data = prop_data.get(ym_mi, {}).get("segments", {}).get(seg, {})
                vf = seg_data.get("rm_fcst_rn")
                vb = seg_data.get("rm_budget_rn")
                if vf is not None: f_sum += int(vf); has_f = True
                if vb is not None: b_sum += int(vb); has_b = True
            return (f_sum if has_f else None, b_sum if has_b else None)

        for seg in SEGMENT_KEYS:
            seg_props = []
            for _, display_name, region, db_props in PROPERTY_DEFS:
                sa = prop_seg_data[display_name][seg]
                s_bud_rn = sa["bud_rn"]; s_bud_rev = sa["bud_rev"]; s_bud_adr = sa["bud_adr"]
                s_act_rn = sa["act_rn"]; s_act_rev = sa["act_rev"]
                s_lst_rn = sa["lst_rn"]; s_lst_rev = sa["lst_rev"]

                s_rns_ach = round((s_act_rn / s_bud_rn * 100), 1) if s_bud_rn > 0 else 0.0
                s_rev_ach = round((s_act_rev / s_bud_rev * 100), 1) if s_bud_rev > 0 else 0.0
                s_act_adr = round((s_act_rev * 1_000_000) / s_act_rn) if s_act_rn > 0 else 0

                # 사업장 total FCST (RM 미커버 월 ratio fallback에 필요)
                p_total = prop_total_by_name.get(display_name, {})
                p_rns_fcst_total = p_total.get("rns_fcst", 0) or 0

                is_past_single_month = (month_idx > 0 and month_idx < now_kst.month)
                is_current_single_month = (month_idx > 0 and month_idx == now_kst.month)
                fcst_source = None  # debug/audit용

                if is_past_single_month:
                    # 과거월: actual = fcst
                    s_rns_f = s_act_rn
                    fcst_source = "past_actual"
                elif is_current_single_month and seg in BUDGET_SEGMENT_KEYS:
                    # 현재월: 사업장 pickup_ratio × 세그 OTB
                    # (manager_keyin/RM 분배는 미래월 계획용 — 현재월에는 OTB 기반이 더 정확)
                    p_pickup_ratio = _compute_ly_pickup_ratio(
                        adj_by_prop, db_bps, db_bp, db_props, last_keys
                    )
                    if p_pickup_ratio and p_pickup_ratio > 0:
                        s_rns_f = round(s_act_rn * p_pickup_ratio)
                        fcst_source = "pickup_ratio"
                    else:
                        s_rns_f = s_act_rn
                        fcst_source = "no_pickup_actual"
                elif seg in BUDGET_SEGMENT_KEYS and month_idx > 0:
                    # 미래 단일월: 원본 세그 FCST 우선 사용 (sanity check 포함)
                    ym = f"2026-{month_idx:02d}"
                    s_rns_f, fcst_source = get_seg_fcst(
                        seg_fcst_data, display_name, ym, seg,
                        p_total_fcst=p_rns_fcst_total, month_int=month_idx,
                        s_bud_rn=s_bud_rn, s_lst_rn=s_lst_rn,
                    )
                    if s_rns_f is None:
                        # 모든 source 부재 → actual 보수적 사용
                        s_rns_f = s_act_rn
                        fcst_source = "no_source_actual"
                elif seg in BUDGET_SEGMENT_KEYS and month_idx == 0:
                    # 전체(연간): 12개월 합산
                    s_rns_f = 0
                    for mi in range(1, 13):
                        ym_mi = f"2026-{mi:02d}"
                        mi_seg_bud = seg_budgets.get(display_name, {}).get(seg, {}).get(BUDGET_MONTH_LABEL[mi - 1], {}).get("rn", 0)
                        # 해당 월의 LY 실적 (sanity bound용)
                        mi_lst_seg = 0
                        for pname in db_props:
                            mi_lst_seg += db_bps.get(pname, {}).get(seg, {}).get(MONTHS_25[mi - 1], {}).get("booking_rn", 0)
                        if mi < now_kst.month:
                            # 과거월: actual
                            mi_act = 0
                            for pname in db_props:
                                m = db_bps.get(pname, {}).get(seg, {}).get(MONTHS_26[mi - 1], {})
                                mi_act += m.get("booking_rn", 0)
                            s_rns_f += mi_act
                        elif mi == now_kst.month:
                            # 현재월: 사업장 pickup_ratio × 세그 OTB
                            mi_pickup_ratio = _compute_ly_pickup_ratio(
                                adj_by_prop, db_bps, db_bp, db_props, [MONTHS_25[mi - 1]]
                            )
                            mi_seg_act = 0
                            for pname in db_props:
                                mi_seg_act += db_bps.get(pname, {}).get(seg, {}).get(MONTHS_26[mi - 1], {}).get("booking_rn", 0)
                            if mi_pickup_ratio and mi_pickup_ratio > 0:
                                s_rns_f += round(mi_seg_act * mi_pickup_ratio)
                            else:
                                s_rns_f += mi_seg_act
                        else:
                            # 미래월: 원본 source (sanity check 적용)
                            mi_seg_f, _ = get_seg_fcst(
                                seg_fcst_data, display_name, ym_mi, seg,
                                p_total_fcst=p_rns_fcst_total / max(1, 12 - now_kst.month + 1),
                                month_int=mi,
                                s_bud_rn=mi_seg_bud, s_lst_rn=mi_lst_seg,
                            )
                            s_rns_f += mi_seg_f if mi_seg_f is not None else 0
                    fcst_source = "annual_aggregate"
                else:
                    # 기타 세그(단체/패키지 등): actual 사용
                    s_rns_f = s_act_rn
                    fcst_source = "non_budget_actual"

                # 매출 FCST: rev = rn × ADR(LY 동일 세그 ADR 우선, 없으면 budget ADR)
                if s_lst_rn > 0:
                    seg_adr_m = s_lst_rev / s_lst_rn  # 백만원/RN
                elif s_bud_rn > 0:
                    seg_adr_m = s_bud_rev / s_bud_rn
                else:
                    seg_adr_m = 0.0
                s_rev_f = s_rns_f * seg_adr_m

                s_fcst_ach = round(s_rns_f / s_bud_rn * 100, 1) if s_bud_rn > 0 else 0.0
                s_rev_fcst_ach = round(s_rev_f / s_bud_rev * 100, 1) if s_bud_rev > 0 else 0.0
                s_adr_fcst = round((s_rev_f * 1_000_000) / s_rns_f) if s_rns_f > 0 else 0

                # RM FCST: PDF segments 원본 셀 (해당 세그 단일값)
                s_rm_rn, s_rm_bud = _rm_seg(display_name, seg)
                s_rm_ach = round(s_rm_rn / s_rm_bud * 100, 1) if (s_rm_rn is not None and s_rm_bud and s_rm_bud > 0) else None

                seg_props.append({
                    "name": display_name,
                    "region": region,
                    "rns_budget":      s_bud_rn,
                    "rns_actual":      s_act_rn,
                    "rns_achievement": s_rns_ach,
                    "rns_last":        s_lst_rn,
                    "rns_yoy":         round((s_act_rn / s_lst_rn - 1) * 100, 1) if s_lst_rn > 0 else 0.0,
                    "rns_fcst":        s_rns_f,
                    "fcst_achievement": s_fcst_ach,
                    "fcst_source":     fcst_source,
                    "rm_fcst_rn":      s_rm_rn,
                    "rm_budget_rn":    s_rm_bud,
                    "rm_fcst_ach":     s_rm_ach,
                    "adr_budget":      round(s_bud_adr),
                    "adr_actual":      s_act_adr,
                    "adr_fcst":        s_adr_fcst,
                    "rev_budget":      round(s_bud_rev * 1_000_000),
                    "rev_actual":      round(s_act_rev * 1_000_000),
                    "rev_last":        round(s_lst_rev * 1_000_000),
                    "rev_fcst":        round(s_rev_f * 1_000_000),
                    "rev_achievement": s_rev_ach,
                    "rev_fcst_achievement": s_rev_fcst_ach,
                    "today_booking":   0,
                    "today_cancel":    0,
                    "today_net":       0,
                    "lead_time":       {},
                })
            by_prop_seg[seg] = seg_props

    return {"byProperty": props, "byPropertySegment": by_prop_seg, "summary": summary, "segmentData": seg_data}


def build_monthly_chart(db_bp, budgets, seg_budgets=None, db_bps=None, adj_by_prop=None):
    """Chart용 월별 집계 (전체 사업장 합산, OTA+G-OTA+Inbound 기준)"""
    result = []
    for i, mk in enumerate(MONTHS_26):
        bud_label = BUDGET_MONTH_LABEL[i]
        mk_25 = MONTHS_25[i]
        bud_rn = 0
        act_rn = 0
        lst_rn = 0
        for sheet_name, display_name, region, db_props in PROPERTY_DEFS:
            if seg_budgets:
                bud_rn += sum_seg_budget(seg_budgets, display_name, [bud_label])["rn"]
            else:
                bud_rn += budgets.get(display_name, {}).get(bud_label, {}).get("rn", 0)
            if db_bps is not None:
                d = sum_db_segments(db_bps, db_props, mk)
            else:
                d = sum_db(db_bp, db_props, mk)
            act_rn += d["rn"]
            if adj_by_prop:
                for pname in db_props:
                    adj_m = adj_by_prop.get(pname, {}).get(mk_25, {})
                    lst_rn += adj_m.get("booking_rn", 0) if adj_m else sum_db(db_bp, [pname], mk_25)["rn"]
            else:
                d_last = sum_db(db_bp, db_props, mk_25)
                lst_rn += d_last["rn"]
        result.append({
            "month": i + 1,
            "label": bud_label,
            "rns_budget": bud_rn,
            "rns_actual": act_rn,
            "rns_last":   lst_rn,
        })
    return result


def get_today_summary(db, now_kst):
    """OTA+G-OTA+Inbound 3개 세그먼트만 합산한 전일 데이터 반환.
    빌드 시점(now_kst)의 당일 데이터는 미완성이므로 전일(days_ago=1)부터 탐색."""
    nd_seg = db.get("net_daily_by_segment", {})
    for days_ago in range(1, 8):
        date_str = (now_kst - timedelta(days=days_ago)).strftime("%Y%m%d")
        pickup, cancel = 0, 0
        for seg in BUDGET_SEGMENT_KEYS:
            entry = nd_seg.get(seg, {}).get(date_str, {})
            pickup += entry.get("pickup_rn", 0)
            cancel += entry.get("cancel_rn", 0)
        if pickup > 0 or cancel > 0:
            return pickup, cancel, pickup - cancel, date_str
    # fallback: 전체 net_daily에서 최근 날짜 (세그먼트 필터 적용)
    net_daily = db.get("net_daily", {})
    if net_daily:
        most_recent = max(net_daily.keys())
        pickup, cancel = 0, 0
        for seg in BUDGET_SEGMENT_KEYS:
            entry = nd_seg.get(seg, {}).get(most_recent, {})
            pickup += entry.get("pickup_rn", 0)
            cancel += entry.get("cancel_rn", 0)
        return pickup, cancel, pickup - cancel, most_recent
    return 0, 0, 0, None


def get_today_booking_by_props(db, date_str, db_props):
    """pickup_daily_by_property_segment에서 OTA+G-OTA+Inbound만 합산."""
    if not date_str:
        return 0, 0
    pdbps = db.get("pickup_daily_by_property_segment", {})
    rn, rev = 0, 0.0
    for pname in db_props:
        for seg in BUDGET_SEGMENT_KEYS:
            d = pdbps.get(pname, {}).get(seg, {}).get(date_str, {})
            rn  += d.get("rn",  0)
            rev += d.get("rev", 0.0)
    return rn, rev


def get_today_cancel_by_props(db, date_str, db_props):
    """cancel_daily_by_property_segment에서 OTA+G-OTA+Inbound만 합산."""
    if not date_str:
        return 0, 0
    cdbps = db.get("cancel_daily_by_property_segment", {})
    rn, rev = 0, 0.0
    for pname in db_props:
        for seg in BUDGET_SEGMENT_KEYS:
            d = cdbps.get(pname, {}).get(seg, {}).get(date_str, {})
            rn  += d.get("rn",  0)
            rev += d.get("rev", 0.0)
    return rn, rev


def get_today_booking_by_props_month(db, date_str, db_props, stay_month):
    """pickup_daily_by_property_segment_month에서 OTA+G-OTA+Inbound만 합산."""
    if not date_str:
        return 0, 0
    pdbpsm = db.get("pickup_daily_by_property_segment_month", {})
    rn, rev = 0, 0.0
    for pname in db_props:
        for seg in BUDGET_SEGMENT_KEYS:
            d = pdbpsm.get(pname, {}).get(seg, {}).get(stay_month, {}).get(date_str, {})
            rn  += d.get("rn",  0)
            rev += d.get("rev", 0.0)
    return rn, rev


def get_today_cancel_by_props_month(db, date_str, db_props, stay_month):
    """cancel_daily_by_property_segment_month에서 OTA+G-OTA+Inbound만 합산."""
    if not date_str:
        return 0, 0
    cdbpsm = db.get("cancel_daily_by_property_segment_month", {})
    rn, rev = 0, 0.0
    for pname in db_props:
        for seg in BUDGET_SEGMENT_KEYS:
            d = cdbpsm.get(pname, {}).get(seg, {}).get(stay_month, {}).get(date_str, {})
            rn  += d.get("rn",  0)
            rev += d.get("rev", 0.0)
    return rn, rev


def get_today_summary_by_month(db, now_kst, stay_month):
    """OTA+G-OTA+Inbound 3개 세그먼트만 합산, 특정 투숙월의 전일 데이터 반환."""
    pdsm = db.get("pickup_daily_by_segment_month", {})
    cdsm = db.get("cancel_daily_by_segment_month", {})
    for days_ago in range(1, 8):
        date_str = (now_kst - timedelta(days=days_ago)).strftime("%Y%m%d")
        pickup, cancel = 0, 0
        for seg in BUDGET_SEGMENT_KEYS:
            pickup += pdsm.get(seg, {}).get(stay_month, {}).get(date_str, {}).get("rn", 0)
            cancel += cdsm.get(seg, {}).get(stay_month, {}).get(date_str, {}).get("rn", 0)
        if pickup > 0 or cancel > 0:
            return pickup, cancel, pickup - cancel
    # fallback
    net_daily_m = db.get("net_daily_by_month", {}).get(stay_month, {})
    if net_daily_m:
        most_recent = max(net_daily_m.keys())
        pickup, cancel = 0, 0
        for seg in BUDGET_SEGMENT_KEYS:
            pickup += pdsm.get(seg, {}).get(stay_month, {}).get(most_recent, {}).get("rn", 0)
            cancel += cdsm.get(seg, {}).get(stay_month, {}).get(most_recent, {}).get("rn", 0)
        return pickup, cancel, pickup - cancel
    return 0, 0, 0


def load_daily_booking():
    """daily_booking.json에서 팔라티움 등 온북 DB 미포함 사업장 데이터 로드.
    Returns: {display_name: {month_idx: {budget_rns, actual_rns, ly_actual, daily_change, ...}}}
    """
    if not DAILY_BOOKING_JSON.exists():
        return {}
    try:
        data = json.loads(DAILY_BOOKING_JSON.read_text(encoding="utf-8"))
    except Exception:
        return {}
    # 온북 DB에 없는 사업장만 수집 (db_props가 빈 사업장)
    no_db_names = {dn for _, dn, _, dp in PROPERTY_DEFS if not dp}
    display_map = {}
    for _, dn, _, _ in PROPERTY_DEFS:
        if dn in no_db_names:
            # daily_booking에서의 이름 매핑
            if "팔라티움" in dn:
                display_map["팔라티움 해운대"] = dn
    result = {}
    for md in data.get("months_detail", []):
        month_idx = md.get("month", 0)
        year = md.get("year", 0)
        if year != 2026 or month_idx < 1 or month_idx > 12:
            continue
        for entry in md.get("properties", []):
            name = entry.get("name", "")
            if name in display_map:
                dn = display_map[name]
                if dn not in result:
                    result[dn] = {}
                result[dn][month_idx] = {
                    "budget_rns":  entry.get("budget_rns", 0),
                    "actual_rns":  entry.get("actual_rns", 0),
                    "ly_actual":   entry.get("ly_actual", 0),
                    "daily_change": entry.get("daily_change", 0),
                    "budget_achievement": entry.get("budget_achievement", 0),
                    "occ_budget":  entry.get("occ_budget", 0),
                    "occ_actual":  entry.get("occ_actual", 0),
                    "occ_ly":      entry.get("occ_ly", 0),
                    "daily_occ":   entry.get("daily_occ", []),
                }
    return result


def overlay_daily_booking(all_months, daily_bk, now_kst):
    """온북 DB 미포함 사업장의 OTB 데이터를 daily_booking.json으로 보정.
    byProperty 내 해당 사업장의 rns_budget/actual/achievement 등을 덮어씀.
    summary 합산도 재계산.
    """
    if not daily_bk:
        return
    no_db_names = {dn for _, dn, _, dp in PROPERTY_DEFS if not dp}

    for m_str, snap in all_months.items():
        # "summary"=전체 합산(0과 동치), "1".."12"=해당 월
        is_summary = (m_str == "summary")
        m_idx = 0 if is_summary else int(m_str)
        props = snap.get("byProperty", [])

        # summary 보정을 위해 추가분 추적
        delta_bud_rn = 0
        delta_act_rn = 0
        delta_lst_rn = 0

        for prop in props:
            dn = prop.get("name", "")
            if dn not in no_db_names or dn not in daily_bk:
                continue

            if m_idx == 0:
                # 전체: 모든 월 합산
                bud_rn = sum(daily_bk[dn].get(mi, {}).get("budget_rns", 0) for mi in range(1, 13))
                act_rn = sum(daily_bk[dn].get(mi, {}).get("actual_rns", 0) for mi in range(1, 13))
                lst_rn = sum(daily_bk[dn].get(mi, {}).get("ly_actual", 0) for mi in range(1, 13))
                today_net = sum(daily_bk[dn].get(mi, {}).get("daily_change", 0) for mi in range(1, 13))
            else:
                bk = daily_bk[dn].get(m_idx, {})
                bud_rn = bk.get("budget_rns", 0)
                act_rn = bk.get("actual_rns", 0)
                lst_rn = bk.get("ly_actual", 0)
                today_net = bk.get("daily_change", 0)

            old_bud = prop.get("rns_budget", 0)
            old_act = prop.get("rns_actual", 0)
            old_lst = prop.get("rns_last", 0)

            prop["rns_budget"] = bud_rn
            prop["rns_actual"] = act_rn
            prop["rns_achievement"] = round(act_rn / bud_rn * 100, 1) if bud_rn > 0 else 0.0
            prop["rns_last"] = lst_rn
            prop["rns_yoy"] = round((act_rn / lst_rn - 1) * 100, 1) if lst_rn > 0 else 0.0
            prop["today_net"] = today_net
            prop["today_booking"] = max(today_net, 0)
            prop["today_cancel"] = max(-today_net, 0)

            # FCST: OTB 그대로 사용 (LY pickup 데이터 부재 — 일별 외삽 폭주 방지)
            #   팔라티움 등 daily_booking 사업장은 LY OTB 동기간 데이터가 없어
            #   pickup_ratio 추정 불가. act_rn(현재 OTB)을 보수적으로 FCST로 사용.
            if m_idx > 0:
                fcst_rn = act_rn
                prop["rns_fcst"] = fcst_rn
                prop["fcst_achievement"] = round(fcst_rn / bud_rn * 100, 1) if bud_rn > 0 else 0.0
                prop["ai_fcst_rn"] = fcst_rn
                prop["ai_fcst_ach"] = prop["fcst_achievement"]

            delta_bud_rn += (bud_rn - old_bud)
            delta_act_rn += (act_rn - old_act)
            delta_lst_rn += (lst_rn - old_lst)

        # summary 재계산
        if delta_bud_rn != 0 or delta_act_rn != 0:
            summary = snap.get("summary", {})
            summary["rns_budget"] = summary.get("rns_budget", 0) + delta_bud_rn
            summary["rns_actual"] = summary.get("rns_actual", 0) + delta_act_rn
            summary["rns_last"]   = summary.get("rns_last", 0) + delta_lst_rn
            sb = summary["rns_budget"]
            sa = summary["rns_actual"]
            sl = summary["rns_last"]
            summary["rns_achievement"] = round(sa / sb * 100, 1) if sb > 0 else 0.0
            summary["rns_yoy"] = round((sa / sl - 1) * 100, 1) if sl > 0 else 0.0


def main():
    print("db_aggregated.json 로드 중...")
    db = json.loads(DB_JSON.read_text(encoding="utf-8"))
    db_bp  = db.get("by_property", {})
    db_seg = db.get("by_segment", {})
    db_bps = db.get("by_property_segment", {})

    # 전체 세그먼트 동적 탐색 (기타 제외, 예산 세그먼트 우선)
    global SEGMENT_KEYS
    all_segs_set = set()
    for prop_segs in db_bps.values():
        all_segs_set.update(prop_segs.keys())
    all_segs_set.discard("기타")  # 기타 제거
    # 예산 세그먼트 우선, 나머지 가나다 정렬
    other_segs = sorted(all_segs_set - set(BUDGET_SEGMENT_KEYS))
    SEGMENT_KEYS = BUDGET_SEGMENT_KEYS + other_segs
    print(f"  세그먼트 탐색: {len(SEGMENT_KEYS)}개 (기타 제외)")

    print("Budget XLSX 로드 중...")
    budgets, seg_budgets = {}, {}
    if BUDGET_XLSX.exists():
        wb = openpyxl.load_workbook(BUDGET_XLSX, read_only=True, data_only=True)
        budgets, seg_budgets = load_budget(wb)
    else:
        print(f"  ⚠ Budget 파일 없음: {BUDGET_XLSX.name} — budget 없이 진행")

    now_kst = datetime.now(KST)

    # YoY 동기간 보정값 로드
    adj_by_prop = {}
    adj_by_segment = {}
    yoy_base_date = ""
    yoy_adj_section = db.get("yoy_adjusted", {})
    if "2025" in yoy_adj_section:
        adj_by_prop    = yoy_adj_section["2025"].get("by_property", {})
        adj_by_segment = yoy_adj_section["2025"].get("by_segment", {})
        yoy_base_date  = yoy_adj_section["2025"].get("base_date_full", "")
    print(f"  YoY 동기간 보정 로드: 사업장 수={len(adj_by_prop)}, 세그먼트 수={len(adj_by_segment)}, base={yoy_base_date}")

    # 사업장별 리드타임 분포
    lead_time_by_prop = db.get("lead_time_by_property", {})

    # RM FCST 로드
    print("RM FCST 로드 중...")
    rm_fcst_props = load_rm_fcst()
    print(f"  RM FCST 사업장 수: {len(rm_fcst_props)}")

    # RM FCST Trend 스냅샷 로드 (리드타임 pickup 패턴 분석용)
    rm_trend_snapshots = []
    rm_trend_path = DATA_DIR / "rm_fcst_trend.json"
    if rm_trend_path.exists():
        try:
            rm_trend_data = json.loads(rm_trend_path.read_text(encoding="utf-8"))
            rm_trend_snapshots = rm_trend_data.get("snapshots", [])
            print(f"  RM FCST Trend 스냅샷: {len(rm_trend_snapshots)}개")
        except Exception:
            print("  RM FCST Trend 로드 실패 — 리드타임 반영 건너뜀")

    # 공휴일 보정 계수 (전월 대상)
    print("  공휴일 보정 계수 계산 중 (로컬 holidays_kr.json)...")
    holiday_factors = build_holiday_factors(target_months=tuple(range(1, 13)), cur_year=2026, base_year=2025)
    print(f"  holiday_factors={holiday_factors}")

    # 오늘(또는 가장 최근) 예약/취소/순증 데이터
    today_booking, today_cancel, today_net, today_date = get_today_summary(db, now_kst)
    print(f"  오늘 데이터 날짜: {today_date}")
    print(f"  today_booking={today_booking}, today_cancel={today_cancel}, today_net={today_net}")

    # 세그먼트별 FCST 원본 로드 (manager_keyin + RM segment FCST 분배본 + historical mix ratios)
    seg_fcst_data = load_segment_fcst()
    _km = sum(len(v) for v in seg_fcst_data.get("manager_keyin", {}).values())
    _rm = sum(len(v) for v in seg_fcst_data.get("rm_seg_fcst", {}).values())
    print(f"  세그먼트 FCST source: manager_keyin={_km}건, rm_seg_fcst={_rm}건, ratios={len(seg_fcst_data.get('ratios', {}))}개 사업장")

    # 월별 스냅샷
    #   "summary" : 전체 합산 (12개월 통합)
    #   "1"~"12"  : 해당 월
    # ※ 하위 호환성: 출력 직전에 "0" 키를 "summary" 별칭으로 추가 (구 소비처용)
    SUMMARY_KEY = "summary"
    all_months = {}
    all_months[SUMMARY_KEY] = build_month_snapshot(
        db_bp, budgets, 0,
        db_seg=db_seg, seg_budgets=seg_budgets, db_bps=db_bps,
        adj_by_prop=adj_by_prop, adj_by_segment=adj_by_segment,
        holiday_factors=holiday_factors,
        lead_time_by_prop=lead_time_by_prop, now_kst=now_kst,
        rm_fcst_props=rm_fcst_props,
        rm_trend_snapshots=rm_trend_snapshots,
        seg_fcst_data=seg_fcst_data,
    )
    for m in range(1, 13):
        all_months[str(m)] = build_month_snapshot(
            db_bp, budgets, m,
            db_seg=db_seg, seg_budgets=seg_budgets, db_bps=db_bps,
            adj_by_prop=adj_by_prop, adj_by_segment=adj_by_segment,
            holiday_factors=holiday_factors,
            lead_time_by_prop=lead_time_by_prop, now_kst=now_kst,
            rm_fcst_props=rm_fcst_props,
            rm_trend_snapshots=rm_trend_snapshots,
            seg_fcst_data=seg_fcst_data,
        )

    # today 데이터를 월 스냅샷에 주입 (월별로 stay_month 필터 적용)
    for m_str, snap in all_months.items():
        is_summary = (m_str == SUMMARY_KEY)
        if is_summary:
            snap["summary"]["today_booking"] = today_booking
            snap["summary"]["today_cancel"]  = today_cancel
            snap["summary"]["today_net"]     = today_net
        else:
            stay_month = f"2026{int(m_str):02d}"
            m_booking, m_cancel, m_net = get_today_summary_by_month(db, now_kst, stay_month)
            snap["summary"]["today_booking"] = m_booking
            snap["summary"]["today_cancel"]  = m_cancel
            snap["summary"]["today_net"]     = m_net

        for prop in snap["byProperty"]:
            db_props = next((d for _, n, _, d in PROPERTY_DEFS if n == prop["name"]), [])
            if is_summary:
                prop_booking, prop_booking_rev = get_today_booking_by_props(db, today_date, db_props)
                prop_cancel,  prop_cancel_rev  = get_today_cancel_by_props(db, today_date, db_props)
            else:
                stay_month = f"2026{int(m_str):02d}"
                prop_booking, prop_booking_rev = get_today_booking_by_props_month(db, today_date, db_props, stay_month)
                prop_cancel,  prop_cancel_rev  = get_today_cancel_by_props_month(db, today_date, db_props, stay_month)
            prop["today_booking"]     = prop_booking
            prop["today_cancel"]      = prop_cancel
            prop["today_net"]         = prop_booking - prop_cancel
            prop["today_booking_rev"] = round(prop_booking_rev * 1_000_000)
            prop["today_cancel_rev"]  = round(prop_cancel_rev  * 1_000_000)
            prop["today_net_rev"]     = round((prop_booking_rev - prop_cancel_rev) * 1_000_000)

        if is_summary:
            snap["summary"]["today_booking_rev"] = sum(p["today_booking_rev"] for p in snap["byProperty"])
            snap["summary"]["today_cancel_rev"]  = sum(p["today_cancel_rev"]  for p in snap["byProperty"])
            snap["summary"]["today_net_rev"]     = sum(p["today_net_rev"]     for p in snap["byProperty"])

        for seg_name, seg_props in snap.get("byPropertySegment", {}).items():
            for prop in seg_props:
                db_props = next((d for _, n, _, d in PROPERTY_DEFS if n == prop["name"]), [])
                # 해당 세그먼트만의 today 데이터 (3세그 합산 X → 개별 세그먼트)
                pdbps = db.get("pickup_daily_by_property_segment", {})
                cdbps = db.get("cancel_daily_by_property_segment", {})
                if is_summary:
                    prop_booking, prop_cancel = 0, 0
                    if today_date:
                        for pname in db_props:
                            prop_booking += pdbps.get(pname, {}).get(seg_name, {}).get(today_date, {}).get("rn", 0) or 0
                            prop_cancel  += cdbps.get(pname, {}).get(seg_name, {}).get(today_date, {}).get("rn", 0) or 0
                else:
                    stay_month = f"2026{int(m_str):02d}"
                    pdbpsm = db.get("pickup_daily_by_property_segment_month", {})
                    cdbpsm = db.get("cancel_daily_by_property_segment_month", {})
                    prop_booking, prop_cancel = 0, 0
                    if today_date:
                        for pname in db_props:
                            prop_booking += pdbpsm.get(pname, {}).get(seg_name, {}).get(stay_month, {}).get(today_date, {}).get("rn", 0) or 0
                            prop_cancel  += cdbpsm.get(pname, {}).get(seg_name, {}).get(stay_month, {}).get(today_date, {}).get("rn", 0) or 0
                prop["today_booking"] = prop_booking
                prop["today_cancel"]  = prop_cancel
                prop["today_net"]     = prop_booking - prop_cancel

        # ── segmentData에 today 주입 ──
        for seg, seg_summary in snap.get("segmentData", {}).items():
            if today_date:
                if is_summary:
                    # 전체: net_daily_by_segment (전월 합산)
                    nds = db.get("net_daily_by_segment", {})
                    pds = db.get("pickup_daily_by_segment", {})
                    cds = db.get("cancel_daily_by_segment", {})
                    nd_entry = nds.get(seg, {}).get(today_date, {})
                    seg_summary["today_booking"] = nd_entry.get("pickup_rn", 0)
                    seg_summary["today_cancel"]  = nd_entry.get("cancel_rn", 0)
                    seg_summary["today_net"]     = nd_entry.get("net_rn", 0)
                    pd_entry = pds.get(seg, {}).get(today_date, {})
                    cd_entry = cds.get(seg, {}).get(today_date, {})
                    seg_summary["today_booking_rev"] = round(pd_entry.get("rev", 0.0) * 1_000_000)
                    seg_summary["today_cancel_rev"]  = round(cd_entry.get("rev", 0.0) * 1_000_000)
                else:
                    # 월별: pickup/cancel_daily_by_segment_month 사용
                    stay_month = f"2026{int(m_str):02d}"
                    pdsm = db.get("pickup_daily_by_segment_month", {})
                    cdsm = db.get("cancel_daily_by_segment_month", {})
                    pd_entry = pdsm.get(seg, {}).get(stay_month, {}).get(today_date, {})
                    cd_entry = cdsm.get(seg, {}).get(stay_month, {}).get(today_date, {})
                    seg_summary["today_booking"] = pd_entry.get("rn", 0) or 0
                    seg_summary["today_cancel"]  = cd_entry.get("rn", 0) or 0
                    seg_summary["today_net"]     = seg_summary["today_booking"] - seg_summary["today_cancel"]
                    seg_summary["today_booking_rev"] = round((pd_entry.get("rev", 0.0) or 0) * 1_000_000)
                    seg_summary["today_cancel_rev"]  = round((cd_entry.get("rev", 0.0) or 0) * 1_000_000)
                seg_summary["today_net_rev"]     = seg_summary["today_booking_rev"] - seg_summary["today_cancel_rev"]

    # ── 온북 DB 미포함 사업장(팔라티움 등) daily_booking.json 보정 ──
    print("Daily Booking 보정 (온북 DB 미포함 사업장)...")
    daily_bk = load_daily_booking()
    if daily_bk:
        overlay_daily_booking(all_months, daily_bk, now_kst)
        print(f"  보정 대상: {list(daily_bk.keys())}")
    else:
        print("  daily_booking.json 없음 — 보정 건너뜀")

    # Chart data (동기간 보정 반영)
    monthly_chart = build_monthly_chart(
        db_bp, budgets, seg_budgets=seg_budgets, db_bps=db_bps, adj_by_prop=adj_by_prop
    )

    # YoY 사업장별 추이 테이블 (자동 롤링)
    yoy_table = build_yoy_table(
        db_bp, budgets, seg_budgets, db_bps, adj_by_prop, holiday_factors,
        months=TARGET_MONTHS, now_kst=now_kst, rm_fcst_props=rm_fcst_props,
        daily_bk=daily_bk, rm_trend_snapshots=rm_trend_snapshots,
    )

    output = {
        "meta": {
            "refreshTime":  now_kst.strftime("%Y-%m-%d %H:%M KST"),
            "baseDate":     now_kst.strftime("%Y-%m-%d"),
            "todayDate":    today_date,
            "yoyBaseDate":  yoy_base_date,
            "dataSource":   "온북 DB + 사업계획 Budget",
        },
        "filters": {
            "months": [{"value": 0, "label": "전체"}] + [
                {"value": i+1, "label": f"{i+1}월"} for i in range(12)
            ],
            "segments": ["전체"] + SEGMENT_KEYS,
        },
        # 전체(기본) 스냅샷 (월 필터=전체 상태)
        "summary":    all_months["summary"]["summary"],
        "byProperty": all_months["summary"]["byProperty"],
        # 월별 분리 데이터 (월 필터 작동용)
        #   "summary" = 전체 합산 / "1".."12" = 해당 월
        #   하위 호환성: "0"은 "summary"의 별칭 (구 소비처 보호용; 신규 코드는 "summary" 사용)
        "allMonths":  {**all_months, "0": all_months["summary"]},
        # Chart
        "monthly":    monthly_chart,
        # YoY 사업장별 추이
        "yoyTable":   yoy_table,
        # 투숙일별 일별 데이터 (stay-date.html 용)
        "stayDateDaily": db.get("stay_date_daily", {}),
    }

    DOCS_DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✓ {OUTPUT_JSON} 생성 완료")
    print(f"  사업장 수: {len(output['byProperty'])}")
    print(f"  총 목표 RNS: {output['summary']['rns_budget']:,}")
    print(f"  총 실적 RNS: {output['summary']['rns_actual']:,}")
    print(f"  달성률: {output['summary']['rns_achievement']}%")
    ai_rn = output['summary'].get('ai_fcst_rn')
    ai_lo = output['summary'].get('ai_fcst_lo')
    ai_hi = output['summary'].get('ai_fcst_hi')
    rm_rn = output['summary'].get('rm_fcst_rn')
    print(f"  AI FCST RN: {ai_rn:,}" if ai_rn else "  AI FCST RN: N/A")
    if ai_lo and ai_hi:
        print(f"  AI FCST CI: [{ai_lo:,} ~ {ai_hi:,}]")
    print(f"  RM FCST RN: {rm_rn:,}" if rm_rn else "  RM FCST RN: N/A")


if __name__ == "__main__":
    main()
