#!/usr/bin/env python3
"""
팔라티움 해운대 by sonofelice 해운대 예약 DB(xlsx) → daily_booking.json 반영
=========================================================
data/palatium_db/예약정보조회*.xlsx 파일을 파싱하여
팔라티움 해운대 by sonofelice 사업장의 RNs, OCC, Budget, 당일변동 등을 산출하고
daily_booking.json에 팔라티움 해운대 by sonofelice 데이터를 추가/갱신합니다.

Budget 출처: data/palatium_db/*사업계획*.xlsx (요약 시트)
OCC 출처:    data/palatium_room_availability.json (parse_palatium_rooms.py 산출)

parse_raw_db.py 종료 후 자동 호출되거나 단독 실행 가능.
"""
import json
import fs_utils  # macOS NFD→NFC 유니코드 정규화
import logging
import re
import sys
import calendar
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

ROOT = Path(__file__).parent.parent
DATA_DIR = ROOT / "data"
DOCS_DATA_DIR = ROOT / "docs" / "data"
PALATIUM_DB_DIR = DATA_DIR / "palatium_db"
DAILY_BOOKING_JSON = DATA_DIR / "daily_booking.json"
DOCS_DAILY_BOOKING = DOCS_DATA_DIR / "daily_booking.json"
ROOM_AVAIL_JSON = DATA_DIR / "palatium_room_availability.json"

PROPERTY_NAME = "팔라티움 해운대 by sonofelice 해운대"
REGION = "south"
TOTAL_ROOMS = 201  # 기본 객실수 (room_availability 없을 때 폴백)

# 예약 상태 분류
ACTIVE_STATUSES = {"Reservation", "In House", "Checked Out", "Assigned Room", "No Show"}
CANCEL_STATUS = "Cancelled Reservation"


def find_latest_xlsx():
    """data/palatium_db/ 에서 최신 예약정보조회 xlsx 파일들을 찾음."""
    import unicodedata
    if not PALATIUM_DB_DIR.exists():
        return []
    files = sorted(
        f for f in PALATIUM_DB_DIR.iterdir()
        if unicodedata.normalize("NFC", f.name).startswith("예약정보조회")
        and f.suffix == ".xlsx"
    )
    if not files:
        return []
    # 파일명 패턴: 예약정보조회MMDD0N.xlsx (예: 예약정보조회051801.xlsx)
    # 또는 구형: 예약정보조회4월01.xlsx
    date_groups = defaultdict(list)
    for f in files:
        name = f.stem  # 예약정보조회051801
        # 패턴1: MMDD 숫자 (6자리 이상)
        m1 = re.search(r'예약정보조회(\d{4,})', name)
        if m1:
            date_key = m1.group(1)[:4]  # MMDD
            date_groups[date_key].append(f)
            continue
        # 패턴2: M월NN (예: 4월01)
        m2 = re.search(r'예약정보조회(\d{1,2})월', name)
        if m2:
            month_num = int(m2.group(1))
            date_key = f"{month_num:02d}00"  # 0400 (일자 불명이면 00)
            date_groups[date_key].append(f)
            continue
        date_groups["0000"].append(f)

    if not date_groups:
        return list(files)
    # 가장 최근 날짜 그룹 반환
    latest_key = max(date_groups.keys())
    return date_groups[latest_key]


def find_budget_xlsx():
    """data/palatium_db/ 에서 사업계획 xlsx 파일 찾기."""
    import unicodedata
    for f in PALATIUM_DB_DIR.iterdir():
        nfc_name = unicodedata.normalize("NFC", f.name)
        if "사업계획" in nfc_name and nfc_name.endswith(".xlsx"):
            return f
    return None


def parse_reservations(xlsx_files):
    """예약정보조회 xlsx 파일들을 파싱하여 월별 RN 집계.
    Returns: {
        'sales_date': str,
        'months': {month_int: {'booking_rn', 'cancel_rn', 'revenue', 'daily_rn': {day: rn}}},
        'today_booking': int, 'today_cancel': int,
    }
    """
    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl 필요: pip install openpyxl")
        return None

    monthly = defaultdict(lambda: {
        "booking_rn": 0, "cancel_rn": 0, "revenue": 0,
        "daily_rn": defaultdict(int),  # day → room nights
    })
    sales_date = None
    today_booking = 0
    today_cancel = 0
    seen_keys = set()  # 중복 방지 (예약번호+도착일)

    for fpath in xlsx_files:
        log.info(f"  파싱: {fpath.name}")
        wb = openpyxl.load_workbook(str(fpath), data_only=True, read_only=True)
        ws = wb.active

        # Sales Date 추출 (2행)
        row2_val = None
        for row in ws.iter_rows(min_row=2, max_row=2, max_col=1, values_only=True):
            row2_val = row[0]
        if row2_val and "Sales Date" in str(row2_val) and not sales_date:
            m = re.search(r'(\d{4}-\d{2}-\d{2})', str(row2_val))
            if m:
                sales_date = m.group(1)

        # 데이터 행 (4행부터)
        for row in ws.iter_rows(min_row=4, values_only=True):
            status = row[0] if row[0] else ""
            if not status or status == "총합계" or "Created by" in str(status):
                continue

            reservation_no = row[2]  # 예약번호
            arrival = row[6]   # 도착일자
            nights = row[8]    # 박수
            room_rev = row[10]  # 객실료
            room_cnt = row[14]  # 객실수

            if not arrival:
                continue

            # 날짜 파싱
            if isinstance(arrival, datetime):
                arr_dt = arrival
            elif isinstance(arrival, date):
                arr_dt = datetime(arrival.year, arrival.month, arrival.day)
            else:
                try:
                    arr_dt = datetime.strptime(str(arrival)[:10], "%Y-%m-%d")
                except (ValueError, TypeError):
                    continue

            # 중복 체크
            dup_key = f"{reservation_no}_{arr_dt.strftime('%Y%m%d')}_{status}"
            if dup_key in seen_keys:
                continue
            seen_keys.add(dup_key)

            month = arr_dt.month
            day = arr_dt.day
            n = max(1, int(nights or 1))
            rc = max(1, int(room_cnt or 1))
            rn = n * rc
            rev = float(room_rev or 0)

            if status == CANCEL_STATUS:
                monthly[month]["cancel_rn"] += rn
                if arr_dt.date() == datetime.strptime(sales_date, "%Y-%m-%d").date() if sales_date else False:
                    today_cancel += rn
            elif status in ACTIVE_STATUSES:
                monthly[month]["booking_rn"] += rn
                monthly[month]["revenue"] += rev
                # 일별 객실수 (도착일 기준으로 배분)
                for d_off in range(n):
                    target_day = day + d_off
                    days_in_month = calendar.monthrange(arr_dt.year, month)[1]
                    if target_day <= days_in_month:
                        monthly[month]["daily_rn"][target_day] += rc

        wb.close()

    return {
        "sales_date": sales_date,
        "months": dict(monthly),
        "today_booking": today_booking,
        "today_cancel": today_cancel,
    }


def load_budget():
    """사업계획 xlsx에서 월별 Budget RNs, OCC%, ADR 로드."""
    budget_file = find_budget_xlsx()
    if not budget_file:
        log.warning("사업계획 xlsx 없음 — Budget 0으로 처리")
        return {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(budget_file), data_only=True, read_only=True)
        ws = wb["요약"]

        result = {}
        rows_data = list(ws.iter_rows(min_row=6, max_row=9, min_col=1, max_col=17, values_only=True))
        # rows_data[0] = month numbers (1..12)
        # rows_data[1] = 판매객실수(실) = Budget RNs
        # rows_data[2] = 투숙률 = OCC Budget (decimal)
        # rows_data[3] = ADR(천원)

        for m in range(1, 13):
            col_idx = m + 3  # 0-based: col 4=1월, col 15=12월 in rows_data
            budget_rn = int(rows_data[1][col_idx] or 0)
            occ_bud = round(float(rows_data[2][col_idx] or 0) * 100, 1)
            adr_bud = round(float(rows_data[3][col_idx] or 0), 1)
            result[m] = {
                "budget_rns": budget_rn,
                "occ_budget": occ_bud,
                "adr_budget": adr_bud,
            }
        wb.close()
        log.info(f"  Budget 로드: {budget_file.name}")
        return result
    except Exception as e:
        log.error(f"Budget 로드 실패: {e}")
        return {}


def load_room_availability():
    """palatium_room_availability.json에서 일별/월별 OCC 데이터 로드."""
    if not ROOM_AVAIL_JSON.exists():
        return {}, {}
    try:
        data = json.loads(ROOM_AVAIL_JSON.read_text(encoding="utf-8"))
        return data.get("monthly", {}), data.get("daily", {})
    except Exception:
        return {}, {}


def build_daily_occ(month: int, year: int, daily_avail: dict) -> list:
    """월의 일별 OCC 리스트 생성 (room_availability.json 기반)."""
    days_in_month = calendar.monthrange(year, month)[1]
    occ_list = []
    for d in range(1, days_in_month + 1):
        key = f"{year}-{month:02d}-{d:02d}"
        if key in daily_avail:
            occ_list.append(round(daily_avail[key].get("occ", 0), 1))
        else:
            occ_list.append(0)
    return occ_list


def update_daily_booking(reservations, budget, room_monthly, room_daily):
    """daily_booking.json에 팔라티움 해운대 by sonofelice 데이터를 추가/갱신."""
    # 기존 daily_booking.json 로드 (없으면 새로 생성)
    if DAILY_BOOKING_JSON.exists():
        data = json.loads(DAILY_BOOKING_JSON.read_text(encoding="utf-8"))
    else:
        data = {"meta": {}, "months_detail": []}

    sales_date = reservations["sales_date"] or datetime.now().strftime("%Y-%m-%d")
    year = int(sales_date[:4])

    # meta 업데이트
    data["meta"]["report_date"] = sales_date
    # months 결정: 현재월 + 다음 2개월
    current_month = int(sales_date[5:7])
    target_months = []
    for i in range(3):
        m = current_month + i
        if m <= 12:
            target_months.append(f"{year}-{m:02d}")
    data["meta"]["months"] = target_months
    data["meta"]["source"] = "Daily_Booking_Report"
    data["meta"]["property_count"] = 25  # 기존 값 유지

    # months_detail에서 팔라티움 해운대 by sonofelice 엔트리 추가/갱신
    for month_key in target_months:
        m = int(month_key.split("-")[1])
        days_in_month = calendar.monthrange(year, m)[1]

        # 기존 month_detail 찾기
        md_entry = None
        for md in data.get("months_detail", []):
            if md.get("month_key") == month_key:
                md_entry = md
                break
        if not md_entry:
            md_entry = {
                "month_key": month_key,
                "month": m,
                "year": year,
                "properties": [],
            }
            data.setdefault("months_detail", []).append(md_entry)

        # 팔라티움 해운대 by sonofelice 프로퍼티 찾기/생성
        prop_entry = None
        for p in md_entry.get("properties", []):
            if p.get("name") == PROPERTY_NAME:
                prop_entry = p
                break

        # 예약 데이터
        res_month = reservations["months"].get(m, {})
        actual_rns = res_month.get("booking_rn", 0)
        cancel_rns = res_month.get("cancel_rn", 0)
        revenue = res_month.get("revenue", 0)

        # Budget
        bud = budget.get(m, {})
        budget_rns = bud.get("budget_rns", 0)
        occ_budget = bud.get("occ_budget", 0)
        adr_budget = bud.get("adr_budget", 0)

        # 달성률
        ach = round(actual_rns / budget_rns * 100, 1) if budget_rns > 0 else 0
        vs_budget = actual_rns - budget_rns

        # OCC (room_availability 기반)
        ra_monthly_key = f"{year}-{m:02d}"
        ra_month = room_monthly.get(ra_monthly_key, {})
        occ_actual = ra_month.get("occ", 0)

        # 일별 OCC
        daily_occ = build_daily_occ(m, year, room_daily)

        # 당일 변동 (대략적 — 정확한 계산은 전일 대비 필요하지만 현재 데이터로는 근사)
        daily_change = 0
        daily_dir = "▲"

        # LY (전년도 데이터 — 현재 없음)
        ly_actual = 0
        occ_ly = 0
        yoy_rns = 0
        yoy_pct = 0.0
        occ_yoy_change = round(occ_actual - occ_ly, 1) if occ_ly > 0 else round(occ_actual - occ_budget, 1)

        new_prop = {
            "name": PROPERTY_NAME,
            "region": REGION,
            "budget_rns": budget_rns,
            "ly_actual": ly_actual,
            "vs_budget": vs_budget,
            "actual_rns": actual_rns,
            "daily_change": daily_change,
            "daily_change_dir": daily_dir,
            "budget_achievement": ach,
            "yoy_rns": yoy_rns,
            "yoy_pct": yoy_pct,
            "total_rn_capacity": ra_month.get("total_sum", TOTAL_ROOMS * days_in_month),
            "budget_per_day": round(budget_rns / days_in_month) if days_in_month > 0 else 0,
            "occ_budget": occ_budget,
            "occ_ly": occ_ly,
            "occ_yoy_change": occ_yoy_change,
            "occ_actual": round(occ_actual, 1),
            "daily_occ": daily_occ,
        }

        if prop_entry:
            prop_entry.update(new_prop)
        else:
            md_entry.setdefault("properties", []).append(new_prop)

    return data


def main():
    log.info("=" * 60)
    log.info("팔라티움 해운대 by sonofelice DB 파싱 시작")
    log.info("=" * 60)

    # 1. xlsx 파일 찾기
    xlsx_files = find_latest_xlsx()
    if not xlsx_files:
        log.warning("예약정보조회 xlsx 파일 없음 — 스킵")
        return False

    log.info(f"예약정보조회 파일: {len(xlsx_files)}개")

    # 2. 예약 데이터 파싱
    reservations = parse_reservations(xlsx_files)
    if not reservations:
        log.error("예약 데이터 파싱 실패")
        return False

    log.info(f"  Sales Date: {reservations['sales_date']}")
    for m, d in sorted(reservations["months"].items()):
        log.info(f"  {m}월: booking={d['booking_rn']}, cancel={d['cancel_rn']}, rev={d['revenue']/1e6:.1f}M")

    # 3. Budget 로드
    budget = load_budget()

    # 4. Room Availability 로드 (OCC 계산용)
    room_monthly, room_daily = load_room_availability()
    if room_monthly:
        log.info(f"  Room Availability: {len(room_monthly)}개월")

    # 5. daily_booking.json 갱신
    data = update_daily_booking(reservations, budget, room_monthly, room_daily)

    # 6. 저장
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    DAILY_BOOKING_JSON.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    log.info(f"✓ 저장: {DAILY_BOOKING_JSON}")

    DOCS_DATA_DIR.mkdir(parents=True, exist_ok=True)
    DOCS_DAILY_BOOKING.write_text(
        json.dumps(data, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    log.info(f"✓ docs/data 동기화: {DOCS_DAILY_BOOKING}")

    # 요약
    for mk in data["meta"].get("months", []):
        m = int(mk.split("-")[1])
        for p in next((md["properties"] for md in data["months_detail"] if md["month_key"] == mk), []):
            if p["name"] == PROPERTY_NAME:
                log.info(f"  [{mk}] actual={p['actual_rns']}, budget={p['budget_rns']}, ach={p['budget_achievement']}%, occ={p['occ_actual']}%")

    log.info("=" * 60)
    log.info("✅ 팔라티움 해운대 by sonofelice DB 파싱 완료")
    log.info("=" * 60)
    return True


if __name__ == "__main__":
    main()
