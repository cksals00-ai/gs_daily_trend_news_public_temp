#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_july_pickup_tracker.py — 7월 동기간 픽업 일자별 관리 엑셀 (송예건 CM 지시, 2026-06-30~)

대상: 7월 동기간 YoY 모니터링 전사업장 (세그 구분없이 전체 = 온라인영업팀 raw_db 전 세그)
  소노캄 비발디파크 / 소노문 단양 / 소노벨 청송 / 소노캄 여수 / 소노캄 거제 / 쏠비치 진도

데이터: data/raw_db (27/28/43/44) 라이브 스냅샷 직접 net 계산 — db_aggregated 이중계상(doubling) 회피.
픽업 정의(하우스 컨벤션, parse_raw_db.pickup_daily_agg/cancel_daily 동일):
  신규(+) = 27/43 + 28/44 행 중 최초입력일자 == 그 날 (gross add; 28/44는 기취소분 복원)
  취소(−) = 28/44 행 중 취소일자 == 그 날
  net = 신규 − 취소.  투숙월 = 판매일자[:6] == 202607(2026) / 202507(2025).
정제: 라인 중복제거, 거래처(회원명==이용자명, 단 Inbound 58 예외) 제거, 매출조정 제거.
RN = 객실수(없으면 1). 세그 구분 없음(전체 합).

동기간 온북(요약):  entry ≤ asof − cancel ≤ asof.
  asof_2026 = 스냅샷일 − 1 (마지막 완전일, 예 20260629).  asof_2025 = 동일 MMDD 전년 (20250629).
  → 전사 OTB 대시보드 todayDate(=어제) 기준과 동일. [[yoy-same-period-base-date]]

서식: GS 소노위크 실적 분석 보고 양식(대명체, 흑색 헤더+백색 텍스트, 헤어라인 테두리,
  회계 숫자서식 _-* #,##0_-, 0.0% 비중, 합계 이중테두리, 제목/번호목차/총평 박스).
날짜 하드코딩 없음: 27 라이브 스냅샷 파일명에서 기준일 자동 추출 → 매일 재실행하면 일자별 자동 연장.
산출물: ~/Desktop/7월동기간_픽업_일자별관리_YYYYMMDD.xlsx

사용: python3 scripts/build_july_pickup_tracker.py
"""
import os, re, sys, glob, json, shutil, unicodedata
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
RAW = PROJECT_DIR / "data" / "raw_db"

WINDOW_DAYS = 30  # 일별 픽업 시트: 마지막 완전일 기준 최근 N일

# 대상 = 전사업장(raw_db 온라인영업팀 기준 25개): (raw_db 정규화명, 표시 라벨=BSR명)
# BSR 페이지 순서(권역별)로 정렬. 일부는 브랜드명 상이(소노문 단양=소노벨 단양 등).
TARGETS = [
    ("소노벨 비발디파크",          "소노벨 비발디파크"),
    ("소노캄 비발디파크",          "소노캄 비발디파크"),
    ("소노펫 비발디파크",          "소노펫 비발디파크"),
    ("소노펠리체 비발디파크",       "소노펠리체 비발디파크"),
    ("소노펠리체 빌리지 비발디파크", "소노펠리체 빌리지 비발디파크"),
    ("소노문 비발디파크",          "소노문 비발디파크"),          # BSR 없음
    ("소노휴 양평",               "소노벨 양평"),
    ("델피노",                   "델피노"),
    ("쏠비치 양양",               "쏠비치 양양"),
    ("쏠비치 삼척",               "쏠비치 삼척"),
    ("소노문 단양",               "소노벨 단양"),
    ("소노벨 경주",               "소노캄 경주"),
    ("소노벨 청송",               "소노벨 청송"),
    ("소노벨 천안",               "소노벨 천안"),
    ("소노벨 변산",               "소노벨 변산"),
    ("소노캄 여수",               "소노캄 여수"),
    ("소노캄 거제",               "소노캄 거제"),
    ("쏠비치 진도",               "쏠비치 진도"),
    ("소노벨 제주",               "소노벨 제주"),
    ("소노캄 제주",               "소노캄 제주"),
    ("소노캄 고양",               "소노캄 고양"),
    ("소노문 해운대",             "소노문 해운대"),
    ("쏠비치 남해",               "쏠비치 남해"),
    ("르네블루",                 "르네블루"),
    ("오션월드빌리지",            "오션월드빌리지"),               # BSR 없음
]
TARGET_SET = {t[0] for t in TARGETS}
LABEL = {t[0]: t[1] for t in TARGETS}

# ─── 세그먼트 (변경예약집계코드 기준, 공식매핑: 회원구분_팀별취합본 '예약집계코드(현행)' 영업자료(대)) ───
#   FIT = 영업자료(대) '03일반' = OTA + G-OTA + 홈페이지 + 제휴사 + 일반(기타)  ← 소노 BSR FIT 정의
#   그 외: 인바운드(02단체/인바운드) → Inbound,  회원(01)·단체(02)·COMP(04)·기타(05) → 기타
#   ※ 제휴사·일반 코드는 GS 온라인영업팀 raw_db엔 거의 없음(타팀 채널) → BSR 대비 미달의 근본원인.
OTA_CODES     = {"51", "53", "72"}                                             # OTA R/O·PKG
GOTA_CODES    = {"A4", "A5"}                                                    # G-OTA R/O·PKG
HOME_CODES    = {"34", "73", "90", "GB"}                                        # 홈페이지 RO·PKG (D멤버스·자사·FIT·글로벌D멤버스)
AFFIL_CODES   = {"03", "04", "23", "45", "46", "47", "52", "60", "81", "93", "95", "CP"}  # 제휴사
GEN_CODES     = {"31", "32", "33", "35", "36", "48", "49", "70"}                # 일반(기타): 세트권·카드대여·할인권·예약부일반 등
INBOUND_CODES = {"54", "58", "A3", "A6", "A7"}                                  # 인바운드

SEGMENTS = ["OTA", "G-OTA", "홈페이지", "제휴사", "일반", "Inbound", "기타"]
TEAM_SEGMENTS = ["OTA", "G-OTA"]                                # GS 관리
HOME_SEGMENTS = ["홈페이지", "제휴사", "일반"]                    # 기타(자체채널)
FIT_SEGMENTS = ["OTA", "G-OTA", "홈페이지", "제휴사", "일반"]     # 합계(FIT) = GS + 기타
DEFAULT_SEGMENTS = FIT_SEGMENTS                                 # 기본 = FIT

def seg_bucket(cnum):
    n = (cnum or "").strip()
    if n in OTA_CODES:     return "OTA"
    if n in GOTA_CODES:    return "G-OTA"
    if n in HOME_CODES:    return "홈페이지"
    if n in AFFIL_CODES:   return "제휴사"
    if n in GEN_CODES:     return "일반"
    if n in INBOUND_CODES: return "Inbound"
    return "기타"

# ───────────────────────── 파일 해소 (macOS NFD/NFC) ─────────────────────────
def _nfc(s): return unicodedata.normalize("NFC", s)

def _list(year):
    d = RAW / str(year)
    return [(str(d / fn), _nfc(fn)) for fn in os.listdir(d) if fn.endswith(".txt")]

def live_files(year):
    fs = []
    for pfx in ("27", "28", "43", "44"):
        c = sorted(fp for fp, n in _list(year) if n.startswith(pfx) and "생성시간" in n)
        if c:
            fs.append(c[-1])
    return fs

def retrans_files(year):
    return sorted(fp for fp, n in _list(year) if "재전송" in n)

# BSR(Booking Status Report) PDF → 전사업장 FIT OTB(2026-07/2025-07). 로컬 실행 전용(pdfplumber).
# raw_db명 → BSR 페이지명. 기본은 TARGETS 라벨, 페이지명이 더 긴 경우만 아래에서 덮어씀.
BSR_PAGE_OVERRIDE = {"르네블루": "르네블루 바이 쏠비치"}
BSR_TAG = {name: BSR_PAGE_OVERRIDE.get(name, lab) for name, lab in TARGETS}
# BSR에 페이지가 없는 사업장 → raw_db 실측으로 폴백
NO_BSR_PAGE = {"소노문 비발디파크", "오션월드빌리지"}

def parse_bsr():
    """최신 'Booking Status Report_YYYY.MM.DD.pdf'의 Segment(OTB) 블록에서
    사업장별 7월 투숙 FIT OTB(2026/2025/차이)를 추출.

    BSR은 15시 기준 스냅샷이며 전년 대조일은 요일 정렬(2026-07-09 목 ↔ 2025-07-10 목).
    실패(라이브러리/파일/파싱 불가) 시 None → 리포트는 raw_db 실측만 사용."""
    pdf_dir = PROJECT_DIR / "data" / "Daily Booking Report PDF"
    if not pdf_dir.is_dir():
        return None
    def fdate(p):
        m = re.search(r"(\d{4})\.(\d{2})\.(\d{2})", p.name)
        return (m.group(1) + m.group(2) + m.group(3)) if m else "0"
    cands = sorted(pdf_dir.glob("Booking Status Report_*.pdf"), key=fdate)
    if not cands:
        return None
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from parse_bsr_segment import parse_bsr as _parse
        raw = _parse(cands[-1], month="07")
    except Exception:
        return None
    pages = raw.get("props") or {}
    fit26, fit25, gap, has_prior = {}, {}, {}, {}
    for name, _lab in TARGETS:
        d = pages.get(BSR_TAG[name])
        if not d:
            continue
        fit26[name] = d["fit26"]; fit25[name] = d["fit25"]
        gap[name] = d["gap"];     has_prior[name] = d["has_prior"]
    if not fit26:
        return None
    return {"date": raw["date"], "prior_date": raw["prior_date"],
            "fit26": fit26, "fit25": fit25, "gap": gap, "has_prior": has_prior,
            "total": raw["total"]}


def _apportion(total, weights):
    """total(정수)을 weights 비율로 정수 배분(최대잔여법). weights 합이 0이면 첫 칸에 전액."""
    keys = list(weights)
    wsum = sum(weights[k] for k in keys)
    if total == 0 or not keys:
        return {k: 0 for k in keys}
    if wsum <= 0:
        out = {k: 0 for k in keys}; out[keys[0]] = total; return out
    exact = {k: total * weights[k] / wsum for k in keys}
    out = {k: int(exact[k]) for k in keys}
    rem = total - sum(out.values())
    for k in sorted(keys, key=lambda k: exact[k] - int(exact[k]), reverse=True)[:rem]:
        out[k] += 1
    return out


def _rescale(v, num, den):
    """v를 num/den 배로 보정(반올림, 0에서 멀어지는 방향). v==den이면 정확히 num."""
    if den == 0 or num == den or v == 0:
        return v
    x = v * num / den
    return int(x + 0.5) if x >= 0 else -int(-x + 0.5)


def bsr_scale(anchor):
    """사업장별 보정 계수 (확정 FIT / 자사 실측 FIT). 추이(일별·누적)에 곱해
    최종 시점 누적이 확정 OTB와 정확히 일치하도록 한다.
    확정치 미제공(src=rawdb) 또는 전년 미집계면 보정 없음(계수 1)."""
    sc = {}
    for name, _lab in TARGETS:
        a = anchor.get(name) or {}
        use = a.get("src") == "bsr"
        n26, d26 = (a.get("fit26", 0), a.get("raw26", 0)) if use else (0, 0)
        n25, d25 = (a.get("fit25", 0), a.get("raw25", 0)) if (use and a.get("has_prior")) else (0, 0)
        sc[name] = {
            "n26": n26 if d26 > 0 else 1, "d26": d26 if d26 > 0 else 1,
            "n25": n25 if d25 > 0 else 1, "d25": d25 if d25 > 0 else 1,
        }
    return sc


def bsr_anchor(bsr, on26, on25):
    """BSR FIT(15시)을 사업장별 권위값으로 삼고, GS(OTA+G-OTA)/기타(홈피·제휴·일반)
    분해는 raw_db FIT 구성비로 비례배분한다.  → 합계는 BSR과 정확히 일치.

    BSR 페이지가 없는 사업장(소노문 비발디파크·오션월드빌리지)은 raw_db 실측 그대로(src=rawdb).
    BSR에 전년이 미집계된 사업장(소노캄 경주·르네블루)은 has_prior=False로 표시."""
    anchor = {}
    for name, _lab in TARGETS:
        raw26 = {s: on26.get(name, {}).get(s, 0) for s in FIT_SEGMENTS}
        raw25 = {s: on25.get(name, {}).get(s, 0) for s in FIT_SEGMENTS}
        r26, r25 = sum(raw26.values()), sum(raw25.values())
        if bsr and name in bsr["fit26"]:
            f26, f25 = bsr["fit26"][name], bsr["fit25"][name]
            src, prior = "bsr", bsr["has_prior"][name]
        else:
            f26, f25, src, prior = r26, r25, "rawdb", True
        s26 = _apportion(f26, raw26 if r26 else {s: 1 for s in FIT_SEGMENTS})
        s25 = _apportion(f25, raw25 if r25 else {s: 1 for s in FIT_SEGMENTS})
        anchor[name] = {
            "fit26": f26, "fit25": f25, "gap": f26 - f25,
            "seg26": s26, "seg25": s25,
            "raw26": r26, "raw25": r25,      # 우리 raw_db 실측(참고·정합률)
            "src": src, "has_prior": prior,
        }
    return anchor

def snapshot_date():
    """최신 27 라이브 스냅샷 파일명에서 기준일(YYYYMMDD) 추출."""
    best = None
    for fp, n in _list(2026):
        if n.startswith("27") and "생성시간" in n:
            m = re.search(r"생성시간\((\d{8})", n)
            if m:
                best = max(best, m.group(1)) if best else m.group(1)
    if not best:
        raise SystemExit("27 라이브 스냅샷을 찾지 못했습니다.")
    return best

def _openf(fp):
    for enc in ("cp949", "euc-kr", "utf-8"):
        try:
            f = open(fp, encoding=enc); f.readline(); f.close()
            return open(fp, encoding=enc)
        except UnicodeDecodeError:
            continue
    raise IOError(f"인코딩 실패: {fp}")

def _norm_prop(s):
    return re.sub(r"^\d+\.\s*", "", s).strip()

# ───────────────────────── 로딩 ─────────────────────────
# row = dict(prop, entry(YYYYMMDD|None), cancel(YYYYMMDD|None), rn, is_cancel)
def load_rows(files, staymon):
    rows = []
    for fp in files:
        ft = os.path.basename(fp)[:2]
        is_cancel = ft in ("28", "44")
        f = _openf(fp)
        hdr = f.readline().strip().split(";")
        idx = {h.strip(): i for i, h in enumerate(hdr)}
        ip = idx.get("영업장명", -1); icp = idx.get("변경사업장명", -1)
        isd = idx.get("판매일자", -1); ici = idx.get("입실일자", -1)
        irooms = idx.get("객실수", -1)
        ient = idx.get("최초입력일자", -1); ican = idx.get("취소일자", -1)
        imn = idx.get("회원명", -1); iun = idx.get("이용자명", -1)
        icn = idx.get("변경예약집계코드", idx.get("예약집계코드", -1))
        seen = set()
        for line in f:
            ls = line.rstrip("\n\r"); h = hash(ls)
            if h in seen:
                continue
            seen.add(h)
            p = line.split(";"); plen = len(p)
            def g(i): return p[i].strip() if 0 <= i < plen else ""
            sell = g(isd) or g(ici)
            if sell[:6] != staymon:
                continue
            prop = _norm_prop(g(icp)) if g(icp) else _norm_prop(g(ip))
            if prop not in TARGET_SET:
                continue
            mn, un, cn = g(imn), g(iun), g(icn)
            seg = seg_bucket(cn)
            # 거래처(회원명==이용자명) 제거는 '기타'(대매점·단체 phantom/할당행)에만 적용.
            # FIT(OTA/G-OTA/홈페이지/제휴사/일반)·Inbound은 회원명==이용자명이 정상(D멤버스·법인·자사 등) → 유지.
            if seg == "기타" and mn and un and mn == un:
                continue
            if "매출조정" in mn or "매출조정" in un:
                continue
            rooms = int(g(irooms)) if g(irooms).isdigit() else 0
            rn = rooms if rooms > 0 else 1
            ent = g(ient)[:8]; can = g(ican)[:8]
            rows.append(dict(prop=prop, seg=seg, rn=rn, is_cancel=is_cancel,
                             entry=ent if len(ent) == 8 else None,
                             cancel=can if (is_cancel and len(can) == 8) else None))
    return rows

# ───────────────────────── 집계 ─────────────────────────
def onbook_at(rows, cutoff):
    """동기간 net 온북: entry ≤ cutoff − cancel ≤ cutoff."""
    net = defaultdict(int)
    for r in rows:
        if r["entry"] and r["entry"] <= cutoff:
            net[r["prop"]] += r["rn"]
        if r["cancel"] and r["cancel"] <= cutoff:
            net[r["prop"]] -= r["rn"]
    return net

def daily_newcancel(rows):
    """(prop, YYYYMMDD) → 신규rn / 취소rn."""
    new = defaultdict(int); cxl = defaultdict(int)
    for r in rows:
        if r["entry"]:
            new[(r["prop"], r["entry"])] += r["rn"]
        if r["cancel"]:
            cxl[(r["prop"], r["cancel"])] += r["rn"]
    return new, cxl

# ───────────────────────── 엑셀 (소노위크 보고 양식) ─────────────────────────
def build_excel(out_path, data_date, asof26, asof25, rows26, rows25, seg_label="OTA+G-OTA", bsr=None, anchor=None, scale=None):
    """소노위크 보고 양식(간소화) — 개요/일별픽업(증감)/전년대비(누적)/상세. 설명문 제거."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FN_R = "대명체 보통"; FN_B = "대명체 굵게"
    WHITE = "FFFFFF"; BLACK = "000000"; RED = "C00000"; GREEN = "375623"; GREY = "808080"
    NF_NUM = r'_-* #,##0_-;\-* #,##0_-;_-* "-"_-;_-@_-'        # 회계(음수 -표기)
    NF_GAP = r'_-* +#,##0_-;_-* \-#,##0_-;_-* "-"_-;_-@_-'     # 부호(+/-) 표기
    NF_PCT = '0.0%'

    def F(sz=10, bold=False, color=BLACK):
        return Font(name=(FN_B if bold else FN_R), size=sz, bold=bold, color=color)
    cen = Alignment(horizontal="center", vertical="center")
    cenw = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rgt = Alignment(horizontal="right", vertical="center")
    lft = Alignment(horizontal="left", vertical="center")
    HAIR = Side(style="hair", color=BLACK); DBL = Side(style="double", color=BLACK)
    box = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
    def topbox(top): return Border(left=HAIR, right=HAIR, top=top, bottom=HAIR)
    HFILL = PatternFill("solid", fgColor=BLACK)

    def hcell(ws, r, c, v):
        cell = ws.cell(r, c, v); cell.font = F(10, color=WHITE); cell.fill = HFILL
        cell.alignment = cenw; cell.border = box; return cell
    def d2k(s): return f"{s[4:6]}/{s[6:8]}"
    WD = "월화수목금토일"
    def wdc(day): return WD[datetime.strptime(day, "%Y%m%d").weekday()]
    def clr(g): return GREEN if g > 0 else (RED if g < 0 else BLACK)

    # 표시 라벨 = TARGETS 라벨(전사업장 커버). raw_db명 → 표시명(BSR명).
    DATA_LBL = dict(LABEL)
    GAEYO_LBL = dict(LABEL)

    wb = openpyxl.Workbook()
    _fixed = datetime.strptime(data_date, "%Y%m%d")
    wb.properties.created = _fixed; wb.properties.modified = _fixed
    wb.properties.creator = "build_july_pickup_tracker"; wb.properties.lastModifiedBy = "build_july_pickup_tracker"

    hi = datetime.strptime(asof26, "%Y%m%d")
    days26 = [(hi - timedelta(days=k)).strftime("%Y%m%d") for k in range(WINDOW_DAYS)][::-1]
    def to25(d): return "2025" + d[4:]
    desc = list(reversed(days26))  # 최근 → 과거

    new26, cxl26 = daily_newcancel(rows26); new25, cxl25 = daily_newcancel(rows25)
    nb26 = onbook_at(rows26, asof26); nb25 = onbook_at(rows25, asof25)
    def net_of(new, cxl, name, day): return new.get((name, day), 0) - cxl.get((name, day), 0)
    # 세그별 동기간 온북 (rows는 FIT=OTA+G-OTA+홈페이지 필터본, r['seg']로 GS/홈피 분리)
    def onbook_seg(rows, cutoff, segs):
        net = defaultdict(int)
        for r in rows:
            if r.get("seg") not in segs: continue
            if r["entry"] and r["entry"] <= cutoff: net[r["prop"]] += r["rn"]
            if r["cancel"] and r["cancel"] <= cutoff: net[r["prop"]] -= r["rn"]
        return net
    TEAM_S = set(TEAM_SEGMENTS); HOME_S = set(HOME_SEGMENTS)
    tm26 = onbook_seg(rows26, asof26, TEAM_S); tm25 = onbook_seg(rows25, asof25, TEAM_S)
    hm26 = onbook_seg(rows26, asof26, HOME_S); hm25 = onbook_seg(rows25, asof25, HOME_S)
    # BSR 정합: 개요의 FIT 합계를 BSR(15시) 값으로 고정하고 GS/기타는 raw_db 구성비로 비례배분
    if anchor:
        for n, _l in TARGETS:
            a = anchor.get(n)
            if not a: continue
            tm26[n] = sum(a["seg26"][s] for s in TEAM_SEGMENTS); hm26[n] = sum(a["seg26"][s] for s in HOME_SEGMENTS)
            tm25[n] = sum(a["seg25"][s] for s in TEAM_SEGMENTS); hm25[n] = sum(a["seg25"][s] for s in HOME_SEGMENTS)

    # ============================== 개요 (GS / 홈페이지 / 합계 FIT [+ BSR]) ==============================
    ws = wb.active; ws.title = "개요"; ws.sheet_view.showGridLines = False
    # 앵커 모드에선 합계(FIT) 자체가 BSR 값 → 별도 BSR 대조 컬럼 불필요
    has_bsr = bool(bsr and bsr.get("fit26") and not anchor)
    BSR_C = 12; STAT_C = 14 if has_bsr else 12  # BSR: L(FIT)·M(달성%),  상태: N or L
    LASTC = STAT_C
    bsr_lbl = f"BSR ({bsr['date'][4:6]}/{bsr['date'][6:8]})" if has_bsr else ""
    ws.column_dimensions["A"].width = 1.2; ws.column_dimensions["B"].width = 16
    for c in range(3, STAT_C): ws.column_dimensions[get_column_letter(c)].width = 8
    ws.column_dimensions[get_column_letter(STAT_C)].width = 12

    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=LASTC)
    t = ws.cell(2, 2, "7월 동기간 比 일자별 픽업 현황"); t.font = F(14); t.alignment = cen
    ws.row_dimensions[2].height = 24
    dcell = ws.cell(3, LASTC, _fixed); dcell.number_format = "yyyy-mm-dd"; dcell.font = F(9); dcell.alignment = rgt
    ws.cell(4, 2, f"- 사업장 : 전사업장 {len(TARGETS)}개").font = F(10)
    ws.cell(5, 2, "- 일  자 : 7월 투숙건").font = F(10)
    if anchor and bsr:
        d, p = bsr["date"], bsr.get("prior_date") or ""
        fmt = lambda s: f"{s[:4]}-{s[4:6]}-{s[6:8]}" if s else "전년 동기"
        ws.cell(6, 2, f"- 기  준 : {fmt(d)} 15시 OTB  vs  {fmt(p)}  |  합계(FIT) = GS(OTA+G-OTA) + 기타(홈페이지·제휴사·일반)").font = F(10)
    else:
        ws.cell(6, 2, "- 기  준 : 전년 동기간 YOY  |  합계(FIT) = GS(OTA+G-OTA) + 기타(홈페이지·제휴사·일반)").font = F(10)
    ws.cell(7, 2, "- 사업장별 7월 동기간 OTB 현황").font = F(11)
    ic = ws.cell(7, LASTC, "[단위 : 실]"); ic.font = F(9); ic.alignment = rgt
    # 2단 헤더
    hcell(ws, 8, 2, "구분"); ws.merge_cells("B8:B9")
    hcell(ws, 8, 3, "GS (OTA+G-OTA)"); ws.merge_cells("C8:E8"); hcell(ws, 8, 4, None); hcell(ws, 8, 5, None)
    hcell(ws, 8, 6, "기타 (홈피·제휴·일반)"); ws.merge_cells("F8:H8"); hcell(ws, 8, 7, None); hcell(ws, 8, 8, None)
    hcell(ws, 8, 9, "합계 (FIT)"); ws.merge_cells("I8:K8"); hcell(ws, 8, 10, None); hcell(ws, 8, 11, None)
    if has_bsr:
        hcell(ws, 8, BSR_C, bsr_lbl); ws.merge_cells(start_row=8, start_column=BSR_C, end_row=8, end_column=BSR_C + 1)
        hcell(ws, 8, BSR_C + 1, None)
        hcell(ws, 9, BSR_C, "FIT"); hcell(ws, 9, BSR_C + 1, "달성%")
    hcell(ws, 8, STAT_C, "상태"); ws.merge_cells(start_row=8, start_column=STAT_C, end_row=9, end_column=STAT_C)
    for base in (3, 6, 9):
        for j, lab in enumerate(("26Y", "25Y", "GAP")): hcell(ws, 9, base + j, lab)

    def grp3(r, base, v26, v25, bold=False, tint=False, no_prior=False):
        """no_prior: 전년 미집계 → '25·GAP은 '—'(0과 비교하면 갭이 허수)."""
        g = v26 - v25
        cells = [(v26, NF_NUM, clr(0)), ("—" if no_prior else v25, NF_NUM, clr(0)),
                 ("—" if no_prior else g, NF_GAP, clr(0 if no_prior else g))]
        for j, (v, nf, col) in enumerate(cells):
            cell = ws.cell(r, base + j, v); cell.alignment = rgt
            if not isinstance(v, str): cell.number_format = nf
            cell.font = F(10, bold=(bold or j == 2), color=col)
            cell.border = topbox(DBL) if bold else box
            if tint: cell.fill = PatternFill("solid", fgColor="F0ECF9")
        return g

    def bsr_cells(r, f26, bfit, bold=False):
        bd = topbox(DBL) if bold else box
        c1 = ws.cell(r, BSR_C, bfit if bfit else "—"); c1.alignment = rgt; c1.border = bd
        c1.font = F(10, bold=bold, color="1F4E78")
        if bfit: c1.number_format = NF_NUM
        ach = (f26 / bfit) if bfit else None
        c2 = ws.cell(r, BSR_C + 1, ach if ach is not None else "—"); c2.alignment = rgt; c2.border = bd
        c2.font = F(10, bold=bold, color="1F4E78")
        if ach is not None: c2.number_format = NF_PCT

    def _no_prior(name):
        a = (anchor or {}).get(name)
        return bool(a and a["has_prior"] is False)

    t_tm26 = t_tm25 = t_hm26 = t_hm25 = t_bsr = 0
    c_tm26 = c_tm25 = c_hm26 = c_hm25 = 0   # 비교가능(전년 보유) 소계
    n_new = 0
    for i, (name, _) in enumerate(TARGETS):
        r = 10 + i
        a26, a25 = tm26.get(name, 0), tm25.get(name, 0)
        h26, h25 = hm26.get(name, 0), hm25.get(name, 0)
        f26, f25 = a26 + h26, a25 + h25; fg = f26 - f25
        npr = _no_prior(name)
        t_tm26 += a26; t_tm25 += a25; t_hm26 += h26; t_hm25 += h25
        if npr:
            n_new += 1
        else:
            c_tm26 += a26; c_tm25 += a25; c_hm26 += h26; c_hm25 += h25
        lab = GAEYO_LBL[name] + ("†" if npr else ("*" if (anchor or {}).get(name, {}).get("src") == "rawdb" else ""))
        lc = ws.cell(r, 2, lab); lc.font = F(10); lc.alignment = lft; lc.border = box
        grp3(r, 3, a26, a25, no_prior=npr); grp3(r, 6, h26, h25, tint=True, no_prior=npr)
        grp3(r, 9, f26, f25, no_prior=npr)
        if has_bsr:
            bfit = bsr["fit26"].get(name, 0); t_bsr += bfit; bsr_cells(r, f26, bfit)
        st = "신규 (전년 미집계)" if npr else ("전년초과 ▲" if fg > 0 else ("동률" if fg == 0 else "전년미달 ▼"))
        sc = ws.cell(r, STAT_C, st); sc.font = F(10, bold=True, color=clr(0 if npr else fg))
        sc.alignment = cen; sc.border = box
    # Total (전체) — 신규 포함 시 전년 비교 불가
    r = 10 + len(TARGETS)
    ft26, ft25 = t_tm26 + t_hm26, t_tm25 + t_hm25; ftg = ft26 - ft25
    lc = ws.cell(r, 2, f"Total (전체 {len(TARGETS)}개)"); lc.font = F(10, bold=True); lc.alignment = lft; lc.border = topbox(DBL)
    grp3(r, 3, t_tm26, t_tm25, bold=True, no_prior=bool(n_new))
    grp3(r, 6, t_hm26, t_hm25, bold=True, tint=True, no_prior=bool(n_new))
    grp3(r, 9, ft26, ft25, bold=True, no_prior=bool(n_new))
    if has_bsr:
        bsr_cells(r, ft26, t_bsr, bold=True)
    stt = "전년 비교 불가" if n_new else ("전년초과 ▲" if ftg > 0 else "전년미달 ▼")
    sc = ws.cell(r, STAT_C, stt); sc.font = F(10, bold=True, color=clr(0 if n_new else ftg))
    sc.alignment = cen; sc.border = topbox(DBL)
    # 비교가능 합계 (전년 보유 사업장만) — 동일 기준 YoY
    if n_new:
        r += 1
        cf26, cf25 = c_tm26 + c_hm26, c_tm25 + c_hm25; cfg = cf26 - cf25
        lc = ws.cell(r, 2, f"비교가능 합계 (전년 보유 {len(TARGETS) - n_new}개)")
        lc.font = F(10, bold=True); lc.alignment = lft; lc.border = box
        grp3(r, 3, c_tm26, c_tm25, bold=True); grp3(r, 6, c_hm26, c_hm25, bold=True, tint=True)
        grp3(r, 9, cf26, cf25, bold=True)
        sc = ws.cell(r, STAT_C, "전년초과 ▲" if cfg > 0 else ("동률" if cfg == 0 else "전년미달 ▼"))
        sc.font = F(10, bold=True, color=clr(cfg)); sc.alignment = cen; sc.border = box
    if anchor:
        note = ("※ 합계(FIT)=사업장별 확정 OTB(15시 기준). GS(OTA+G-OTA)·기타(홈피·제휴·일반)는 자사 구성비로 비례배분 → GS+기타=FIT.  "
                "†=전년 미집계(브랜드 전환) → '25·GAP 및 전체 Total의 전년 비교 제외, 별도 '비교가능 합계' 참조.  *=확정치 미제공 → 자사 실측")
    else:
        note = "※ FIT=영업자료(대) 03일반=GS(OTA+G-OTA)+기타(홈페이지·제휴사·일반). 제휴사·일반 코드는 온라인팀 raw_db엔 거의 없어 전채널 대비 미달"
    ws.cell(r + 1, 2, note).font = F(9, color=GREY)

    # ===== 공용: 사업장별 3컬럼(26Y/25Y/GAP) 추이 시트 =====
    def trend_sheet(title_text, sheet_name, value_fn):
        ws = wb.create_sheet(sheet_name); ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 1.2; ws.column_dimensions["B"].width = 13
        lastc = 2 + len(TARGETS) * 3 + 3
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=lastc)
        ws.cell(2, 2, title_text).font = F(13)
        top, sub = 4, 5
        hcell(ws, top, 2, "기준일" if "누적" in sheet_name else "예약일")
        ws.merge_cells(start_row=top, start_column=2, end_row=sub, end_column=2)
        col = 3; pc = {}
        for name, _ in TARGETS:
            hcell(ws, top, col, DATA_LBL[name]); ws.merge_cells(start_row=top, start_column=col, end_row=top, end_column=col + 2)
            hcell(ws, top, col + 1, None); hcell(ws, top, col + 2, None)
            for j, lab in enumerate(("26Y", "25Y", "GAP")):
                hcell(ws, sub, col + j, lab); ws.column_dimensions[get_column_letter(col + j)].width = 7.5
            pc[name] = col; col += 3
        hcell(ws, top, col, "Total"); ws.merge_cells(start_row=top, start_column=col, end_row=top, end_column=col + 2)
        hcell(ws, top, col + 1, None); hcell(ws, top, col + 2, None)
        for j, lab in enumerate(("26Y", "25Y", "GAP")):
            hcell(ws, sub, col + j, lab); ws.column_dimensions[get_column_letter(col + j)].width = 8
        sc = col; rr = sub + 1
        for day in desc:
            ws.cell(rr, 2, f"{d2k(day)} ({wdc(day)})").font = F(9)
            ws.cell(rr, 2).alignment = cen; ws.cell(rr, 2).border = box
            s26 = s25 = 0
            for name, _ in TARGETS:
                v26, v25 = value_fn(name, day); g = v26 - v25; s26 += v26; s25 += v25
                for j, (v, nf, isg) in enumerate([(v26, NF_NUM, 0), (v25, NF_NUM, 0), (g, NF_GAP, 1)]):
                    cell = ws.cell(rr, pc[name] + j, v); cell.alignment = rgt; cell.border = box
                    cell.number_format = nf; cell.font = F(9, color=(clr(g) if isg else BLACK))
            g = s26 - s25
            for j, (v, nf, isg) in enumerate([(s26, NF_NUM, 0), (s25, NF_NUM, 0), (g, NF_GAP, 1)]):
                cell = ws.cell(rr, sc + j, v); cell.alignment = rgt; cell.border = box
                cell.number_format = nf; cell.font = F(9, bold=True, color=(clr(g) if isg else BLACK))
            rr += 1
        ws.freeze_panes = ws.cell(sub + 1, 3).coordinate

    # BSR 보정: 추이·상세 값도 확정 OTB 계수로 스케일 (누적 최종일 = 확정 OTB와 정확히 일치)
    S = scale or {}
    def s26v(name, v):
        f = S.get(name); return _rescale(v, f["n26"], f["d26"]) if f else v
    def s25v(name, v):
        f = S.get(name); return _rescale(v, f["n25"], f["d25"]) if f else v

    # 보정된 누적 온북(캐시) — 일별 net은 이 곡선의 차분으로 산출해 반올림 오차를 없앤다.
    nb_cache26 = {}; nb_cache25 = {}
    def cum_fn(name, day):
        if day not in nb_cache26: nb_cache26[day] = onbook_at(rows26, day)
        d25 = to25(day)
        if d25 not in nb_cache25: nb_cache25[d25] = onbook_at(rows25, d25)
        return s26v(name, nb_cache26[day].get(name, 0)), s25v(name, nb_cache25[d25].get(name, 0))
    def prevd(day): return (datetime.strptime(day, "%Y%m%d") - timedelta(days=1)).strftime("%Y%m%d")
    def net_fn(name, day):
        c, p = cum_fn(name, day), cum_fn(name, prevd(day))
        return c[0] - p[0], c[1] - p[1]

    # 일별 픽업 (증감): 올해 일별 net vs 전년 동일자
    trend_sheet("일별 픽업 (증감) — 7월 투숙, 전년 동기간 比", "일별픽업(증감)", net_fn)
    # 전년대비 누적 온북 추이
    trend_sheet("전년 동기간 대비 누적 온북 추이 (일자별 cutoff, 확정 OTB 보정)", "전년대비(누적추이)", cum_fn)

    # ============================== 상세 (오름차순) ==============================
    ws4 = wb.create_sheet("상세(신규·취소)"); ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 1.2
    ws4.merge_cells("B2:G2")
    ws4.cell(2, 2, "일자×사업장 상세 — 신규 / 취소 / net / 당일 누적온북 (2026 7월 투숙)").font = F(13)
    for i, (h, w) in enumerate([("예약 유입일", 16), ("사업장", 18), ("신규(+)", 10), ("취소(−)", 10), ("net", 9), ("당일 누적온북", 13)]):
        hcell(ws4, 4, 2 + i, h); ws4.column_dimensions[get_column_letter(2 + i)].width = w
    rr = 5
    for day in days26:  # 과거 → 최근
        nbd = onbook_at(rows26, day)
        for name, _ in TARGETS:
            # net은 보정 누적의 차분으로 확정, 신규만 보정 후 취소를 역산 → 당일 누적온북과 정합
            net = net_fn(name, day)[0]
            nw = s26v(name, new26.get((name, day), 0)); cx = nw - net
            row = [(f"{d2k(day)} ({wdc(day)})", cen, None, F(9)), (DATA_LBL[name], lft, None, F(9)),
                   (nw, rgt, NF_NUM, F(9)), (-cx, rgt, NF_NUM, F(9)),
                   (net, rgt, NF_NUM, F(9, bold=True, color=clr(net))), (s26v(name, nbd.get(name, 0)), rgt, NF_NUM, F(9))]
            for i, (v, al, nf, ft) in enumerate(row):
                cell = ws4.cell(rr, 2 + i, v); cell.alignment = al; cell.font = ft; cell.border = box
                if nf: cell.number_format = nf
            rr += 1
    ws4.freeze_panes = "A5"

    wb.save(out_path)

# ───────────────────────── JSON (대시보드 HTML용) ─────────────────────────
def build_payload(data_date, asof26, asof25, rows26, rows25, anchor=None, scale=None):
    """세그먼트(OTA/G-OTA/Inbound/기타)별로 분해해 payload 생성.
    HTML이 선택된 세그를 합산해 표/총평을 재계산(기본 OTA+G-OTA)."""
    hi = datetime.strptime(asof26, "%Y%m%d")
    days26 = [(hi - timedelta(days=k)).strftime("%Y%m%d") for k in range(WINDOW_DAYS)][::-1]
    ordered = list(reversed(days26))  # 최근 → 과거
    def to25(d): return "2025" + d[4:]
    WD = "월화수목금토일"
    def wdc(day): return WD[datetime.strptime(day, "%Y%m%d").weekday()]

    # 세그별 동기간 net 온북: {prop: {seg: net_rn}}
    def onbook_seg(rows, cutoff):
        out = defaultdict(lambda: defaultdict(int))
        for r in rows:
            if r["entry"] and r["entry"] <= cutoff:
                out[r["prop"]][r["seg"]] += r["rn"]
            if r["cancel"] and r["cancel"] <= cutoff:
                out[r["prop"]][r["seg"]] -= r["rn"]
        return out
    # 세그별 일별 신규/취소: {(prop,seg,day): rn}
    def daily_seg(rows):
        new = defaultdict(int); cxl = defaultdict(int)
        for r in rows:
            if r["entry"]:  new[(r["prop"], r["seg"], r["entry"])] += r["rn"]
            if r["cancel"]: cxl[(r["prop"], r["seg"], r["cancel"])] += r["rn"]
        return new, cxl

    def segmap(d): return {s: d.get(s, 0) for s in SEGMENTS}

    on26 = onbook_seg(rows26, asof26); on25 = onbook_seg(rows25, asof25)
    new26, cxl26 = daily_seg(rows26); new25, cxl25 = daily_seg(rows25)

    # 요약: 사업장별 세그별 온북 (HTML이 선택 세그 합산)
    # anchor가 있으면 FIT 세그는 BSR 정합값(seg26/seg25)으로 치환 → 합계가 BSR과 일치.
    # Inbound·기타는 BSR FIT 밖이므로 raw_db 실측 유지.
    summary = []
    for n, l in TARGETS:
        s26, s25 = segmap(on26.get(n, {})), segmap(on25.get(n, {}))
        item = {"name": n, "label": l, "on26": s26, "on25": s25}
        if anchor and n in anchor:
            a = anchor[n]
            for s in FIT_SEGMENTS:
                s26[s] = a["seg26"][s]; s25[s] = a["seg25"][s]
            item["anchor"] = {k: a[k] for k in ("fit26", "fit25", "gap", "raw26", "raw25", "src", "has_prior")}
        summary.append(item)

    # BSR 보정 계수 — FIT 세그(GS·기타)에만 적용. Inbound·기타는 FIT 밖이라 실측 유지.
    SC = scale or {}
    def k26(n, s, v):
        f = SC.get(n); return _rescale(v, f["n26"], f["d26"]) if (f and s in FIT_SEGMENTS) else v
    def k25(n, s, v):
        f = SC.get(n); return _rescale(v, f["n25"], f["d25"]) if (f and s in FIT_SEGMENTS) else v

    # 보정된 누적 온북: {day: {prop: {seg: [온북26, 온북25]}}}
    # 최종일(asof26)의 FIT 세그는 앵커 값으로 고정 → 요약·확정 OTB와 정확히 일치.
    prev_day = (datetime.strptime(days26[0], "%Y%m%d") - timedelta(days=1)).strftime("%Y%m%d")
    cum_days = [prev_day] + days26          # 과거 → 최근 (첫날 net 계산용으로 하루 더)
    cum_scaled = {}
    for day in cum_days:
        c26 = onbook_seg(rows26, day); c25 = onbook_seg(rows25, to25(day))
        last = (day == asof26)
        per = {}
        for n, _ in TARGETS:
            a = (anchor or {}).get(n)
            row = {}
            for s in SEGMENTS:
                v26 = c26.get(n, {}).get(s, 0); v25 = c25.get(n, {}).get(s, 0)
                if last and a and s in FIT_SEGMENTS:
                    row[s] = [a["seg26"][s], a["seg25"][s] if a["has_prior"] else k25(n, s, v25)]
                else:
                    row[s] = [k26(n, s, v26), k25(n, s, v25)]
            per[n] = row
        cum_scaled[day] = per

    # 일별 픽업: per[prop][seg]=[신규, 취소].
    # net(=신규−취소)은 보정 누적의 차분으로 확정하고, 신규만 계수 보정한 뒤 취소를 역산한다.
    # → 일별 net 합계가 누적 곡선과 정확히 일치(반올림 오차 0).
    def _daily(day, prev, newmap, cxlmap, key_day, kfn, idx):
        per = {}
        for n, _ in TARGETS:
            row = {}
            for s in SEGMENTS:
                net = cum_scaled[day][n][s][idx] - cum_scaled[prev][n][s][idx]
                nw = kfn(n, s, newmap.get((n, s, key_day), 0))
                row[s] = [nw, nw - net]
            per[n] = row
        return per

    daily = []
    for day in ordered:                      # 최근 → 과거
        prev = (datetime.strptime(day, "%Y%m%d") - timedelta(days=1)).strftime("%Y%m%d")
        d25 = to25(day)
        daily.append({"d": day, "wd": wdc(day), "d25": d25,
                      "per":   _daily(day, prev, new26, cxl26, day, k26, 0),
                      "per25": _daily(day, prev, new25, cxl25, d25, k25, 1)})

    cumulative = [{"d": day, "wd": wdc(day), "per": cum_scaled[day]} for day in ordered]

    return {
        "meta": {"data_date": data_date, "asof26": asof26, "asof25": asof25,
                 "window_days": WINDOW_DAYS,
                 "segments": SEGMENTS, "default_segments": DEFAULT_SEGMENTS,
                 "team_segments": TEAM_SEGMENTS, "home_segments": HOME_SEGMENTS, "fit_segments": FIT_SEGMENTS,
                 "targets": [{"name": n, "label": l} for n, l in TARGETS]},
        "summary": summary, "daily": daily, "cumulative": cumulative,
    }


# ───────────────────────── main ─────────────────────────
def main():
    data_date = snapshot_date()                       # 예: 20260630
    asof26 = (datetime.strptime(data_date, "%Y%m%d") - timedelta(days=1)).strftime("%Y%m%d")  # 마지막 완전일
    asof25 = "2025" + asof26[4:]
    print(f"스냅샷일={data_date}  동기간 기준(완전일)={asof26} / 전년={asof25}")

    print("2026 7월 투숙 로딩...", flush=True)
    rows26 = load_rows(live_files(2026), "202607")
    print(f"  rows={len(rows26):,}")
    print("2025 7월 투숙 로딩...", flush=True)
    rows25 = load_rows(retrans_files(2025) + live_files(2025), "202507")
    print(f"  rows={len(rows25):,}")

    nb26 = onbook_at(rows26, asof26); nb25 = onbook_at(rows25, asof25)
    print("\n[검증] 동기간 온북 갭 (세그 구분없이 전체 RN):")
    print(f"  {'사업장':18}{'26온북':>8}{'25온북':>8}{'갭':>8}")
    t26 = t25 = 0
    for name, lab in TARGETS:
        v26, v25 = nb26.get(name, 0), nb25.get(name, 0); t26 += v26; t25 += v25
        print(f"  {lab:18}{v26:>8,}{v25:>8,}{v26 - v25:>8,}")
    print(f"  {'합계':18}{t26:>8,}{t25:>8,}{t26 - t25:>8,}")

    # 1) 대시보드 배포본 (docs/data): HTML이 fetch + 엑셀 다운로드
    docs_data = PROJECT_DIR / "docs" / "data"
    docs_data.mkdir(parents=True, exist_ok=True)
    xlsx_docs = docs_data / "july_pickup.xlsx"
    # 엑셀은 기본 세그(OTA+G-OTA) 기준 — 화면 기본 뷰와 동일. JSON은 전 세그 보존(화면 필터용).
    xl_seg = set(DEFAULT_SEGMENTS); xl_label = "+".join(DEFAULT_SEGMENTS)
    rows26_xl = [r for r in rows26 if r["seg"] in xl_seg]
    rows25_xl = [r for r in rows25 if r["seg"] in xl_seg]
    bsr = parse_bsr()   # 최신 BSR PDF(15시)의 사업장별 FIT OTB — 로컬 실행 시에만

    # FIT 세그 raw_db 실측(비례배분 가중치 + 정합률 산출용)
    def _onbook_seg(rows, cutoff):
        out = defaultdict(lambda: defaultdict(int))
        for r in rows:
            if r["seg"] not in FIT_SEGMENTS: continue
            if r["entry"] and r["entry"] <= cutoff:  out[r["prop"]][r["seg"]] += r["rn"]
            if r["cancel"] and r["cancel"] <= cutoff: out[r["prop"]][r["seg"]] -= r["rn"]
        return out
    fon26 = _onbook_seg(rows26, asof26); fon25 = _onbook_seg(rows25, asof25)
    anchor = bsr_anchor(bsr, fon26, fon25)

    if bsr:
        print(f"\n[BSR 정합] {bsr['date']} 15시 OTB  ↔  전년 {bsr['prior_date']}  (7월 투숙 FIT)")
        print(f"  {'사업장':22}{'2026':>8}{'2025':>8}{'차이':>8}{'raw_db26':>10}{'정합률':>8}")
        A26 = A25 = 0
        for name, lab in TARGETS:
            a = anchor[name]; A26 += a["fit26"]; A25 += a["fit25"]
            rate = f"{a['raw26']/a['fit26']*100:5.1f}%" if a["fit26"] else "    —"
            flag = "" if a["src"] == "bsr" else "  (raw_db)"
            if a["src"] == "bsr" and not a["has_prior"]: flag = "  (전년 미집계)"
            print(f"  {lab:22}{a['fit26']:>8,}{a['fit25']:>8,}{a['gap']:>+8,}{a['raw26']:>10,}{rate:>8}{flag}")
        print(f"  {'합계':22}{A26:>8,}{A25:>8,}{A26-A25:>+8,}")
        bt = bsr["total"]
        bsr_sum26 = sum(bsr["fit26"].values())
        print(f"  · BSR 사업장 합(우리 25개 매칭분) = {bsr_sum26:,} / BSR Total 페이지 = {bt['fit26']:,}")
    else:
        print("BSR: PDF 없음/파싱불가 — raw_db 실측만 사용")

    scale = bsr_scale(anchor)
    build_excel(str(xlsx_docs), data_date, asof26, asof25, rows26_xl, rows25_xl,
                seg_label=xl_label, bsr=bsr, anchor=anchor, scale=scale)
    payload = build_payload(data_date, asof26, asof25, rows26, rows25, anchor=anchor, scale=scale)
    payload["meta"]["scaled"] = bool(bsr)
    if bsr:
        payload["bsr"] = {"date": bsr["date"], "prior_date": bsr["prior_date"],
                          "fit26": bsr["fit26"], "fit25": bsr["fit25"],
                          "gap": bsr["gap"], "has_prior": bsr["has_prior"],
                          "total": bsr["total"]}
    json_docs = docs_data / "july_pickup.json"
    json.dump(payload, open(json_docs, "w", encoding="utf-8"), ensure_ascii=False)
    print(f"\n대시보드 배포본 → {json_docs}\n               → {xlsx_docs}")

    # 2) 사용자 로컬 사본 (Desktop, 날짜 파일명) — 자동 파이프라인(daily_update.sh)에선 생략
    desk = Path(os.path.expanduser("~/Desktop"))
    if os.environ.get("PICKUP_NO_DESKTOP"):
        print("로컬 사본       → (PICKUP_NO_DESKTOP: 생략)")
    elif desk.is_dir():
        out_path = desk / f"7월동기간_픽업_일자별관리_{data_date}.xlsx"
        shutil.copyfile(xlsx_docs, out_path)
        print(f"로컬 사본       → {out_path}")

if __name__ == "__main__":
    main()
