#!/usr/bin/env python3
"""
CRM 코드 사전(데이터 딕셔너리) 추출: 27/28/43/44 raw_db 전 연도 스캔
- CRM에 쓰일 수 있는 구분/코드 항목별로 (코드, 명칭) 도메인 + 파일타입별 등장 빈도 집계.
- 코드 컬럼 명칭이 타입별로 다름(27/43=변경예약집계코드/원예약집계코드, 28/44=예약집계코드) → variants로 흡수.
- 빈도는 행 단위(원본, 중복 미제거) — 목적은 '값 도메인 열거'이므로 distinct 값에는 영향 없음.
- 출력: 개요, 필드맵(컬럼 존재 매트릭스), 항목별 시트.
"""
import sys, glob, os
from pathlib import Path
from collections import defaultdict, Counter

PROJECT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT / "scripts"))
from parse_raw_db import classify_segment

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path.home() / "Desktop" / "CRM_코드사전_27_28_43_44.xlsx"
FTYPES = ["27", "28", "43", "44"]
FT_IDX = {t: i for i, t in enumerate(FTYPES)}

# ─── CRM 차원 정의 ─────────────────────────────────────────────
# (시트명, [ (코드컬럼 or None, 명칭컬럼) variants ... ], 설명, 세그먼트해석여부)
DIMS = [
    ("예약집계코드(현행)", [("변경예약집계코드", "변경예약집계코드명"), ("예약집계코드", "예약집계명")],
     "예약 집계 분류코드(현행/변경). 세그먼트 분류의 핵심 키. 27/43=변경예약집계코드, 28/44=예약집계코드", True),
    ("원예약집계코드", [("원예약집계코드", "원예약집계명")],
     "원(최초) 예약 집계 분류코드. 27/43에만 존재", True),
    ("예약구분", [("예약구분", "예약구분명")], "예약 구분 코드/명", False),
    ("사용구분", [("사용구분", "사용구분명")], "사용 구분 코드/명", False),
    ("고객유형", [(None, "고객유형")], "고객 유형 (CRM 핵심 분류)", False),
    ("단체구분", [(None, "단체구분")], "단체/개인 등 단체 구분", False),
    ("예약상태", [(None, "예약상태")], "예약 상태(노쇼 등)", False),
    ("투숙여부", [(None, "투숙여부")], "투숙 여부", False),
    ("선결제여부", [(None, "선결제여부")], "선결제 여부", False),
    ("시즌", [("시즌코드", "시즌명")], "시즌 코드/명", False),
    ("요금시즌", [("요금시즌코드", "요금시즌명")], "요금 시즌 코드/명", False),
    ("패키지분류", [("패키지분류코드", "패키지분류코드명")], "패키지 분류 코드/명 (27/28만)", False),
    ("객실타입", [(None, "객실타입")], "객실 타입", False),
    ("평형", [(None, "평형")], "평형", False),
    ("동", [(None, "동")], "동(건물)", False),
    ("타입", [(None, "타입")], "타입", False),
    ("거래처(AGENT)", [("AGENT코드", "AGENT명")], "거래처/채널 AGENT 코드/명", False),
    ("영업장", [("영업장", "영업장명")], "영업장(사업장) 코드/명", False),
    ("변경사업장명", [(None, "변경사업장명")], "정규화 사업장명(변경)", False),
]

# dim별 집계: key=(code, name) -> [c27,c28,c43,c44]
dim_data = {d[0]: defaultdict(lambda: [0, 0, 0, 0]) for d in DIMS}
# 컬럼 존재 매트릭스: col -> set(filetypes)
col_presence = defaultdict(set)
file_log = []

def detect_ft(name):
    b = os.path.basename(name)
    for t in FTYPES:
        if b.startswith(t + "."):
            return t
    return None

files = []
for y in ["2022", "2023", "2024", "2025", "2026"]:
    for fp in sorted(glob.glob(str(PROJECT / f"data/raw_db/{y}/*.txt"))):
        if detect_ft(fp):
            files.append(fp)

print(f"스캔 대상 파일: {len(files)}개")
for fp in files:
    ft = detect_ft(fp)
    fi = FT_IDX[ft]
    enc_ok = None
    for enc in ["cp949", "euc-kr", "utf-8"]:
        try:
            with open(fp, encoding=enc) as f:
                header = [h.strip() for h in f.readline().rstrip("\n\r").split(";")]
            enc_ok = enc
            break
        except UnicodeDecodeError:
            continue
    if not enc_ok:
        print("  ENC FAIL:", fp); continue
    ix = {h: i for i, h in enumerate(header)}
    for h in header:
        col_presence[h].add(ft)

    # 이 파일에서 각 dim이 사용할 (code_idx, name_idx) 결정
    resolved = {}
    for sheet, variants, _desc, _seg in DIMS:
        for code_col, name_col in variants:
            if name_col in ix and (code_col is None or code_col in ix):
                ci = ix[code_col] if code_col else -1
                ni = ix[name_col]
                resolved[sheet] = (ci, ni)
                break
    maxidx = max([max(c, n) for c, n in resolved.values()] + [0])

    rows = 0
    with open(fp, encoding=enc_ok) as f:
        f.readline()
        for line in f:
            p = line.rstrip("\n\r").split(";")
            if len(p) <= maxidx:
                continue
            rows += 1
            for sheet, (ci, ni) in resolved.items():
                name = p[ni].strip()
                code = p[ci].strip() if ci >= 0 else ""
                if not code and not name:
                    continue
                dim_data[sheet][(code, name)][fi] += 1
    file_log.append((os.path.basename(fp), ft, enc_ok, rows))
    print(f"  [{ft}] {rows:>9,} rows  {os.path.basename(fp)[:55]}")

# ─── 엑셀 작성 ─────────────────────────────────────────────
wb = openpyxl.Workbook()
HF = PatternFill("solid", fgColor="1F4E78"); HFONT = Font(bold=True, color="FFFFFF")
SUBF = PatternFill("solid", fgColor="FCE4D6"); BOLD = Font(bold=True)
THIN = Side(style="thin", color="D0D0D0"); BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CEN = Alignment(horizontal="center", vertical="center"); WRAP = Alignment(wrap_text=True, vertical="top")

def hdr_style(ws, ncol, row=1):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c); cell.fill = HF; cell.font = HFONT; cell.alignment = CEN; cell.border = BORD

def safe_title(s):
    for ch in '[]:*?/\\':
        s = s.replace(ch, "")
    return s[:31]

# 시트1: 개요
ws = wb.active; ws.title = "개요"
ws.append(["CRM 코드 사전 — 27/28/43/44 raw_db 전 연도(2022~2026) 스캔"])
ws.cell(row=1, column=1).font = Font(bold=True, size=13)
ws.append([])
ws.append(["#", "항목(시트)", "코드컬럼", "고유값수", "27", "28", "43", "44", "설명"])
hdr_style(ws, 9, row=3)
r = 4
for i, (sheet, variants, desc, seg) in enumerate(DIMS, 1):
    d = dim_data[sheet]
    nuniq = len(d)
    tots = [sum(v[k] for v in d.values()) for k in range(4)]
    has_code = "Y" if variants[0][0] else "—"
    ws.append([i, sheet, has_code, nuniq] + [t if t else "" for t in tots] + [desc])
    for c in range(1, 10):
        ws.cell(row=r, column=c).border = BORD
    ws.cell(row=r, column=9).alignment = WRAP
    r += 1
ws.append([])
ws.append(["기준", "빈도=행 단위 원본 카운트(중복 미제거). 목적=값 도메인 열거. '—'=해당 타입에 컬럼 없음/값 없음"])
ws.append(["세그먼트해석", "예약집계코드/원예약집계코드 시트의 '세그먼트해석' = parse_raw_db classify_segment (53/72→OTA, A4/A5→G-OTA, 58→Inbound)"])
ws.column_dimensions["A"].width = 4; ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 9; ws.column_dimensions["D"].width = 9
for col in "EFGH": ws.column_dimensions[col].width = 9
ws.column_dimensions["I"].width = 70
ws.freeze_panes = "A4"

# 시트2: 필드맵 (전체 컬럼 × 타입 존재)
ws2 = wb.create_sheet("필드맵")
ws2.append(["컬럼명", "27", "28", "43", "44", "타입수"])
hdr_style(ws2, 6)
# 원본 헤더 순서 보존: 27 헤더 우선 + 나머지 추가
order = []
seen = set()
for fp in files:
    ft = detect_ft(fp)
    for enc in ["cp949", "euc-kr", "utf-8"]:
        try:
            with open(fp, encoding=enc) as f:
                hdr = [h.strip() for h in f.readline().rstrip("\n\r").split(";")]
            break
        except UnicodeDecodeError:
            continue
    for h in hdr:
        if h not in seen:
            seen.add(h); order.append(h)
rr = 2
for h in order:
    pres = col_presence[h]
    ws2.append([h] + ["O" if t in pres else "" for t in FTYPES] + [len(pres)])
    for c in range(1, 7):
        ws2.cell(row=rr, column=c).border = BORD
        if c >= 2: ws2.cell(row=rr, column=c).alignment = CEN
    rr += 1
ws2.column_dimensions["A"].width = 24
for col in "BCDE": ws2.column_dimensions[col].width = 6
ws2.column_dimensions["F"].width = 8
ws2.freeze_panes = "A2"

# 항목별 시트
for sheet, variants, desc, seg in DIMS:
    ws = wb.create_sheet(safe_title(sheet))
    has_code = variants[0][0] is not None
    cols = (["코드"] if has_code else []) + ["명칭"] + (["세그먼트해석"] if seg else []) + ["27", "28", "43", "44", "합계"]
    ws.append([desc]); ws.cell(row=1, column=1).font = Font(italic=True, color="666666")
    ws.append(cols)
    hdr_style(ws, len(cols), row=2)
    # 정렬: 합계 desc
    items = sorted(dim_data[sheet].items(), key=lambda kv: -sum(kv[1]))
    r = 3
    for (code, name), cnt in items:
        total = sum(cnt)
        row = []
        if has_code: row.append(code)
        row.append(name)
        if seg:
            row.append(classify_segment(code, name, "", ""))
        row += [c if c else "" for c in cnt] + [total]
        ws.append(row)
        for c in range(1, len(cols) + 1):
            ws.cell(row=r, column=c).border = BORD
        r += 1
    # 너비
    ci = 1
    if has_code: ws.column_dimensions[get_column_letter(ci)].width = 14; ci += 1
    ws.column_dimensions[get_column_letter(ci)].width = 30; ci += 1
    if seg: ws.column_dimensions[get_column_letter(ci)].width = 12; ci += 1
    for _ in range(5):
        ws.column_dimensions[get_column_letter(ci)].width = 9; ci += 1
    ws.freeze_panes = "A3"

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"\n저장: {OUT}")
print(f"항목 수: {len(DIMS)}  |  스캔 파일: {len(files)}")
for sheet, *_ in [(d[0],) for d in DIMS]:
    print(f"  {sheet}: 고유 {len(dim_data[sheet])}개")
