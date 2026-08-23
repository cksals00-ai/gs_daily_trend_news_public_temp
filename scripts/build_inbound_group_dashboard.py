#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
인바운드 단체(1박 20실 이상) 5개년 실적 대시보드 (43번 예약자료 기준)
================================================================================
소스: data/raw_db/20*/43.*  (Inbound=변경예약집계코드 58). 43(예약)만, 취소 미차감.
필터: 세그먼트=Inbound AND 행(=예약-투숙일) 객실수 >= 20  (1박당 20객실 이상 그룹)
지표: 객실수(실)=RN 합, ADR(천원)=공급가/RN/1000, REV(백만)=공급가/1e6
      공급가 = 1박객실료 × 객실수 ÷ 1.1 (VAT 제외, 프로젝트 관례)
차원: 연도(판매일자[:4]) · 월 · 국적(회원명 괄호 추출) · 여행사(회원명 정리)
출력: ~/Desktop/인바운드_단체20실_5개년_대시보드.xlsx
      시트: 대시보드 / 여행사별 / 국적별 / 연월별 / DB / 기준
양식: 첨부 고객지표 대시보드 템플릿 시각문법(헤더바·회색라인표·데이터바·라인/스택바) 준용, 폰트=대명체.
"""
import glob, sys, re
from pathlib import Path
from collections import defaultdict, Counter

sys.path.insert(0, "scripts")
from parse_raw_db import classify_segment

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.formatting.rule import DataBarRule

OUT = Path.home() / "Desktop" / "인바운드_단체20실_5개년_대시보드.xlsx"
FN = "대명체"  # 폰트: 대명체 보통

# ── 국적/권역 매핑 ──
COUNTRIES = ['대만','중국','일본','홍콩','마카오','말레이시아','베트남','태국','필리핀',
             '싱가포르','싱가폴','인도네시아','몽골','인도','캄보디아','라오스','미얀마',
             '유럽','미주','미국','호주','뉴질랜드','러시아','중동','다국적']
CO_NORM = {'미국':'미주','싱가폴':'싱가포르'}
REGION = {'대만':'중화권','중국':'중화권','홍콩':'중화권','마카오':'중화권','일본':'일본',
          '말레이시아':'동남아','베트남':'동남아','태국':'동남아','필리핀':'동남아','싱가포르':'동남아',
          '인도네시아':'동남아','캄보디아':'동남아','라오스':'동남아','미얀마':'동남아',
          '몽골':'기타아시아','인도':'기타아시아','유럽':'구미주','미주':'구미주','호주':'구미주',
          '뉴질랜드':'구미주','러시아':'기타','중동':'기타','다국적':'기타','동남아(미상)':'동남아','미상':'기타'}

def extract_country(memn, danche):
    for pc in re.findall(r'\((.*?)\)', memn):
        for c in COUNTRIES:
            if c in pc:
                return CO_NORM.get(c, c)
    for c in COUNTRIES:
        if c in memn:
            return CO_NORM.get(c, c)
    if '동남아' in danche:
        return '동남아(미상)'
    if '일본' in danche:
        return '일본'
    return '미상'

def clean_agency(memn):
    s = memn
    for cut in ['(', '_', ' - ']:
        i = s.find(cut)
        if i > 0:
            s = s[:i]
    # 공백 뒤 국가 키워드 절단 (예: '오션인터내셔널 중국 수정제약' → '오션인터내셔널')
    for c in COUNTRIES:
        i = s.find(' ' + c)
        if i > 0:
            s = s[:i]
    return s.strip() or memn

# ── 데이터 스캔 ──
rows = []  # dict per qualifying booking-night
for fp in sorted(glob.glob("data/raw_db/20*/43.*")):
    with open(fp, encoding="cp949") as f:
        hdr = [h.strip() for h in f.readline().split(";")]
        ix = {h: i for i, h in enumerate(hdr)}
        for line in f:
            p = line.rstrip("\n\r").split(";")
            if len(p) < len(hdr):
                continue
            if classify_segment(p[ix['변경예약집계코드']].strip(), '', '', '') != 'Inbound':
                continue
            rooms = int(p[ix['객실수']].strip() or 0)
            if rooms < 20:
                continue
            sd = p[ix['판매일자']].strip()
            if len(sd) < 6:
                continue
            memn = p[ix['회원명']].strip()
            dan = p[ix['단체구분']].strip()
            rate = int(p[ix['1박객실료']].strip() or 0)
            rev = int(rate * rooms / 1.1)
            country = extract_country(memn, dan)
            rows.append({
                'year': sd[:4], 'ym': sd[:6], 'month': int(sd[4:6]), 'date': sd[:8],
                'member': memn, 'agency': clean_agency(memn), 'country': country,
                'region': REGION.get(country, '기타'), 'danche': dan,
                'rn': rooms, 'rate': rate, 'rev': rev,
                'prop': (p[ix['변경사업장명']].strip() or p[ix['영업장명']].strip()),
                'key': p[ix['KEY_RSV_NO']].strip(),
            })

YEARS = sorted({r['year'] for r in rows})
print(f"대상 booking-night: {len(rows):,}  연도: {YEARS}")

def agg(dim):
    d = defaultdict(lambda: {'rn': 0, 'rev': 0, 'cnt': 0})
    for r in rows:
        k = dim(r)
        d[k]['rn'] += r['rn']; d[k]['rev'] += r['rev']; d[k]['cnt'] += 1
    return d

by_year = agg(lambda r: r['year'])
by_ym = agg(lambda r: (r['year'], r['month']))
by_month = agg(lambda r: r['month'])
by_country = agg(lambda r: r['country'])
by_country_year = agg(lambda r: (r['country'], r['year']))
by_country_month = agg(lambda r: (r['country'], r['month']))
by_agency = agg(lambda r: r['agency'])
by_agency_year = agg(lambda r: (r['agency'], r['year']))

tot_rn = sum(r['rn'] for r in rows)
tot_rev = sum(r['rev'] for r in rows)
tot_cnt = len(rows)
def adr_k(rev, rn):
    return round(rev / rn / 1000) if rn else 0
def rev_m(rev):
    return round(rev / 1e6, 1)

# 여행사 대표국적
ag_country = {}
for a in by_agency:
    c = Counter()
    for r in rows:
        if r['agency'] == a:
            c[r['country']] += r['rn']
    ag_country[a] = c.most_common(1)[0][0] if c else ''

countries_sorted = sorted(by_country, key=lambda c: -by_country[c]['rn'])
agencies_sorted = sorted(by_agency, key=lambda a: -by_agency[a]['rn'])
TOP_NAT = countries_sorted[:5]

# ================= 엑셀 =================
wb = openpyxl.Workbook()
# 워크북 기본(Normal) 스타일 폰트를 대명체로 → 미지정 셀까지 전부 대명체 적용
try:
    wb._named_styles['Normal'].font = Font(name=FN, size=10)
except Exception:
    pass

# 팔레트 (템플릿 근접 + 약간 리프레시)
C_TITLE = "463F3A"   # 타이틀/헤더 바 (dark taupe)
C_HEAD  = "6E6259"   # 표 헤더
C_CARD  = "8A7A6D"   # 카드 헤더
C_BAND  = "EFEAE5"   # 밴드/소계
C_BAND2 = "F6F3F0"
C_LINE  = "CFC7BF"   # 표 라인 (연회색)
C_ACC   = "9C8878"   # 데이터바/강조
C_RN    = "6E6259"   # 차트 RN
C_REV   = "B08968"   # 차트 REV
NAT_COLORS = ["6E6259", "B08968", "A3A0A0", "C4A484", "8A9A8B", "7D6B5D"]

thin = Side(style="thin", color=C_LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
def F(size=10, bold=False, color="2B2B2B"):
    return Font(name=FN, size=size, bold=bold, color=color)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

def cell(ws, r, c, v=None, font=None, fill=None, align=None, border=True, numfmt=None):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font = font or F()
    if fill:
        cl.fill = PatternFill("solid", fgColor=fill)
    cl.alignment = align or LEFT
    if border:
        cl.border = BORDER
    if numfmt:
        cl.number_format = numfmt
    return cl

def merge_cell(ws, r1, c1, r2, c2, v, font=None, fill=None, align=None, border=True):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cl = cell(ws, r1, c1, v, font, fill, align or CEN, border)
    if border:
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(row=rr, column=cc).border = BORDER
    return cl

def section(ws, r, c1, c2, title):
    """섹션 제목 + 하단 회색 라인 (템플릿의 밑줄 제목)."""
    cl = ws.cell(row=r, column=c1, value=title); cl.font = F(13, True, C_TITLE); cl.alignment = LEFT
    for cc in range(c1, c2 + 1):
        ws.cell(row=r, column=cc).border = Border(bottom=Side(style="medium", color=C_ACC))

NUM = "#,##0"
NUM1 = "#,##0.0"

# ---------------- 시트1: 대시보드 ----------------
ws = wb.create_sheet("대시보드")
ws.sheet_view.showGridLines = False
LASTCOL = 16
ws.column_dimensions["A"].width = 2.5
for c in range(2, LASTCOL + 1):
    ws.column_dimensions[get_column_letter(c)].width = 10.5

# 타이틀 바
merge_cell(ws, 2, 2, 3, 11, "[ 인바운드 단체 · 1박 20실 이상 ]  5개년 여행사·국적별 실적 현황",
           F(16, True, "FFFFFF"), C_TITLE, Alignment(horizontal="left", vertical="center", indent=1))
# 필터형 박스 (장식)
def pill(col, label, val):
    merge_cell(ws, 2, col, 3, col, label, F(9, True, "FFFFFF"), C_HEAD, CEN)
    merge_cell(ws, 2, col + 1, 3, col + 2, val, F(9, True, "3B3B3B"), "FFFFFF", CEN)
pill(12, "기간", f"{YEARS[0]}~{YEARS[-1]}")
ws.row_dimensions[2].height = 16
ws.row_dimensions[3].height = 16

r = 5
section(ws, r, 2, 11, "종합 요약"); r += 1
# KPI 카드 (라벨행 + 값행)
cards = [
    ("객실수(실) 누계", f"{tot_rn:,}", "실"),
    ("REV 누계(백만)", f"{rev_m(tot_rev):,.1f}", "백만원"),
    ("평균 ADR(천원)", f"{adr_k(tot_rev, tot_rn):,}", "천원"),
    ("케이스(1박20실↑)", f"{tot_cnt:,}", "건"),
    ("여행사 수", f"{len(by_agency):,}", "개사"),
    ("국적 수", f"{len(by_country):,}", "개국"),
]
cw = 3  # 카드 폭(열)
c0 = 2
for i, (lab, val, unit) in enumerate(cards):
    cc = c0 + (i % 3) * cw
    rr = r + (i // 3) * 3
    merge_cell(ws, rr, cc, rr, cc + cw - 1, lab, F(10, True, "FFFFFF"), C_CARD, CEN)
    merge_cell(ws, rr + 1, cc, rr + 1, cc + cw - 1, val, F(20, True, C_TITLE), "FFFFFF", CEN)
    merge_cell(ws, rr + 2, cc, rr + 2, cc + cw - 1, unit, F(8, False, "8A8A8A"), C_BAND2, CEN)
    ws.row_dimensions[rr + 1].height = 26
r = r + 6 + 1

# ── 연도별 실적 추이 (표 + 라인차트) ──
section(ws, r, 2, 11, "연도별 실적 추이"); r += 1
hdr_row = r
heads = ["연도", "객실수(실)", "REV(백만)", "ADR(천원)", "케이스", "전년比RN"]
for j, h in enumerate(heads):
    merge_cell(ws, r, 2 + j, r, 2 + j, h, F(10, True, "FFFFFF"), C_HEAD, CEN)
r += 1
data_start = r
prev = None
for y in YEARS:
    d = by_year[y]
    yoy = (d['rn'] - prev) if prev is not None else None
    cell(ws, r, 2, y + ("(부분)" if y == "2026" else ""), F(10, True), C_BAND, LEFT)
    cell(ws, r, 3, d['rn'], F(10), None, RIGHT, numfmt=NUM)
    cell(ws, r, 4, rev_m(d['rev']), F(10), None, RIGHT, numfmt=NUM1)
    cell(ws, r, 5, adr_k(d['rev'], d['rn']), F(10), None, RIGHT, numfmt=NUM)
    cell(ws, r, 6, d['cnt'], F(10), None, RIGHT, numfmt=NUM)
    cell(ws, r, 7, ("—" if yoy is None else f"{'▲' if yoy>=0 else '▼'} {abs(yoy):,}"),
         F(10, False, ("9C0006" if (yoy or 0) < 0 else "1F6E43")), None, RIGHT)
    prev = d['rn']
    r += 1
# 합계행
cell(ws, r, 2, "합계", F(10, True), C_BAND, LEFT)
cell(ws, r, 3, tot_rn, F(10, True), C_BAND, RIGHT, numfmt=NUM)
cell(ws, r, 4, rev_m(tot_rev), F(10, True), C_BAND, RIGHT, numfmt=NUM1)
cell(ws, r, 5, adr_k(tot_rev, tot_rn), F(10, True), C_BAND, RIGHT, numfmt=NUM)
cell(ws, r, 6, tot_cnt, F(10, True), C_BAND, RIGHT, numfmt=NUM)
cell(ws, r, 7, "", F(10), C_BAND, RIGHT)
data_end = r - 1
r += 1

# 라인차트: 연도별 RN & REV (이중축)
lc = LineChart(); lc.title = "연도별 객실수(실) · REV(백만)"; lc.style = 2
lc.height = 6.2; lc.width = 12.5
cats = Reference(ws, min_col=2, min_row=data_start, max_row=data_end)
rn_ref = Reference(ws, min_col=3, min_row=hdr_row, max_row=data_end)
lc.add_data(rn_ref, titles_from_data=True); lc.set_categories(cats)
lc2 = LineChart()
rev_ref = Reference(ws, min_col=4, min_row=hdr_row, max_row=data_end)
lc2.add_data(rev_ref, titles_from_data=True); lc2.y_axis.axId = 200
lc2.y_axis.title = "REV(백만)"; lc.y_axis.title = "객실수(실)"
lc2.y_axis.crosses = "max"
# 결합 전 각 차트 series 스타일링
lc.series[0].graphicalProperties.line.solidFill = C_RN
lc.series[0].graphicalProperties.line.width = 26000
lc.series[0].marker.symbol = "circle"; lc.series[0].marker.size = 6; lc.series[0].smooth = False
lc2.series[0].graphicalProperties.line.solidFill = C_REV
lc2.series[0].graphicalProperties.line.width = 26000
lc2.series[0].marker.symbol = "circle"; lc2.series[0].marker.size = 6; lc2.series[0].smooth = False
lc += lc2
lc.x_axis.delete = False; lc.y_axis.delete = False
ws.add_chart(lc, f"I{hdr_row-1}")
r += 1

# ── 국적별 현황 (표 + 데이터바) ──
section(ws, r, 2, 11, "국적별 현황 (객실수 실)"); r += 1
heads = ["국적", "권역", "객실수(실)", "구성비", "REV(백만)", "ADR(천원)", "케이스"]
widths = [1, 1, 1, 1, 1, 1, 1]
for j, h in enumerate(heads):
    merge_cell(ws, r, 2 + j, r, 2 + j, h, F(10, True, "FFFFFF"), C_HEAD, CEN)
r += 1
nat_start = r
for c in countries_sorted:
    d = by_country[c]
    cell(ws, r, 2, c, F(10, True), C_BAND2, LEFT)
    cell(ws, r, 3, REGION.get(c, '기타'), F(9, False, "6A6A6A"), None, CEN)
    cell(ws, r, 4, d['rn'], F(10), None, RIGHT, numfmt=NUM)
    cell(ws, r, 5, round(d['rn'] / tot_rn * 100, 1), F(10), None, RIGHT, numfmt='0.0"%"')
    cell(ws, r, 6, rev_m(d['rev']), F(10), None, RIGHT, numfmt=NUM1)
    cell(ws, r, 7, adr_k(d['rev'], d['rn']), F(10), None, RIGHT, numfmt=NUM)
    cell(ws, r, 8, d['cnt'], F(10), None, RIGHT, numfmt=NUM)
    r += 1
nat_end = r - 1
# 데이터바 (객실수 열 D=4)
ws.conditional_formatting.add(f"D{nat_start}:D{nat_end}",
    DataBarRule(start_type="num", start_value=0, end_type="max",
                color=C_ACC, showValue=True, minLength=None, maxLength=None))
r += 1

# ── 국적 × 월 스택 막대 (Top5 국적) ──
section(ws, r, 2, 11, "월별 · 국적(Top5) 객실수 추이"); r += 1
stk_hdr = r
cell(ws, r, 2, "월", F(10, True, "FFFFFF"), C_HEAD, CEN)
for j, c in enumerate(TOP_NAT):
    cell(ws, r, 3 + j, c, F(10, True, "FFFFFF"), C_HEAD, CEN)
cell(ws, r, 3 + len(TOP_NAT), "기타", F(10, True, "FFFFFF"), C_HEAD, CEN)
r += 1
stk_start = r
for m in range(1, 13):
    cell(ws, r, 2, f"{m}월", F(10, True), C_BAND2, CEN)
    row_total = by_month[m]['rn'] if m in by_month else 0
    top_sum = 0
    for j, c in enumerate(TOP_NAT):
        v = by_country_month.get((c, m), {'rn': 0})['rn']
        top_sum += v
        cell(ws, r, 3 + j, v, F(9), None, RIGHT, numfmt=NUM)
    cell(ws, r, 3 + len(TOP_NAT), max(0, row_total - top_sum), F(9), None, RIGHT, numfmt=NUM)
    r += 1
stk_end = r - 1
bc = BarChart(); bc.type = "col"; bc.grouping = "stacked"; bc.overlap = 100
bc.title = "월별 국적 구성 (객실수 실)"; bc.height = 6.4; bc.width = 13
bc.style = 2
cats = Reference(ws, min_col=2, min_row=stk_start, max_row=stk_end)
data = Reference(ws, min_col=3, max_col=3 + len(TOP_NAT), min_row=stk_hdr, max_row=stk_end)
bc.add_data(data, titles_from_data=True); bc.set_categories(cats)
for i, s in enumerate(bc.series):
    s.graphicalProperties.solidFill = NAT_COLORS[i % len(NAT_COLORS)]
    s.graphicalProperties.line.solidFill = "FFFFFF"
bc.y_axis.delete = False; bc.x_axis.delete = False
ws.add_chart(bc, f"I{stk_hdr-1}")
r += 1

# ── 여행사 Top15 (표 + 데이터바) ──
section(ws, r, 2, 11, "여행사별 Top 15 (객실수 실)"); r += 1
heads = ["순위", "여행사", "대표국적", "객실수(실)", "REV(백만)", "ADR(천원)", "케이스"]
for j, h in enumerate(heads):
    merge_cell(ws, r, 2 + j, r, 2 + j, h, F(10, True, "FFFFFF"), C_HEAD, CEN)
r += 1
ag_start = r
for i, a in enumerate(agencies_sorted[:15], 1):
    d = by_agency[a]
    cell(ws, r, 2, i, F(10, True), C_BAND2, CEN)
    cell(ws, r, 3, a, F(10), None, LEFT)
    cell(ws, r, 4, ag_country.get(a, ''), F(9, False, "6A6A6A"), None, CEN)
    cell(ws, r, 5, d['rn'], F(10), None, RIGHT, numfmt=NUM)
    cell(ws, r, 6, rev_m(d['rev']), F(10), None, RIGHT, numfmt=NUM1)
    cell(ws, r, 7, adr_k(d['rev'], d['rn']), F(10), None, RIGHT, numfmt=NUM)
    cell(ws, r, 8, d['cnt'], F(10), None, RIGHT, numfmt=NUM)
    r += 1
ag_end = r - 1
ws.conditional_formatting.add(f"E{ag_start}:E{ag_end}",
    DataBarRule(start_type="num", start_value=0, end_type="max", color=C_ACC, showValue=True))
ws.column_dimensions["D"].width = 22  # 여행사명 넓게 (열 D=4? 아래 재조정)

# 대시보드 열폭 미세조정
ws.column_dimensions["B"].width = 8
ws.column_dimensions["C"].width = 12
for c in range(4, 9):
    ws.column_dimensions[get_column_letter(c)].width = 11
ws.column_dimensions["D"].width = 13

# ---------------- 공통: 표 시트 빌더 ----------------
def write_table(ws, headers, data_rows, widths=None, title=None, numfmt_cols=None):
    numfmt_cols = numfmt_cols or {}
    r0 = 1
    if title:
        cell(ws, 1, 1, title, F(13, True, C_TITLE), None, LEFT, border=False)
        r0 = 2
    for j, h in enumerate(headers, 1):
        cell(ws, r0, j, h, F(10, True, "FFFFFF"), C_HEAD, CEN)
    rr = r0 + 1
    for row in data_rows:
        for j, v in enumerate(row, 1):
            al = RIGHT if isinstance(v, (int, float)) and not isinstance(v, bool) else LEFT
            nf = numfmt_cols.get(j)
            cell(ws, rr, j, v, F(10), (C_BAND2 if j == 1 else None), al, numfmt=nf)
        rr += 1
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=r0 + 1, column=2)
    return rr

# ---------------- 시트2: 여행사별 (여행사 × 연도 + 합) ----------------
ws2 = wb.create_sheet("여행사별")
ws2.sheet_view.showGridLines = False
heads = ["순위", "여행사", "대표국적"]
for y in YEARS:
    heads += [f"{y} RN"]
heads += ["RN 합", "REV(백만)", "ADR(천원)", "케이스"]
data_rows = []
for i, a in enumerate(agencies_sorted, 1):
    d = by_agency[a]
    row = [i, a, ag_country.get(a, '')]
    for y in YEARS:
        row.append(by_agency_year.get((a, y), {'rn': 0})['rn'])
    row += [d['rn'], rev_m(d['rev']), adr_k(d['rev'], d['rn']), d['cnt']]
    data_rows.append(row)
nfc = {j: NUM for j in range(4, 4 + len(YEARS) + 1)}
nfc[4 + len(YEARS) + 1] = NUM1
nfc[4 + len(YEARS) + 2] = NUM
nfc[4 + len(YEARS) + 3] = NUM
write_table(ws2, heads, data_rows,
            widths=[6, 26, 10] + [9] * len(YEARS) + [10, 11, 11, 9],
            title="여행사별 실적 (5개년, 객실수 실 기준 정렬)", numfmt_cols=nfc)

# ---------------- 시트3: 국적별 (국적 × 연도 + 월) ----------------
ws3 = wb.create_sheet("국적별")
ws3.sheet_view.showGridLines = False
heads = ["국적", "권역"] + [f"{y} RN" for y in YEARS] + ["RN 합", "REV(백만)", "ADR(천원)", "케이스"]
data_rows = []
for c in countries_sorted:
    d = by_country[c]
    row = [c, REGION.get(c, '기타')]
    for y in YEARS:
        row.append(by_country_year.get((c, y), {'rn': 0})['rn'])
    row += [d['rn'], rev_m(d['rev']), adr_k(d['rev'], d['rn']), d['cnt']]
    data_rows.append(row)
nfc = {j: NUM for j in range(3, 3 + len(YEARS) + 1)}
nfc[3 + len(YEARS) + 1] = NUM1
nfc[3 + len(YEARS) + 2] = NUM
nfc[3 + len(YEARS) + 3] = NUM
end = write_table(ws3, heads, data_rows,
            widths=[14, 9] + [9] * len(YEARS) + [10, 11, 11, 9],
            title="국적별 실적 (5개년)", numfmt_cols=nfc)

# ---------------- 시트4: 연월별 (연도 × 월 매트릭스 RN, REV) ----------------
ws4 = wb.create_sheet("연월별")
ws4.sheet_view.showGridLines = False
cell(ws4, 1, 1, "연월별 실적 매트릭스 (상단 객실수 실 / 하단 REV 백만)", F(13, True, C_TITLE), None, LEFT, border=False)
# RN 매트릭스
r = 3
cell(ws4, r, 1, "객실수(실)", F(10, True, "FFFFFF"), C_TITLE, CEN)
for m in range(1, 13):
    cell(ws4, r, 1 + m, f"{m}월", F(10, True, "FFFFFF"), C_HEAD, CEN)
cell(ws4, r, 14, "연합계", F(10, True, "FFFFFF"), C_HEAD, CEN)
r += 1
for y in YEARS:
    cell(ws4, r, 1, y, F(10, True), C_BAND, CEN)
    ytot = 0
    for m in range(1, 13):
        v = by_ym.get((y, m), {'rn': 0})['rn']; ytot += v
        cell(ws4, r, 1 + m, v if v else "", F(9), None, RIGHT, numfmt=NUM)
    cell(ws4, r, 14, ytot, F(10, True), C_BAND, RIGHT, numfmt=NUM)
    r += 1
# 월합계
cell(ws4, r, 1, "월합계", F(10, True), C_BAND, CEN)
for m in range(1, 13):
    cell(ws4, r, 1 + m, by_month.get(m, {'rn': 0})['rn'], F(10, True), C_BAND, RIGHT, numfmt=NUM)
cell(ws4, r, 14, tot_rn, F(10, True), C_ACC, RIGHT, numfmt=NUM)
r += 2
# REV 매트릭스
cell(ws4, r, 1, "REV(백만)", F(10, True, "FFFFFF"), C_TITLE, CEN)
for m in range(1, 13):
    cell(ws4, r, 1 + m, f"{m}월", F(10, True, "FFFFFF"), C_HEAD, CEN)
cell(ws4, r, 14, "연합계", F(10, True, "FFFFFF"), C_HEAD, CEN)
r += 1
for y in YEARS:
    cell(ws4, r, 1, y, F(10, True), C_BAND, CEN)
    ytot = 0
    for m in range(1, 13):
        v = by_ym.get((y, m), {'rev': 0})['rev']; ytot += v
        cell(ws4, r, 1 + m, (rev_m(v) if v else ""), F(9), None, RIGHT, numfmt=NUM1)
    cell(ws4, r, 14, rev_m(ytot), F(10, True), C_BAND, RIGHT, numfmt=NUM1)
    r += 1
ws4.column_dimensions["A"].width = 11
for m in range(1, 14):
    ws4.column_dimensions[get_column_letter(1 + m)].width = 8.5

# ---------------- 시트5: DB ----------------
ws5 = wb.create_sheet("DB")
ws5.sheet_view.showGridLines = False
db_heads = ["연도", "판매일자", "세그먼트", "여행사(정리)", "회원명(원본)", "국적", "권역", "단체구분",
            "사업장", "객실수(실)RN", "1박객실료", "REV공급가(원)", "REV(백만)", "ADR(천원)", "KEY_RSV_NO"]
for j, h in enumerate(db_heads, 1):
    cell(ws5, 1, j, h, F(10, True, "FFFFFF"), C_HEAD, CEN)
rr = 2
for r_ in sorted(rows, key=lambda x: (x['date'], -x['rn'])):
    vals = [r_['year'], r_['date'], "Inbound", r_['agency'], r_['member'], r_['country'], r_['region'],
            r_['danche'], r_['prop'], r_['rn'], r_['rate'], r_['rev'], round(r_['rev']/1e6, 3),
            adr_k(r_['rev'], r_['rn']), r_['key']]
    for j, v in enumerate(vals, 1):
        al = RIGHT if j in (10, 11, 12, 13, 14) else LEFT
        cell(ws5, rr, j, v, F(9), None, al)
    rr += 1
ws5.freeze_panes = "A2"
for j, w in enumerate([7, 11, 9, 22, 30, 10, 9, 14, 16, 12, 11, 14, 10, 10, 14], 1):
    ws5.column_dimensions[get_column_letter(j)].width = w
ws5.auto_filter.ref = f"A1:{get_column_letter(len(db_heads))}1"

# ---------------- 시트6: 기준 ----------------
ws6 = wb.create_sheet("기준")
ws6.sheet_view.showGridLines = False
notes = [
    ["항목", "내용"],
    ["대상", "인바운드 단체(1박 20객실 이상) — 43번 예약자료 5개년"],
    ["소스", "data/raw_db/20*/43.[57,87]온라인영업팀 예약자료 (Inbound=변경예약집계코드 58). 43만 사용, 취소(44) 미차감"],
    ["케이스 정의", "세그먼트=Inbound AND 행(예약×투숙일) 객실수 ≥ 20. (KEY_RSV_NO×판매일자 그룹과 1:1 일치 확인)"],
    ["객실수(실) RN", "객실수 합(1박=1행 단위). '실'=예약 객실수(43 기준). 취소 미차감 그로스"],
    ["REV(백만)", "공급가 ÷ 1e6. 공급가 = 1박객실료 × 객실수 ÷ 1.1 (VAT 제외, 프로젝트 관례)"],
    ["ADR(천원)", "공급가 합 ÷ RN ÷ 1000"],
    ["연도/월", "판매일자(투숙일) 기준. 2026은 부분(연초~추출시점)"],
    ["국적", "회원명 괄호 내 국가 추출(대만·중국·일본·말레이시아·태국·베트남·홍콩·필리핀·인도네시아·미주·유럽·호주 등). 없으면 단체구분(동남아/일본)→'동남아(미상)', 그 외 '미상'"],
    ["여행사", "회원명에서 괄호/팀·행사 접미사(_,' - ')·공백뒤 국가어 제거한 기본 상호. 팀 분리건은 동일 여행사로 병합"],
    ["권역", "중화권(대만·중국·홍콩)/동남아/일본/구미주(미주·유럽·호주)/기타아시아/기타"],
    ["", ""],
    ["검증: 총 RN", tot_rn],
    ["검증: 총 REV공급가(원)", tot_rev],
    ["검증: 총 REV(백만)", rev_m(tot_rev)],
    ["검증: 케이스 수", tot_cnt],
    ["양식", "첨부 고객지표 대시보드 시각문법(헤더바·회색라인표·데이터바·라인/스택바) 준용, 폰트=대명체, 색상 근접 팔레트"],
]
for i, row in enumerate(notes, 1):
    for j, v in enumerate(row, 1):
        fill = C_HEAD if i == 1 else None
        fnt = F(10, True, "FFFFFF") if i == 1 else F(10)
        cell(ws5 := ws6, i, j, v, fnt, fill, (LEFT if j == 2 else CEN))
    ws6.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
ws6.column_dimensions["A"].width = 20
ws6.column_dimensions["B"].width = 100

# 기본시트 제거
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"저장: {OUT}")

# ── 교차검증 ──
db_rn = sum(r['rn'] for r in rows)
db_rev = sum(r['rev'] for r in rows)
print("\n=== 교차검증 ===")
print(f"RN   대시보드/DB = {tot_rn:,} / {db_rn:,}  일치={tot_rn==db_rn}")
print(f"REV  대시보드/DB = {tot_rev:,} / {db_rev:,}  일치={tot_rev==db_rev}")
print("\n연도별:", {y: by_year[y]['rn'] for y in YEARS})
print("국적 Top:", [(c, by_country[c]['rn']) for c in countries_sorted[:8]])
