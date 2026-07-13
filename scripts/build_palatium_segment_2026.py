#!/usr/bin/env python3
"""
긴급 추출: 팔라티움 해운대 by sonofelice 2026년 2~6월 세그먼트별 실적(RN·매출) → xlsx
================================================================================
소스: data/palatium_db/예약정보조회2026010701~07.xlsx  (Sales Date 2026-07-07 = 과거 실적 배치)
  · 이 배치가 도착월 1~6월(대부분 Checked Out)을 담은 유일한 Feb~June 소스.
  · 071301~05 배치(Sales Date 7/13)는 7~12월 forward 온북이라 2~5월 미포함 → 미사용.
세그먼트: 팔라티움 거래처(15)/시장(24) 기준 분류. 거래처로 OTA/G-OTA 판별, 그 외는 시장으로 버킷.
RN = 박수 × 객실수 (도착월 귀속, 파서 parse_palatium_db 관례 동일).
매출 = 객실료(10) 총액(부킹 전체, 도착월 귀속). 공급가(÷1.1) 병기.
상태: 실적=active(Checked Out/In House/Reservation/Assigned Room/No Show), 취소(Cancelled) 제외.

⚠️ 소스 불완전: 추출본이 낡은/잘린 스냅샷이라 2~6월 실적이 실제와 다를 수 있음(특히 2월 과대 가능).
   정확값은 다올비전 완전본 재추출 필요. 상세는 '기준' 시트 상단.
"""
import glob, unicodedata
from pathlib import Path
from collections import defaultdict
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT = Path(__file__).resolve().parent.parent
PDB = PROJECT / "data" / "palatium_db"
OUT = Path.home() / "Desktop" / "팔라티움_해운대_2026_2-6월_세그먼트별_실적.xlsx"

ACTIVE = {"Reservation", "In House", "Checked Out", "Assigned Room", "No Show"}
CANCEL = "Cancelled Reservation"
MONTHS = [2, 3, 4, 5, 6]

# ─── 세그먼트 분류 (거래처 우선 → 시장 폴백) ───
GOTA_KW = ["아고다", "트립닷컴", "익스피디아", "부킹닷컴", "트립비토즈", "호텔패스", "호텔스닷컴", "expedia", "agoda", "booking"]
OTA_KW = ["여기어때", "야놀자", "놀유니버스", "웹투어", "인터파크", "네이버", "쿠팡", "타이드스퀘어"]

def classify_segment(vendor, market, path, rate):
    v = (vendor or "").strip()
    mk = (market or "").strip()
    for kw in GOTA_KW:
        if kw in v:
            return "G-OTA(해외OTA)"
    for kw in OTA_KW:
        if kw in v:
            return "OTA(국내OTA)"
    # 거래처가 자사직판/기타 → 시장으로 버킷
    u = mk.upper()
    if ("소노회원" in mk) or ("D멤버스" in mk) or ("MEMBERSHIP" in u) or ("멤버스" in mk):
        return "회원/D멤버스(자사)"
    if "국내여행사" in mk:
        return "국내여행사"
    if ("MICE" in u) or ("마이스" in mk) or ("GROUP" in u) or ("단체" in mk):
        return "MICE/단체"
    if ("WALK" in u) or ("워크" in mk):
        return "WALK-IN"
    if ("COMPLIMENT" in u) or ("COMP" in u) or ("HOUSE USE" in u) or ("EMPLOYEE" in u) or ("임직원" in mk):
        return "COMP/하우스/임직원"
    if "FOREIGN" in u or "INDIVI" in u:
        return "FIT(개인)"
    if "해외여행사" in mk:
        return "G-OTA(해외OTA)"  # 거래처 미매칭이나 해외여행사 시장 → G-OTA 귀속
    return "기타/미분류"

# 세그먼트 표시 순서
SEG_ORDER = ["G-OTA(해외OTA)", "OTA(국내OTA)", "회원/D멤버스(자사)", "국내여행사",
             "MICE/단체", "WALK-IN", "COMP/하우스/임직원", "FIT(개인)", "기타/미분류"]
def seg_rank(s):
    return SEG_ORDER.index(s) if s in SEG_ORDER else len(SEG_ORDER)

def am(v):
    if isinstance(v, (datetime, date)):
        return v.month, v.year
    try:
        d = datetime.strptime(str(v)[:10], "%Y-%m-%d")
        return d.month, d.year
    except (ValueError, TypeError):
        return None, None

# ─── 파싱 ───
files = sorted(f for f in glob.glob(str(PDB / "*.xlsx"))
               if unicodedata.normalize("NFC", Path(f).name).startswith("예약정보조회2026010"))
HDR = None
agg = defaultdict(lambda: {"rn": 0, "rev": 0})   # (month, seg) -> rn, rev(원 총액)
seg_map = defaultdict(lambda: {"rn": 0, "rev": 0})  # (seg, vendor, market) -> for 기준
db_rows = []
tot_rn = tot_rev = 0
seen = set()

for fp in files:
    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if i == 3:
            HDR = [str(c).strip() if c is not None else "" for c in row]
        if i < 4:
            continue
        st = row[0] if row[0] else ""
        if not st or st == "총합계" or "Created by" in str(st):
            continue
        if st == CANCEL or st not in ACTIVE:
            continue
        m, y = am(row[6])
        if y != 2026 or m not in MONTHS:
            continue
        resno = row[2]
        dup = f"{resno}_{am(row[6])}_{st}_{row[9]}_{row[6]}"
        if dup in seen:
            continue
        seen.add(dup)
        nights = max(1, int(row[8] or 1))
        rooms = max(1, int(row[14] or 1))
        rn = nights * rooms
        rev = float(row[10] or 0)            # 객실료 총액(부킹 전체)
        vendor = str(row[15] or "").strip()
        market = str(row[24] or "").strip()
        path = str(row[25] or "").strip()
        rate = str(row[22] or "").strip()
        seg = classify_segment(vendor, market, path, rate)
        agg[(m, seg)]["rn"] += rn
        agg[(m, seg)]["rev"] += rev
        seg_map[(seg, vendor, market)]["rn"] += rn
        seg_map[(seg, vendor, market)]["rev"] += rev
        tot_rn += rn
        tot_rev += rev
        db_rows.append(list(row) + [seg, rn, rev, round(rev / 1.1 / 1_000_000, 4)])
    wb.close()

print(f"소스 파일: {len(files)}개  |  필터 통과 행: {len(db_rows):,}")
print(f"합계 RN: {tot_rn:,}  매출총액: {tot_rev:,.0f}원 = {tot_rev/1e6:,.1f}백만 (공급가 {tot_rev/1.1/1e6:,.1f}백만)")

# ─── 엑셀 ───
wb = openpyxl.Workbook()
HF = PatternFill("solid", fgColor="7030A0"); HFONT = Font(bold=True, color="FFFFFF")
SUBF = PatternFill("solid", fgColor="E4DFEC"); MONF = PatternFill("solid", fgColor="FCE4D6")
WARNF = PatternFill("solid", fgColor="FFF2CC"); BOLD = Font(bold=True)
THIN = Side(style="thin", color="C9C9C9"); BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CEN = Alignment(horizontal="center", vertical="center"); RIGHT = Alignment(horizontal="right"); WRAP = Alignment(wrap_text=True, vertical="top")

CAVEAT = ("⚠️ 대외비·관리자용 | 소스 불완전 주의: 팔라티움 예약 추출본이 낡은/잘린 스냅샷 상태라 "
          "2~6월 실적(특히 2월 과대 가능 ~+13%)이 실제와 다를 수 있음. 정확값은 다올비전 완전본 재추출 필요. "
          "본 자료는 현재 파일 실측값 기준(추정·더미 없음).")

def hdr_style(ws, ncol, row=1):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c); cell.fill = HF; cell.font = HFONT; cell.alignment = CEN; cell.border = BORD

# ── 시트1: 집계 ──
ws = wb.active; ws.title = "집계"
ws.append([CAVEAT]); ws.merge_cells("A1:F1")
ws.cell(row=1, column=1).fill = WARNF; ws.cell(row=1, column=1).font = Font(bold=True, color="9C0006")
ws.cell(row=1, column=1).alignment = WRAP
ws.row_dimensions[1].height = 42
ws.append([])
cols = ["투숙월", "세그먼트", "RN(객실수)", "매출_총액(백만)", "매출_공급가(÷1.1,백만)", "구성비(RN%)"]
ws.append(cols); hdr_style(ws, len(cols), row=3)
r = 4
grand_rn = grand_rev = 0
for m in MONTHS:
    segs = sorted([(seg, v) for (mm, seg), v in agg.items() if mm == m], key=lambda x: seg_rank(x[0]))
    m_rn = sum(v["rn"] for _, v in segs); m_rev = sum(v["rev"] for _, v in segs)
    for seg, v in segs:
        pct = round(v["rn"] / m_rn * 100, 1) if m_rn else 0
        ws.append([f"{m}월", seg, v["rn"], round(v["rev"]/1e6, 1), round(v["rev"]/1.1/1e6, 1), pct])
        for c in range(1, len(cols)+1):
            ws.cell(row=r, column=c).border = BORD
            if c >= 3: ws.cell(row=r, column=c).alignment = RIGHT
        r += 1
    # 월 소계
    ws.append([f"▶ {m}월 합계", "", m_rn, round(m_rev/1e6, 1), round(m_rev/1.1/1e6, 1), 100.0])
    for c in range(1, len(cols)+1):
        ws.cell(row=r, column=c).fill = MONF; ws.cell(row=r, column=c).font = BOLD; ws.cell(row=r, column=c).border = BORD
        if c >= 3: ws.cell(row=r, column=c).alignment = RIGHT
    r += 1
    grand_rn += m_rn; grand_rev += m_rev
# 총합계
ws.append(["■ 총합계 (2~6월)", "", grand_rn, round(grand_rev/1e6, 1), round(grand_rev/1.1/1e6, 1), 100.0])
for c in range(1, len(cols)+1):
    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFC000"); ws.cell(row=r, column=c).font = Font(bold=True, size=12); ws.cell(row=r, column=c).border = BORD
    if c >= 3: ws.cell(row=r, column=c).alignment = RIGHT
r += 2
# 세그먼트×월 매트릭스 (RN)
ws.cell(row=r, column=1, value="[참고] 세그먼트 × 월 RN 매트릭스").font = BOLD; r += 1
hdr2 = ["세그먼트"] + [f"{m}월" for m in MONTHS] + ["2~6월合"]
ws.append(hdr2)
for c in range(1, len(hdr2)+1):
    ws.cell(row=r, column=c).fill = HF; ws.cell(row=r, column=c).font = HFONT; ws.cell(row=r, column=c).alignment = CEN; ws.cell(row=r, column=c).border = BORD
r += 1
segs_all = sorted({seg for (_, seg) in agg}, key=seg_rank)
for seg in segs_all:
    rowvals = [seg] + [agg.get((m, seg), {"rn":0})["rn"] for m in MONTHS]
    rowvals.append(sum(rowvals[1:]))
    ws.append(rowvals)
    for c in range(1, len(hdr2)+1):
        ws.cell(row=r, column=c).border = BORD
        if c >= 2: ws.cell(row=r, column=c).alignment = RIGHT
    r += 1
for i, w in enumerate([18, 22, 12, 16, 20, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A4"

# ── 시트2: DB ──
ws2 = wb.create_sheet("DB")
db_hdr = (HDR or [f"col{i}" for i in range(38)]) + ["세그먼트", "_RN", "_매출총액(원)", "_매출공급가(백만)"]
ws2.append(db_hdr); hdr_style(ws2, len(db_hdr))
for row in db_rows:
    ws2.append(row)
ws2.freeze_panes = "B2"
seg_col = len(HDR) + 1
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions[get_column_letter(16)].width = 24  # 거래처
ws2.column_dimensions[get_column_letter(seg_col)].width = 20
ws2.auto_filter.ref = f"A1:{get_column_letter(len(db_hdr))}1"

# ── 시트3: 기준 (caveat + 세그먼트 매핑) ──
ws3 = wb.create_sheet("기준")
ws3.append([CAVEAT]); ws3.merge_cells("A1:E1")
ws3.cell(row=1, column=1).fill = WARNF; ws3.cell(row=1, column=1).font = Font(bold=True, color="9C0006"); ws3.cell(row=1, column=1).alignment = WRAP
ws3.row_dimensions[1].height = 56
notes = [
    ["항목", "내용"],
    ["대상", "팔라티움 해운대 by sonofelice — 2026년 2~6월 (도착월 기준)"],
    ["소스", "data/palatium_db/예약정보조회2026010701~07.xlsx (Sales Date 2026-07-07, 과거 실적 배치)"],
    ["소스 주의①", "파일명 '2026010701'은 1월처럼 보이나 내부 Sales Date=2026-07-07. 071301~05 배치는 Sales Date 7-13이며 7~12월 forward 온북 → 2~5월 미포함이라 미사용."],
    ["소스 주의②", "추출본이 낡은/잘린 스냅샷 상태 → 2~6월 실적이 실제와 다를 수 있음(특히 2월 과대 가능). 정확값=다올비전 완전본 재추출 필요."],
    ["세그먼트 정의", "거래처(원본 col15) 우선 판별 → OTA/G-OTA. 거래처가 자사직판/기타면 시장(col24)으로 버킷."],
    ["G-OTA(해외OTA)", "거래처: 아고다·트립닷컴·익스피디아·부킹닷컴·트립비토즈·호텔패스 (시장=해외여행사, 경로=Channel)"],
    ["OTA(국내OTA)", "거래처: 여기어때·야놀자(놀유니버스)·웹투어·타이드스퀘어 등"],
    ["회원/D멤버스(자사)", "거래처=팔라티움 직판 & 시장=소노회원(분양)/D멤버스(온라인)/PALATIUM MEMBERSHIP"],
    ["국내여행사", "시장=국내여행사"],
    ["MICE/단체", "시장=MICE(마이스)/Group"],
    ["WALK-IN", "시장/경로=WALK-IN"],
    ["COMP/하우스/임직원", "시장=COMPLIMENTARY/HOUSE USE/PALATIUM EMPLOYEE/소노임직원"],
    ["FIT(개인)", "시장=Foreign Individual Traveler"],
    ["RN", "박수 × 객실수, 도착월 귀속 (parse_palatium_db 관례 동일)"],
    ["매출_총액", "객실료(col11 객실료=부킹 전체 총액), 도착월 귀속. 원 단위 → 백만 환산"],
    ["매출_공급가", "매출_총액 ÷ 1.1 (객실료 VAT 포함 가정). 팔라티움 파서는 총액(VAT미조정) 사용이 기본이라 총액을 헤드라인, 공급가는 병기."],
    ["상태(실적)", "active(Checked Out·In House·Reservation·Assigned Room·No Show) 포함, Cancelled 제외"],
    ["중복제거", "예약번호+도착일+상태+객실타입 키 기준"],
    ["", ""],
    ["검증: 집계 RN 합", tot_rn],
    ["검증: 집계 매출총액(원)", round(tot_rev)],
    ["검증: 집계 매출총액(백만)", round(tot_rev/1e6, 1)],
    ["검증: DB 행수", len(db_rows)],
]
r3 = 2
for row in notes:
    ws3.append(row)
    if r3 == 2:
        for c in (1, 2):
            ws3.cell(row=r3, column=c).fill = HF; ws3.cell(row=r3, column=c).font = HFONT; ws3.cell(row=r3, column=c).alignment = CEN
    ws3.cell(row=r3, column=2).alignment = WRAP
    r3 += 1
ws3.column_dimensions["A"].width = 22; ws3.column_dimensions["B"].width = 95

# ── 시트4: 세그먼트 구성(거래처/시장 → 세그먼트, 2~6월 합) ──
ws4 = wb.create_sheet("세그먼트구성")
ws4.append(["세그먼트", "거래처", "시장", "RN", "매출총액(백만)", "매출공급가(백만)"])
hdr_style(ws4, 6)
rr = 2
for (seg, vendor, market), v in sorted(seg_map.items(), key=lambda kv: (seg_rank(kv[0][0]), -kv[1]["rn"])):
    ws4.append([seg, vendor, market, v["rn"], round(v["rev"]/1e6, 2), round(v["rev"]/1.1/1e6, 2)])
    for c in range(1, 7):
        ws4.cell(row=rr, column=c).border = BORD
        if c >= 4: ws4.cell(row=rr, column=c).alignment = RIGHT
    rr += 1
for i, w in enumerate([20, 26, 22, 10, 15, 18], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = "A2"

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"\n저장: {OUT}")

# ─── 교차검증 ───
db_rn = sum(row[-3] for row in db_rows)
db_rev = sum(row[-2] for row in db_rows)
print("\n=== 교차검증 (집계 vs DB) ===")
print(f"RN     집계={tot_rn:,}  DB={db_rn:,}  일치={tot_rn==db_rn}")
print(f"매출총액 집계={tot_rev:,.2f}  DB={db_rev:,.2f}  일치={abs(tot_rev-db_rev)<1}")
print("\n=== 월×세그먼트 RN ===")
for m in MONTHS:
    line = "  ".join(f"{seg.split('(')[0]}={agg.get((m,seg),{'rn':0})['rn']}" for seg in segs_all if agg.get((m,seg),{'rn':0})['rn'])
    print(f"  {m}월: {line}")
