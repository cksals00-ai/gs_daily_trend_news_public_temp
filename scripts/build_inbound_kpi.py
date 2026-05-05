"""
build_inbound_kpi.py — 인바운드 도전목표 KPI 엑셀 → JSON 변환

Source:  data/Inbound/(인바운드)2026_도전목표KPI.xlsx
Output:  docs/data/inbound_kpi.json

산출 구조 (사업장×국가×여행사×월 풀그리드, Budget+도전목표 둘 다)

  totals     : 회사 전체 RNs/ADR/REV (Budget · 도전목표 · 25년 실적)
  by_year    : (단일 키 '2026' — 다년 확장 대비) 위 totals 를 연도별로 박싱
  properties : 사업장별 KPI (Budget · 도전 · 25Y actual · 27/28 stretch · 핵심 전략 텍스트)
  grid       : [{property, country, agency, type, budget:{total,monthly}, stretch:{total,monthly}}]
  meta       : 시트별 행수/컬럼 수, source path, generated_at

소스 시트 매핑:
- 도전목표(3개년)        → totals_3yr   (회사 단위 RNs/ADR/매출 7개 KPI 추이)
- 총괄(사업장별 전략)     → properties.<>.strategy_text + 25/26/27/28 stretch
- 총괄(사업장별 실적)     → properties.<>.monthly (사업장×월 budget/도전/전년)
- 월별 KPI(도전목표)      → grid (사업장×국가×여행사×월, stretch)
- 월별 KPI              → grid (사업장×국가×여행사×월, budget)  ← stretch grid 와 키로 머지

규칙:
- '기타' 카테고리는 묶음용이 아니라 패키지/FIT 그룹의 정식 라벨 → 그대로 노출
- 0/NULL 행도 보존 (진도/남해 0 목표도 의미 있음)
- 화폐 단위: RNs=실, ADR=천원/박, REV=백만원 (엑셀 시트 단위 주석 그대로)
"""

from __future__ import annotations
import json
import warnings
from datetime import datetime
from pathlib import Path

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
SRC  = ROOT / "data" / "Inbound" / "(인바운드)2026_도전목표KPI.xlsx"
OUT  = ROOT / "docs" / "data" / "inbound_kpi.json"

PROPERTIES_ORDER = [
    "고양", "해운대", "비발디", "단양", "여수", "델피노", "삼척", "벨제주",
    "양평", "캄제주", "천안", "변산", "거제", "청송", "양양", "경주", "진도", "남해",
]


def _to_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "")
    if not s or s in {"-", "#DIV/0!", "N/A"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _round(v, ndigits=1):
    if v is None:
        return None
    f = float(v)
    if abs(f - round(f)) < 1e-9:
        return int(round(f))
    return round(f, ndigits)


# ── 월별 KPI 시트 파서 (도전목표/budget 동일 스키마) ────────────────────
# iter_rows 튜플 인덱스 (col A 가 비어있어 idx0=None):
#   1=순번  2=사업장  3=RNs  4=ADR  5=REV  6=국가  7=여행사  8=전략구분
#   9=계  10~21=1월~12월
def parse_monthly_kpi(ws):
    rows = list(ws.iter_rows(values_only=True))
    out = {}  # (property, country, agency) → {type, total, monthly[12]}
    cur_property = None
    cur_country = None
    for r in rows[5:]:  # 데이터 행은 R6(TOTAL) 부터 (R5는 month header)
        if r is None or len(r) < 10:
            continue
        # property switch: idx1 이 숫자 또는 'TOTAL'
        if r[1] is not None and str(r[1]).strip() not in {"", "TOTAL"}:
            # idx2 가 사업장명
            if r[2]:
                cur_property = str(r[2]).strip()
                cur_country = None
        # country switch: idx6 에 새 국가 (단, 빈 문자열 제외)
        if r[6]:
            cur_country = str(r[6]).strip()
        # agency row: idx7 이 '계' 가 아닌 거래처명
        agency = (str(r[7]).strip() if r[7] else "")
        if not agency or agency == "계":
            continue
        if cur_property is None or cur_country is None:
            continue
        agency_type = (str(r[8]).strip() if r[8] else "일반")
        total = _to_num(r[9])
        # 월별 1월~12월 = 인덱스 10~21
        monthly = []
        for ci in range(10, 22):
            v = r[ci] if ci < len(r) else None
            monthly.append(_round(_to_num(v) or 0, 1))
        total = _round(total or 0, 1)
        key = (cur_property, cur_country, agency)
        out[key] = {
            "type": agency_type,
            "total": total,
            "monthly": monthly,
        }
    return out


# ── 총괄(사업장별 실적) 시트 파서 ─────────────────────────────────────
# iter_rows 인덱스 (idx0=None):  1=구분  2=사업장
#   Total 블록 5cols  : 3=Budget  4=도전  5=목표比  6=전년  7=전년比
#   1월 블록 5cols    : 8=Budget  9=도전 10=목표比 11=전년 12=전년比
#   1월 중복 블록     : 13~17  ← 헤더 오류로 1월이 두번 등재됨, 무시
#   2월 블록          : 18~22 ; 이후 5cols 씩 ... 12월 = 68~72
def parse_property_actual(ws):
    rows = list(ws.iter_rows(values_only=True))
    by_prop = {}
    # 데이터 시작 = R6 (R4·R5 헤더). idx 5 부터.
    for r in rows[5:]:
        if r is None or len(r) < 8:
            continue
        if not r[2]:
            continue
        prop = str(r[2]).strip()
        if prop == "" or prop == "사업장":
            continue
        budget_total  = _to_num(r[3])
        stretch_total = _to_num(r[4])
        ly_total      = _to_num(r[6])
        # 월별: 1월=8, 2월=18, 3월=23, ..., 12월=68. 즉 starts = [8, 18, 23, 28, ..., 68]
        starts = [8] + [18 + 5 * i for i in range(11)]
        monthly = []
        for col in starts:
            if col + 4 < len(r):
                monthly.append({
                    "budget":  _round(_to_num(r[col])     or 0, 1),
                    "stretch": _round(_to_num(r[col + 1]) or 0, 1),
                    "ly":      _round(_to_num(r[col + 3]) or 0, 1),
                })
            else:
                monthly.append({"budget": 0, "stretch": 0, "ly": 0})
        if prop == "Total":
            by_prop["__TOTAL__"] = {
                "budget_rns_total":  _round(budget_total, 1),
                "stretch_rns_total": _round(stretch_total, 1),
                "ly_rns_total":      _round(ly_total, 1),
                "monthly_rns":       monthly,
            }
            continue
        by_prop[prop] = {
            "budget_rns_total":  _round(budget_total, 1),
            "stretch_rns_total": _round(stretch_total, 1),
            "ly_rns_total":      _round(ly_total, 1),
            "monthly_rns":       monthly,
        }
    return by_prop


# ── 총괄(사업장별 전략) 시트 파서 ──────────────────────────────────────
# iter_rows 인덱스 (idx0=None):
#   1=순번  2=사업장
#   3~5  = 2019 Actual (RNs, ADR, REV)
#   6~8  = 2025 Actual
#   9~11 = 25Y vs 19Y
#   12~14= 2026 Budget (RNs, ADR, REV)
#   15~17= 26Y Stretch
#   18~21= 25년 비, 보조컬럼
#   22~25= 달성 KEY 전략 텍스트 (병합 셀)
#   26~28= 27Y Stretch (RNs, ADR, REV)
#   29   = 28Y Stretch RNs
def parse_property_strategy(ws):
    rows = list(ws.iter_rows(values_only=True))
    by_prop = {}
    # 데이터 시작 = R15 (R12 메인헤더 → R13 서브헤더 → R14 = 계 = idx 13)
    # R14 가 '계' total 행이므로 일단 포함해서 처리하다 prop=='계' 면 skip
    for r in rows[13:]:
        if r is None or len(r) < 18:
            continue
        if not r[2]:
            continue
        prop = str(r[2]).strip()
        if prop in {"", "사업장", "계"}:
            continue
        y19  = {"rns": _round(_to_num(r[3]),  1),
                "adr": _round(_to_num(r[4]),  1),
                "rev": _round(_to_num(r[5]),  1)}
        ly25 = {"rns": _round(_to_num(r[6]),  1),
                "adr": _round(_to_num(r[7]),  1),
                "rev": _round(_to_num(r[8]),  1)}
        b26  = {"rns": _round(_to_num(r[12]), 1),
                "adr": _round(_to_num(r[13]), 1),
                "rev": _round(_to_num(r[14]), 1)}
        s26  = {"rns": _round(_to_num(r[15]), 1),
                "adr": _round(_to_num(r[16]), 1),
                "rev": _round(_to_num(r[17]), 1)}
        s27  = {"rns": _round(_to_num(r[26]) if len(r) > 26 else None, 1),
                "adr": _round(_to_num(r[27]) if len(r) > 27 else None, 1),
                "rev": _round(_to_num(r[28]) if len(r) > 28 else None, 1)}
        s28  = {"rns": _round(_to_num(r[29]) if len(r) > 29 else None, 1),
                "adr": _round(_to_num(r[30]) if len(r) > 30 else None, 1),
                "rev": _round(_to_num(r[31]) if len(r) > 31 else None, 1)}
        # 달성 KEY 전략 = idx 20 (병합 셀, 텍스트 한 줄)
        strat_text = ""
        if len(r) > 20 and r[20]:
            strat_text = str(r[20]).strip()
        # 사업장 이슈사항 = idx 33 (수도권 호텔수 多 ... 등)
        issue_text = ""
        if len(r) > 33 and r[33]:
            issue_text = str(r[33]).strip()
        by_prop[prop] = {
            "y19_actual":   y19,
            "ly25_actual":  ly25,
            "y26_budget":   b26,
            "y26_stretch":  s26,
            "y27_stretch":  s27,
            "y28_stretch":  s28,
            "strategy_text": strat_text,
            "issue_text":    issue_text,
        }
    return by_prop


# ── 도전목표(3개년) 시트 파서 ────────────────────────────────────────
# iter_rows 인덱스 (idx0=None): 1=구분  2=19Y  3=20  4=21  5=22  6=23Y  7=24Y
#                              8=25Y  9=26Y  10=27Y  11=28Y
# R7=객실수, R9=ADR, R11=객실매출, R13=스포츠매출, R15=오션+아쿠아, R17=식음,
# R19=부대업장, R21=기타, R23=총매출 → list index 6,8,10,...,22
def parse_3yr(ws):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 24:
        return {}

    def trip(row_idx):
        r = rows[row_idx]
        return {
            "y19": _round(_to_num(r[2]),  1) if len(r) >  2 else None,
            "y23": _round(_to_num(r[6]),  1) if len(r) >  6 else None,
            "y24": _round(_to_num(r[7]),  1) if len(r) >  7 else None,
            "y25": _round(_to_num(r[8]),  1) if len(r) >  8 else None,
            "y26": _round(_to_num(r[9]),  1) if len(r) >  9 else None,
            "y27": _round(_to_num(r[10]), 1) if len(r) > 10 else None,
            "y28": _round(_to_num(r[11]), 1) if len(r) > 11 else None,
        }

    return {
        "rns":          trip(6),
        "adr":          trip(8),
        "rev_room":     trip(10),
        "rev_sports":   trip(12),
        "rev_aqua":     trip(14),
        "rev_fnb":      trip(16),
        "rev_facility": trip(18),
        "rev_etc":      trip(20),
        "rev_total":    trip(22),
    }


# ── 결합/머지 ─────────────────────────────────────────────────────────
def merge_grid(stretch_kpi, budget_kpi):
    """stretch (월별 KPI(도전목표)) + budget (월별 KPI) 키로 머지.
    키 = (사업장, 국가, 여행사). budget 시트는 일부 사업장이 '그외 사업장' 으로
    묶여 있어 stretch 와 키가 다를 수 있음 → stretch 기준으로 budget lookup."""
    grid = []
    for key, sd in stretch_kpi.items():
        prop, country, agency = key
        bd = budget_kpi.get(key)
        item = {
            "property": prop,
            "country": country,
            "agency": agency,
            "type": sd["type"],  # 전략 / 일반
            "stretch": {"total": sd["total"], "monthly": sd["monthly"]},
            "budget":  ({"total": bd["total"], "monthly": bd["monthly"]} if bd
                        else {"total": 0, "monthly": [0] * 12}),
        }
        grid.append(item)
    # property order, country, agency 순 정렬
    prop_idx = {p: i for i, p in enumerate(PROPERTIES_ORDER)}
    grid.sort(key=lambda g: (prop_idx.get(g["property"], 99), g["country"], -g["stretch"]["total"]))
    return grid


def derive_strategic_agencies(grid):
    """grid 의 type 컬럼 기반 → 사업장별 전략여행사 리스트 (26년 초기값)."""
    out = {}
    for it in grid:
        if it["type"] != "전략":
            continue
        out.setdefault(it["property"], []).append(it["agency"])
    # dedup, preserve order
    return {p: list(dict.fromkeys(lst)) for p, lst in out.items()}


def main():
    if not SRC.exists():
        raise SystemExit(f"[build_inbound_kpi] 소스 파일 없음: {SRC}")

    print(f"[build_inbound_kpi] 로드: {SRC}")
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

    sheets = {sn for sn in wb.sheetnames}
    required = ["월별 KPI(도전목표)", "월별 KPI", "총괄(사업장별 실적)",
                "총괄(사업장별 전략)", "도전목표(3개년)"]
    missing = [s for s in required if s not in sheets]
    if missing:
        raise SystemExit(f"[build_inbound_kpi] 필수 시트 누락: {missing}")

    print("  · 월별 KPI(도전목표) 파싱…")
    stretch_kpi = parse_monthly_kpi(wb["월별 KPI(도전목표)"])
    print(f"    → {len(stretch_kpi)}개 (사업장×국가×여행사) 행")

    print("  · 월별 KPI(Budget) 파싱…")
    budget_kpi = parse_monthly_kpi(wb["월별 KPI"])
    print(f"    → {len(budget_kpi)}개 행")

    print("  · 총괄(사업장별 실적) 파싱…")
    prop_actual = parse_property_actual(wb["총괄(사업장별 실적)"])
    print(f"    → {len(prop_actual)}개 사업장")

    print("  · 총괄(사업장별 전략) 파싱…")
    prop_strategy = parse_property_strategy(wb["총괄(사업장별 전략)"])
    print(f"    → {len(prop_strategy)}개 사업장")

    print("  · 도전목표(3개년) 파싱…")
    totals_3yr = parse_3yr(wb["도전목표(3개년)"])

    print("  · grid 머지…")
    grid = merge_grid(stretch_kpi, budget_kpi)

    # 사업장 통합 — actual·strategy·monthly 합산
    properties = {}
    for prop in PROPERTIES_ORDER:
        a = prop_actual.get(prop, {})
        s = prop_strategy.get(prop, {})
        properties[prop] = {
            "y19_actual":     s.get("y19_actual",  {}),
            "ly25_actual":    s.get("ly25_actual", {}),
            "y26_budget":     s.get("y26_budget",  {}),
            "y26_stretch":    s.get("y26_stretch", {}),
            "y27_stretch":    s.get("y27_stretch", {}),
            "y28_stretch":    s.get("y28_stretch", {}),
            "strategy_text":  s.get("strategy_text", ""),
            "issue_text":     s.get("issue_text", ""),
            "monthly_rns":    a.get("monthly_rns", []),
            "budget_rns_total":  a.get("budget_rns_total"),
            "stretch_rns_total": a.get("stretch_rns_total"),
            "ly_rns_total":      a.get("ly_rns_total"),
        }

    # totals (월별 KPI 시트 R6 = TOTAL 행 직접 읽음)
    # iter_rows tuple: idx 1=TOTAL, 3=RNs, 4=ADR, 5=REV, 9=계, 10~21=1~12월
    def tot_row(ws, row_num):
        r = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
        return {
            "rns_total": _round(_to_num(r[3]), 1) if len(r) > 3 else None,
            "adr":       _round(_to_num(r[4]), 1) if len(r) > 4 else None,
            "rev":       _round(_to_num(r[5]), 1) if len(r) > 5 else None,
            "monthly":   [_round(_to_num(r[ci]) or 0, 1) for ci in range(10, 22)] if len(r) >= 22 else [],
        }

    totals = {
        "stretch": tot_row(wb["월별 KPI(도전목표)"], 6),
        "budget":  tot_row(wb["월별 KPI"], 6),
        "ly25_rns_total": prop_actual.get("__TOTAL__", {}).get("ly_rns_total"),
    }

    strategic_agencies_2026 = derive_strategic_agencies(grid)

    out = {
        "version": "1.0",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_file": str(SRC.relative_to(ROOT)),
        "year": 2026,
        "_schema": {
            "totals": "회사 전체 RNs/ADR/REV — Budget · 도전목표 · 25년 실적",
            "totals_3yr": "도전목표(3개년) 시트 — 7개 KPI × 19/23/24/25/26/27/28",
            "properties": "사업장별 KPI — 19/25 actual + 26 Budget + 26/27/28 Stretch + 월별 RNs Budget·도전·전년",
            "grid": "사업장×국가×여행사×월 풀그리드 (Budget + 도전목표 + 전략구분)",
            "strategic_agencies_2026": "26년 초기 전략여행사 리스트 — 월별 KPI(도전목표) 시트 H컬럼('전략' 표시) 기반",
        },
        "_units": {
            "RNs": "실 (객실수)",
            "ADR": "천원/박",
            "REV": "백만원 (VAT 제외)",
        },
        "totals":      totals,
        "totals_3yr":  totals_3yr,
        "properties":  properties,
        "grid":        grid,
        "strategic_agencies_2026": strategic_agencies_2026,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    sz = OUT.stat().st_size
    print(f"[build_inbound_kpi] 저장: {OUT}  ({sz:,} bytes)")
    print(f"  · grid rows         : {len(grid)}")
    print(f"  · properties        : {len(properties)}")
    print(f"  · stretch RNs total : {totals['stretch']['rns_total']}")
    print(f"  · budget  RNs total : {totals['budget']['rns_total']}")
    print(f"  · 전략여행사 사업장수: {len(strategic_agencies_2026)}")


if __name__ == "__main__":
    main()
